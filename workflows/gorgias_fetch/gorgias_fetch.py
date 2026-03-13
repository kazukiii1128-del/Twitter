"""
gorgias_fetch.py
Fetches tickets or customer info from the Gorgias REST API and exports to Excel.

Usage:
    python workflows/gorgias_fetch/gorgias_fetch.py --tickets [--status open|closed|all] [--limit N] [--customer-id ID]
    python workflows/gorgias_fetch/gorgias_fetch.py --customer --email someone@example.com
    python workflows/gorgias_fetch/gorgias_fetch.py --customer --id 12345

Output:
    .tmp/gorgias_tickets_{status}_YYYY-MM-DD_HHMMSS.xlsx
    .tmp/gorgias_customer_{identifier}_YYYY-MM-DD_HHMMSS.xlsx
"""

import os
import sys
import argparse
import requests
from requests.auth import HTTPBasicAuth

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font

# --- Config ---

load_dotenv()

DOMAIN  = os.getenv("GORGIAS_DOMAIN")
EMAIL   = os.getenv("GORGIAS_EMAIL")
API_KEY = os.getenv("GORGIAS_API_KEY")

if not DOMAIN:
    raise ValueError("GORGIAS_DOMAIN not found in .env")
if not EMAIL:
    raise ValueError("GORGIAS_EMAIL not found in .env (needed for Basic Auth)")
if not API_KEY:
    raise ValueError("GORGIAS_API_KEY not found in .env")

BASE_URL = f"https://{DOMAIN}.gorgias.com/api"
AUTH     = HTTPBasicAuth(EMAIL, API_KEY)
HEADERS  = {"Accept": "application/json"}


# --- API helper ---

def api_get(endpoint: str, params: dict = None) -> dict:
    """Make a GET request to the Gorgias API. Returns parsed JSON."""
    url = f"{BASE_URL}/{endpoint.lstrip('/')}"
    resp = requests.get(url, auth=AUTH, headers=HEADERS, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


# --- Ticket fetching ---

def normalize_ticket(t: dict) -> dict:
    """Flatten nested Gorgias ticket fields into a simple dict."""
    customer = t.get("customer") or {}
    assignee = t.get("assignee_user") or {}
    tags = t.get("tags") or []
    return {
        "id": t.get("id", ""),
        "status": t.get("status", ""),
        "subject": t.get("subject", ""),
        "channel": t.get("channel", ""),
        "created_datetime": t.get("created_datetime", ""),
        "updated_datetime": t.get("updated_datetime", ""),
        "assignee": assignee.get("name", ""),
        "customer_name": customer.get("name", ""),
        "customer_email": (customer.get("email") or ""),
        "tags": ", ".join(tags) if isinstance(tags, list) else str(tags),
        "messages_count": t.get("messages_count", ""),
        "ticket_url": f"https://{DOMAIN}.gorgias.com/app/tickets/{t.get('id', '')}",
    }


TICKET_COLUMNS = [
    ("Ticket ID",       "id"),
    ("Status",          "status"),
    ("Subject",         "subject"),
    ("Channel",         "channel"),
    ("Created At",      "created_datetime"),
    ("Updated At",      "updated_datetime"),
    ("Assignee",        "assignee"),
    ("Customer Name",   "customer_name"),
    ("Customer Email",  "customer_email"),
    ("Tags",            "tags"),
    ("Message Count",   "messages_count"),
    ("Ticket URL",      "ticket_url"),
]


def fetch_tickets(status: str = "open", limit: int = 100, customer_id: int = None) -> list[dict]:
    """Fetch tickets from Gorgias with optional filters. Returns list of normalized dicts."""
    tickets = []
    page = 1
    per_page = min(limit, 100)  # Gorgias max per-page is 100

    while True:
        params = {"limit": per_page, "page": page}
        if status != "all":
            params["status"] = status
        if customer_id:
            params["customer_id"] = customer_id

        print(f"  Fetching tickets page {page} (limit={per_page}, status={status})...")
        data = api_get("tickets", params)

        batch = data.get("data", [])
        tickets.extend(batch)
        print(f"  Got {len(batch)} tickets (total so far: {len(tickets)})")

        if len(batch) < per_page or len(tickets) >= limit:
            break
        page += 1

    return [normalize_ticket(t) for t in tickets[:limit]]


# --- Customer fetching ---

def normalize_customer(c: dict) -> dict:
    """Flatten nested Gorgias customer fields into a simple dict."""
    tags = c.get("tags") or []
    channels = c.get("channels") or []
    channel_types = ", ".join(ch.get("type", "") for ch in channels if ch.get("type"))
    return {
        "id": c.get("id", ""),
        "name": c.get("name", ""),
        "email": (c.get("email") or ""),
        "phone": c.get("phone", "") or "",
        "created_datetime": c.get("created_datetime", ""),
        "updated_datetime": c.get("updated_datetime", ""),
        "external_id": c.get("external_id", "") or "",
        "note": c.get("note", "") or "",
        "tags": ", ".join(tags) if isinstance(tags, list) else str(tags),
        "channels": channel_types,
    }


CUSTOMER_COLUMNS = [
    ("Customer ID",  "id"),
    ("Name",         "name"),
    ("Email",        "email"),
    ("Phone",        "phone"),
    ("Created At",   "created_datetime"),
    ("Updated At",   "updated_datetime"),
    ("External ID",  "external_id"),
    ("Note",         "note"),
    ("Tags",         "tags"),
    ("Channels",     "channels"),
]


def fetch_customer_by_id(customer_id: int) -> dict | None:
    """Fetch a single customer by numeric ID."""
    print(f"  Fetching customer ID {customer_id}...")
    data = api_get(f"customers/{customer_id}")
    return normalize_customer(data)


def fetch_customer_by_email(email: str) -> dict | None:
    """Search for a customer by email address. Returns first match or None."""
    print(f"  Searching customer by email: {email}...")
    data = api_get("customers", params={"email": email})
    results = data.get("data", [])
    if not results:
        return None
    return normalize_customer(results[0])


# --- Excel writer ---

def write_excel(records: list[dict], columns: list[tuple], sheet_title: str, output_path: str):
    """Write records to an Excel file with bold headers and auto-width columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Header row
    headers = [col[0] for col in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for record in records:
        row = [record.get(col[1], "") for col in columns]
        ws.append(row)

    # Auto-width (approximate)
    for col_idx, (header, _) in enumerate(columns, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[col_letter].width = max_len + 2

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# --- Main ---

def main():
    parser = argparse.ArgumentParser(description="Fetch Gorgias tickets or customer info")
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument("--tickets",  action="store_true", help="Fetch ticket list")
    mode_group.add_argument("--customer", action="store_true", help="Fetch customer info")

    parser.add_argument("--status",      default="open", choices=["open", "closed", "all"],
                        help="Ticket status filter (default: open)")
    parser.add_argument("--limit",       type=int, default=100,
                        help="Max number of tickets to fetch (default: 100)")
    parser.add_argument("--customer-id", type=int, dest="customer_id",
                        help="Filter tickets by customer ID (tickets mode) or lookup customer (customer mode)")
    parser.add_argument("--email",       help="Customer lookup by email (customer mode)")
    parser.add_argument("--id",          type=int, dest="id",
                        help="Customer lookup by numeric ID (customer mode)")

    args = parser.parse_args()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")

    try:
        if args.tickets:
            print(f"=== Step 1: Fetching tickets (status={args.status}, limit={args.limit}) ===")
            tickets = fetch_tickets(
                status=args.status,
                limit=args.limit,
                customer_id=args.customer_id,
            )
            print(f"\nTotal tickets fetched: {len(tickets)}")

            if not tickets:
                print("No tickets found. Exiting.")
                return

            output_path = os.path.join(".tmp", "gorgias_fetch", f"gorgias_tickets_{args.status}_{timestamp}.xlsx")
            print(f"\n=== Step 2: Writing Excel to {output_path} ===")
            write_excel(tickets, TICKET_COLUMNS, "Tickets", output_path)
            print(f"Done. {len(tickets)} tickets saved to {output_path}")

        elif args.customer:
            if args.id:
                print(f"=== Step 1: Fetching customer (ID={args.id}) ===")
                customer = fetch_customer_by_id(args.id)
                identifier = str(args.id)
            elif args.email:
                print(f"=== Step 1: Fetching customer (email={args.email}) ===")
                customer = fetch_customer_by_email(args.email)
                identifier = args.email.replace("@", "_at_")
            else:
                parser.error("--customer requires --email or --id")
                return

            if not customer:
                print("Customer not found. Exiting.")
                return

            output_path = os.path.join(".tmp", "gorgias_fetch", f"gorgias_customer_{identifier}_{timestamp}.xlsx")
            print(f"\n=== Step 2: Writing Excel to {output_path} ===")
            write_excel([customer], CUSTOMER_COLUMNS, "Customer", output_path)
            print(f"Done. Customer saved to {output_path}")

    except requests.exceptions.HTTPError as e:
        print(f"ERROR: API returned {e.response.status_code}: {e.response.text}")
        sys.exit(1)
    except requests.exceptions.ConnectionError:
        print(f"ERROR: Could not connect to {BASE_URL}. Check GORGIAS_DOMAIN in .env.")
        sys.exit(1)
    except requests.exceptions.Timeout:
        print("ERROR: Request timed out. Check network or increase timeout.")
        sys.exit(1)


if __name__ == "__main__":
    main()
