"""
gorgias_analyze.py
Finds the top N Gorgias tickets with the most message exchanges,
then extracts customer name, order number, date history, conversation,
and complaint keywords (via Claude AI) into an Excel report.

Usage:
    python workflows/gorgias_analyze/gorgias_analyze.py [--min-messages N] [--top N] [--status all|open|closed]

Defaults:
    --min-messages 3   (tickets with at least 3 messages)
    --top 5            (pick the 5 most active)
    --status all

Output:
    .tmp/gorgias_analysis_YYYY-MM-DD_HHMMSS.xlsx
"""

import os
import re
import sys
import argparse
import requests
from requests.auth import HTTPBasicAuth

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment
import anthropic

# --- Config ---

load_dotenv()

DOMAIN        = os.getenv("GORGIAS_DOMAIN")
EMAIL         = os.getenv("GORGIAS_EMAIL")
API_KEY       = os.getenv("GORGIAS_API_KEY")
ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY")

if not DOMAIN:
    raise ValueError("GORGIAS_DOMAIN not found in .env")
if not EMAIL:
    raise ValueError("GORGIAS_EMAIL not found in .env (needed for Basic Auth)")
if not API_KEY:
    raise ValueError("GORGIAS_API_KEY not found in .env")
BASE_URL     = f"https://{DOMAIN}.gorgias.com/api"
AUTH         = HTTPBasicAuth(EMAIL, API_KEY)
HEADERS      = {"Accept": "application/json"}
AI_CLIENT    = anthropic.Anthropic(api_key=ANTHROPIC_KEY) if ANTHROPIC_KEY else None

if not ANTHROPIC_KEY:
    print("INFO: ANTHROPIC_API_KEY not set — keyword extraction will be skipped.")

# Regex patterns for order number extraction from message body
ORDER_RE = re.compile(
    r'(?:#\s*(\d{4,}))|(?:order(?:\s+(?:no|num|number|id|#))?[:\s#]+(\d{4,}))|(?:ORD[-\s]?(\d{4,}))',
    re.IGNORECASE
)


# --- API helper ---

def api_get(endpoint: str, params: dict = None) -> dict:
    """Make a GET request to the Gorgias API. Returns parsed JSON."""
    url = f"{BASE_URL}/{endpoint.lstrip('/')}"
    resp = requests.get(url, auth=AUTH, headers=HEADERS, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


# --- Step 1: Fetch and filter tickets ---

def fetch_all_tickets(status: str = "all") -> list[dict]:
    """Fetch all tickets via cursor-based pagination. Returns raw ticket dicts."""
    all_tickets = []
    per_page = 100
    cursor = None
    page_num = 1

    while True:
        params = {"limit": per_page, "order_by": "created_datetime:desc"}
        if cursor:
            params["cursor"] = cursor

        print(f"  Fetching ticket list page {page_num}...")
        data = api_get("tickets", params)
        batch = data.get("data", [])
        all_tickets.extend(batch)
        print(f"  Got {len(batch)} tickets (total: {len(all_tickets)})")

        # Gorgias cursor pagination: get next cursor from meta
        meta = data.get("meta") or {}
        next_cursor = meta.get("next_cursor")
        if not batch or not next_cursor:
            break
        cursor = next_cursor
        page_num += 1

    return all_tickets


def select_top_tickets(tickets: list[dict], min_messages: int, top: int, status: str = "all") -> list[dict]:
    """Filter by status (client-side), min message count, sort by message count desc, return top N."""
    filtered = tickets
    if status != "all":
        filtered = [t for t in filtered if t.get("status") == status]
    filtered = [t for t in filtered if (t.get("messages_count") or 0) >= min_messages]
    filtered.sort(key=lambda t: t.get("messages_count") or 0, reverse=True)
    return filtered[:top]


# --- Step 2: Fetch ticket details + messages ---

def fetch_ticket_detail(ticket_id: int) -> dict:
    """Fetch full ticket details including custom_fields."""
    return api_get(f"tickets/{ticket_id}")


def fetch_ticket_messages(ticket_id: int) -> list[dict]:
    """Fetch all messages for a ticket via cursor pagination."""
    all_messages = []
    cursor = None

    while True:
        params = {"limit": 30}  # Gorgias messages API max per page is 30
        if cursor:
            params["cursor"] = cursor

        data = api_get(f"tickets/{ticket_id}/messages", params=params)
        batch = data.get("data", [])
        all_messages.extend(batch)

        meta = data.get("meta") or {}
        next_cursor = meta.get("next_cursor")
        if not batch or not next_cursor:
            break
        cursor = next_cursor

    return all_messages


# --- Step 3: Extract order number ---

def extract_order_number(ticket_detail: dict, messages: list[dict]) -> str:
    """
    Try to find an order number:
    1. From ticket custom_fields where name contains 'order'
    2. From message body text via regex
    Returns the first match found, or empty string.
    """
    # 1. Custom fields
    custom_fields = ticket_detail.get("custom_fields") or []
    for field in custom_fields:
        name = (field.get("name") or "").lower()
        if "order" in name:
            value = field.get("value")
            if value:
                return str(value)

    # 2. Message body regex
    for msg in messages:
        body = msg.get("body_text") or ""
        match = ORDER_RE.search(body)
        if match:
            # Return whichever capture group matched
            return next(g for g in match.groups() if g is not None)

    return ""


# --- Step 4: Format conversation and date log ---

def format_conversation(messages: list[dict]) -> tuple[str, str]:
    """
    Returns (date_log, conversation_text).
    date_log: one line per message with [date] sender: short preview
    conversation_text: full thread with sender labels
    """
    date_lines = []
    conv_lines = []

    for msg in messages:
        created = msg.get("created_datetime") or ""
        # Trim to readable format: 2026-01-05T14:22:31 -> 2026-01-05 14:22
        try:
            dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
            ts = dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            ts = created[:16] if len(created) >= 16 else created

        # Determine sender label
        from_agent = msg.get("from_agent", False)
        sender_obj = msg.get("sender") or {}
        sender_type = sender_obj.get("type") or ""

        if from_agent or sender_type == "agent":
            label = "상담원"
        else:
            label = "고객"

        body = (msg.get("body_text") or "").strip()
        # Remove excessive whitespace
        body = re.sub(r'\n{3,}', '\n\n', body)
        # Short preview for date log (first 60 chars)
        preview = body[:60].replace('\n', ' ')
        if len(body) > 60:
            preview += "..."

        date_lines.append(f"[{ts}] {label}: {preview}")
        conv_lines.append(f"[{ts}] {label}:\n{body}")

    return "\n".join(date_lines), "\n\n---\n\n".join(conv_lines)


# --- Step 5: Extract complaint keywords via Claude ---

def extract_keywords(conversation_text: str, ticket_subject: str) -> str:
    """
    Use Claude Haiku to extract main complaint keywords from a conversation.
    Returns comma-separated keywords, or empty string on failure.
    """
    # Truncate conversation to keep token cost low (max ~3000 chars)
    truncated = conversation_text[:3000]
    if len(conversation_text) > 3000:
        truncated += "\n...(이하 생략)"

    prompt = (
        f"아래는 고객 지원 이메일 대화입니다. 제목: {ticket_subject}\n\n"
        f"{truncated}\n\n"
        "위 대화에서 고객의 주요 불만/컴플레인 키워드를 5개 이내로 추출해줘. "
        "한국어 단어나 짧은 구문으로, 쉼표로 구분해서 답해줘. 키워드만 출력해."
    )

    if not AI_CLIENT:
        return "(ANTHROPIC_API_KEY 없음 — 키워드 추출 생략)"

    try:
        resp = AI_CLIENT.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=150,
            messages=[{"role": "user", "content": prompt}]
        )
        return resp.content[0].text.strip()
    except Exception as e:
        print(f"  WARN: Claude keyword extraction failed: {e}")
        return ""


# --- Step 6: Excel writer ---

COLUMNS = [
    ("상대 이름",            "customer_name"),
    ("계정 오더넘버",         "order_number"),
    ("이메일 회신 날짜별 기록", "date_log"),
    ("대화 내용",             "conversation"),
    ("주요 컴플레인 키워드",   "keywords"),
]


def write_excel(records: list[dict], output_path: str):
    """Write analysis records to Excel with wrap_text for multi-line cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "분석 결과"

    # Header row
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths (fixed — content is too variable for auto-width)
    col_widths = [20, 15, 35, 60, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Data rows
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for record in records:
        row_data = [record.get(col[1], "") for col in COLUMNS]
        ws.append(row_data)
        row_num = ws.max_row
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_num, column=col_idx).alignment = wrap_align
        # Set row height to give multi-line cells room
        ws.row_dimensions[row_num].height = 120

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# --- Main ---

def main():
    parser = argparse.ArgumentParser(
        description="Analyze top Gorgias tickets with most message exchanges"
    )
    parser.add_argument("--min-messages", type=int, default=3, dest="min_messages",
                        help="Minimum message count to qualify (default: 3)")
    parser.add_argument("--top",          type=int, default=5,
                        help="Number of top tickets to analyze (default: 5)")
    parser.add_argument("--status",       default="all", choices=["all", "open", "closed"],
                        help="Ticket status filter (default: all)")
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_path = os.path.join(".tmp", "gorgias_analyze", f"gorgias_analysis_{timestamp}.xlsx")

    try:
        # Step 1: Fetch and filter
        print(f"=== Step 1: Fetching all tickets (status={args.status}) ===")
        all_tickets = fetch_all_tickets(status=args.status)
        print(f"\nTotal tickets fetched: {len(all_tickets)}")

        candidates = select_top_tickets(all_tickets, args.min_messages, args.top, args.status)
        status_count = len([t for t in all_tickets if args.status == "all" or t.get("status") == args.status])
        qualify_count = len([t for t in all_tickets
                             if (args.status == "all" or t.get("status") == args.status)
                             and (t.get("messages_count") or 0) >= args.min_messages])
        print(f"Status={args.status} tickets: {status_count}")
        print(f"Tickets with >= {args.min_messages} messages: {qualify_count}")
        print(f"Selected top {len(candidates)} for analysis\n")

        if not candidates:
            print(f"No tickets found with >= {args.min_messages} messages. Exiting.")
            return

        # Step 2-5: Analyze each ticket
        print(f"=== Step 2: Analyzing {len(candidates)} tickets ===")
        records = []

        for i, ticket in enumerate(candidates, start=1):
            tid = ticket.get("id")
            subject = ticket.get("subject") or ""
            customer = ticket.get("customer") or {}
            customer_name = customer.get("name") or customer.get("email") or str(tid)
            msg_count = ticket.get("messages_count") or 0

            print(f"\n  [{i}/{len(candidates)}] Ticket #{tid} — {subject[:60]} ({msg_count} messages)")

            # Fetch detail + messages
            detail = fetch_ticket_detail(tid)
            messages = fetch_ticket_messages(tid)
            print(f"    Messages fetched: {len(messages)}")

            # Order number
            order_number = extract_order_number(detail, messages)
            print(f"    Order number: {order_number or '(not found)'}")

            # Format conversation
            date_log, conversation = format_conversation(messages)

            # Claude keyword extraction
            print(f"    Extracting complaint keywords via Claude...")
            keywords = extract_keywords(conversation, subject)
            print(f"    Keywords: {keywords or '(none extracted)'}")

            records.append({
                "customer_name": customer_name,
                "order_number":  order_number,
                "date_log":      date_log,
                "conversation":  conversation,
                "keywords":      keywords,
            })

        # Step 3: Write Excel
        print(f"\n=== Step 3: Writing Excel to {output_path} ===")
        write_excel(records, output_path)
        print(f"Done. {len(records)} tickets analyzed and saved to {output_path}")

    except requests.exceptions.HTTPError as e:
        print(f"ERROR: API returned {e.response.status_code}: {e.response.text}")
        sys.exit(1)
    except requests.exceptions.ConnectionError:
        print(f"ERROR: Could not connect to {BASE_URL}. Check GORGIAS_DOMAIN in .env.")
        sys.exit(1)
    except requests.exceptions.Timeout:
        print("ERROR: Request timed out. Check network or increase timeout.")
        sys.exit(1)


if __name__ == "__main__":
    main()
