"""
sync_influencer_notion.py

Reads influencer data from Google Sheets (read-only) and syncs key fields
to a Notion database. Supports both Master DB tab (personal info) and
brand-specific tabs (campaign data).

Output:
  - Notion Database (influencer campaign records)
  - .tmp/sync_influencer_notion/sync_report_YYYY-MM-DD.xlsx (local report)

Usage:
    python tools/sync_influencer_notion.py --discover
    python tools/sync_influencer_notion.py --dry-run
    python tools/sync_influencer_notion.py --sync
    python tools/sync_influencer_notion.py --sync --master-only

Prerequisites:
    - .env: NOTION_API_TOKEN, INFLUENCER_SHEET_ID, INFLUENCER_NOTION_DB_ID
    - .env: GOOGLE_SERVICE_ACCOUNT_PATH=credentials/google_service_account.json
    - Service Account email shared as viewer on the Google Sheet
    - Notion Integration shared with the target database
"""

import os
import sys
import argparse
import time
import re
import csv
import io
from datetime import datetime
from collections import defaultdict

import requests
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment

# Force UTF-8 output on Windows
if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

load_dotenv()

NOTION_TOKEN = os.getenv("NOTION_API_TOKEN")
NOTION_DB_ID = os.getenv("INFLUENCER_NOTION_DB_ID", "abb8fcc1be0041c598bbe7635413091c")
SHEET_ID = os.getenv("INFLUENCER_SHEET_ID", "1DPI_zxG6XiCliyi7Vw6YY_ojue4ZYor-vDX6EN7nUOY")
MASTER_GID = os.getenv("INFLUENCER_SHEET_GID", "1592924077")
SA_PATH = os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH", "credentials/google_service_account.json")

if not NOTION_TOKEN:
    raise ValueError("NOTION_API_TOKEN not found in .env")

NOTION_HEADERS = {
    "Authorization": f"Bearer {NOTION_TOKEN}",
    "Notion-Version": "2022-06-28",
    "Content-Type": "application/json",
}

# Output directory
from output_utils import get_output_path

# Brand name extraction from tab names (e.g., "Grosmimi Cp" -> "Grosmimi")
BRAND_TAB_MAP = {
    "Grosmimi": "Grosmimi",
    "CHA&MOM": "CHA&MOM",
    "Baby Rabbit": "Baby Rabbit",
    "BabyRabbit": "Baby Rabbit",
    "Naeiae": "Naeiae",
    "Alpremio": "Alpremio",
    "Commemoi": "Commemoi",
    "Beemymagic": "Beemymagic",
    "Onzenna": "Onzenna",
}

# Tabs to skip (aggregate / shared tabs, not individual brand campaigns)
SKIP_TABS = {"JS(All)", "(All) 공유용"}

# Stage mapping: Google Sheets Status -> Notion Stage
STAGE_MAP = {
    "outreach": "Outreach",
    "outreached": "Outreach",
    "negotiating": "Negotiating",
    "soft confirmed": "Soft Confirmed",
    "confirmed": "Soft Confirmed",
    "contracted": "Contracted",
    "content planning": "Content Planning",
    "content reviewing": "Content Reviewing",
    "posted": "Posted",
    "campaign done": "Posted",
    "completed": "Posted",
}

# Collaboration count mapping
COLLAB_COUNT_MAP = {
    1: "1st time",
    2: "2nd time",
    3: "3rd time",
}


# ---------------------------------------------------------------------------
# Notion API helpers
# ---------------------------------------------------------------------------


def notion_api(method, endpoint, json_body=None, max_retries=3):
    """Make a Notion API call with retry logic."""
    url = f"https://api.notion.com/v1/{endpoint}"
    for attempt in range(max_retries):
        try:
            if method == "GET":
                resp = requests.get(url, headers=NOTION_HEADERS, timeout=30)
            elif method == "POST":
                resp = requests.post(url, headers=NOTION_HEADERS, json=json_body, timeout=30)
            elif method == "PATCH":
                resp = requests.patch(url, headers=NOTION_HEADERS, json=json_body, timeout=30)
            else:
                raise ValueError(f"Unsupported method: {method}")

            if resp.status_code == 429:
                retry_after = float(resp.headers.get("Retry-After", 1))
                print(f"    Rate limited. Waiting {retry_after}s...")
                time.sleep(retry_after)
                continue

            if resp.status_code == 409:
                print(f"    Conflict (409). Retrying in 1s...")
                time.sleep(1)
                continue

            resp.raise_for_status()
            return resp.json()

        except requests.exceptions.RequestException as e:
            if attempt < max_retries - 1:
                wait = 2 ** attempt
                print(f"    API error (attempt {attempt + 1}): {e}. Retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise
    return None


def read_notion_db_schema():
    """Retrieve Notion database properties."""
    data = notion_api("GET", f"databases/{NOTION_DB_ID}")
    if not data:
        return {}
    props = {}
    for name, config in data.get("properties", {}).items():
        prop_info = {"type": config.get("type"), "id": config.get("id")}
        if config.get("type") == "select":
            prop_info["options"] = [o["name"] for o in config.get("select", {}).get("options", [])]
        elif config.get("type") == "multi_select":
            prop_info["options"] = [o["name"] for o in config.get("multi_select", {}).get("options", [])]
        elif config.get("type") == "status":
            prop_info["options"] = [o["name"] for o in config.get("status", {}).get("options", [])]
        props[name] = prop_info
    return props


def read_notion_pages():
    """Query all pages from Notion database. Returns list of page objects."""
    all_pages = []
    start_cursor = None
    while True:
        body = {"page_size": 100}
        if start_cursor:
            body["start_cursor"] = start_cursor
        data = notion_api("POST", f"databases/{NOTION_DB_ID}/query", body)
        if not data:
            break
        all_pages.extend(data.get("results", []))
        if not data.get("has_more"):
            break
        start_cursor = data.get("next_cursor")
    return all_pages


def read_notion_users():
    """List workspace users for Owner (people) field mapping."""
    users = {}
    start_cursor = None
    while True:
        endpoint = "users"
        if start_cursor:
            endpoint += f"?start_cursor={start_cursor}"
        data = notion_api("GET", endpoint)
        if not data:
            break
        for user in data.get("results", []):
            if user.get("type") == "person":
                name = user.get("name", "")
                users[name.lower()] = user["id"]
        if not data.get("has_more"):
            break
        start_cursor = data.get("next_cursor")
    return users


def extract_notion_page_data(page):
    """Extract key fields from a Notion page object into a flat dict."""
    props = page.get("properties", {})
    result = {"_page_id": page["id"]}

    for name, prop in props.items():
        ptype = prop.get("type", "")
        val = None
        if ptype == "title":
            val = "".join([t.get("plain_text", "") for t in prop.get("title", [])])
        elif ptype == "rich_text":
            val = "".join([t.get("plain_text", "") for t in prop.get("rich_text", [])])
        elif ptype == "number":
            val = prop.get("number")
        elif ptype == "select":
            sel = prop.get("select")
            val = sel.get("name", "") if sel else ""
        elif ptype == "multi_select":
            val = [s.get("name", "") for s in prop.get("multi_select", [])]
        elif ptype == "url":
            val = prop.get("url")
        elif ptype == "date":
            d = prop.get("date")
            val = d.get("start", "") if d else ""
        elif ptype == "status":
            s = prop.get("status")
            val = s.get("name", "") if s else ""
        elif ptype == "checkbox":
            val = prop.get("checkbox")
        elif ptype == "people":
            val = [p.get("name", "") for p in prop.get("people", [])]
        elif ptype == "email":
            val = prop.get("email")
        result[name] = val

    return result


# ---------------------------------------------------------------------------
# Google Sheets helpers
# ---------------------------------------------------------------------------


def _build_prefixed_headers(cat_row, col_row):
    """Build headers with category prefixes to handle duplicate column names.
    E.g., '@ID (link)' appears under TikTok, Instagram, YouTube ->
    'TikTok:@ID (link)', 'Instagram:@ID (link)', 'YouTube:@ID (link)'
    """
    headers = []
    seen = {}
    current_cat = ""

    for i in range(len(col_row)):
        if i < len(cat_row) and cat_row[i].strip():
            current_cat = cat_row[i].strip().replace("\n", " ")
        col_name = col_row[i].strip().replace("\n", " ")

        if not col_name:
            headers.append("")
            continue

        # Check if this column name already appeared
        if col_name in seen:
            # Prefix with category
            prefixed = f"{current_cat}:{col_name}" if current_cat else col_name
            headers.append(prefixed)
        else:
            headers.append(col_name)
        seen[col_name] = seen.get(col_name, 0) + 1

    return headers


def get_gspread_client():
    """Initialize gspread client with service account."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("ERROR: gspread or google-auth not installed. Run: pip install gspread")
        return None

    if not os.path.exists(SA_PATH):
        print(f"ERROR: Service Account JSON not found at: {SA_PATH}")
        print("  Set up a Google Service Account and save the JSON file.")
        print("  Then set GOOGLE_SERVICE_ACCOUNT_PATH in .env")
        return None

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(SA_PATH, scopes=scopes)
    return gspread.authorize(creds)


def read_master_db_gspread(gc):
    """Read Master DB tab using gspread. Returns (headers, rows_as_dicts)."""
    sh = gc.open_by_key(SHEET_ID)

    # Find Master DB tab by GID
    target_ws = None
    for ws in sh.worksheets():
        if str(ws.id) == MASTER_GID:
            target_ws = ws
            break

    if not target_ws:
        print(f"  WARNING: Could not find Master DB tab with GID {MASTER_GID}")
        return [], []

    print(f"  Reading Master DB tab: '{target_ws.title}' ({target_ws.row_count} rows)")
    all_values = target_ws.get_all_values()

    if len(all_values) < 3:
        print("  WARNING: Master DB tab has fewer than 3 rows")
        return [], []

    # Row 1 = category headers (merged), Row 2 = column names
    cat_row = all_values[0]
    col_row = all_values[1]

    # Build combined headers with category prefixes for duplicates
    headers = _build_prefixed_headers(cat_row, col_row)

    # Parse data rows (from row 3 onwards)
    rows = []
    for row_idx in range(2, len(all_values)):
        row = all_values[row_idx]
        row_dict = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = val.strip() if val else ""
        if row_dict.get("Creator ID", "").strip():  # Skip empty rows
            rows.append(row_dict)

    return headers, rows


def read_master_db_csv():
    """Fallback: Read Master DB via CSV export (no auth needed if sheet is shared)."""
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={MASTER_GID}"
    resp = requests.get(url, timeout=30)
    if resp.status_code != 200 or "text/csv" not in resp.headers.get("content-type", ""):
        print("  ERROR: Cannot access Google Sheet via CSV export.")
        return [], []

    text = resp.content.decode("utf-8")
    reader = csv.reader(io.StringIO(text))
    all_values = list(reader)

    if len(all_values) < 3:
        return [], []

    cat_row = all_values[0]
    col_row = all_values[1]

    headers = _build_prefixed_headers(cat_row, col_row)

    rows = []
    for row_idx in range(2, len(all_values)):
        row = all_values[row_idx]
        row_dict = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = val.strip() if val else ""
        if row_dict.get("Creator ID", "").strip():
            rows.append(row_dict)

    return headers, rows


def read_brand_tabs(gc):
    """Read all brand-specific tabs. Returns list of (brand_name, headers, rows)."""
    sh = gc.open_by_key(SHEET_ID)
    brand_tabs = []

    for ws in sh.worksheets():
        if str(ws.id) == MASTER_GID:
            continue  # Skip Master DB

        tab_name = ws.title.strip()

        # Skip aggregate/shared tabs
        if tab_name in SKIP_TABS:
            print(f"  Skipping tab '{tab_name}' (aggregate/shared)")
            continue

        # Detect brand from tab name
        brand = None
        for brand_key in BRAND_TAB_MAP:
            if brand_key.lower() in tab_name.lower():
                brand = BRAND_TAB_MAP[brand_key]
                break

        if not brand:
            print(f"  Skipping tab '{tab_name}' (no matching brand)")
            continue

        print(f"  Reading brand tab: '{tab_name}' -> Brand: {brand}")
        try:
            all_values = ws.get_all_values()
        except Exception as e:
            print(f"    ERROR reading tab '{tab_name}': {e}")
            continue

        if len(all_values) < 2:
            print(f"    WARNING: Tab '{tab_name}' has fewer than 2 rows, skipping")
            continue

        # Find header row (first row with recognizable column names)
        header_row_idx = 0
        for idx, row in enumerate(all_values[:5]):
            row_text = " ".join(row).lower()
            if "creator" in row_text or "flight" in row_text or "platform" in row_text:
                header_row_idx = idx
                break

        headers = [h.strip().replace("\n", " ") for h in all_values[header_row_idx]]

        rows = []
        for row_idx in range(header_row_idx + 1, len(all_values)):
            row = all_values[row_idx]
            row_dict = {"_brand": brand, "_tab": tab_name}
            has_data = False
            for col_idx, val in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    cleaned = val.strip() if val else ""
                    row_dict[headers[col_idx]] = cleaned
                    if cleaned:
                        has_data = True
            if has_data:
                rows.append(row_dict)

        brand_tabs.append((brand, headers, rows))
        print(f"    Found {len(rows)} rows")

    return brand_tabs


# ---------------------------------------------------------------------------
# Data merging & mapping
# ---------------------------------------------------------------------------


def find_column(row_dict, candidates):
    """Find the first matching column from a list of candidate names."""
    for c in candidates:
        for key in row_dict:
            if key.lower().strip() == c.lower().strip():
                return row_dict[key]
    return ""


def parse_date(date_str):
    """Parse various date formats to YYYY-MM-DD."""
    if not date_str or not date_str.strip():
        return None
    date_str = date_str.strip()

    formats = [
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%d/%m/%Y",
        "%B %d, %Y",
        "%b %d, %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_fee(fee_str):
    """Extract numeric amount from fee string like '$150', '$2,000/reel'."""
    if not fee_str:
        return None
    # Remove $ and commas, take first number
    match = re.search(r"\$?([\d,]+(?:\.\d+)?)", fee_str.replace(",", ""))
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            pass
    return None


def earlier_date(date1, date2):
    """Return the earlier of two YYYY-MM-DD date strings."""
    if not date1:
        return date2
    if not date2:
        return date1
    return min(date1, date2)


def merge_data(master_rows, brand_tabs_data):
    """
    Merge Master DB and brand tab data.
    Aggregates all campaigns per Creator ID into one record.
    Multi-select fields (Brand, Platform, Product) accumulate across campaigns.
    Single-value fields (Fee, Stage, Content Link) use the latest campaign.
    """
    # Build Master DB lookup by Creator ID
    master_lookup = {}
    for row in master_rows:
        cid = row.get("Creator ID", "").strip()
        if cid:
            master_lookup[cid] = row

    # Also build Master DB lookup by social handle (for brand tab matching)
    master_by_handle = {}
    for row in master_rows:
        for key in ["Instagram:@ID (link)", "@ID (link)", "Tiktok:@ID (link)", "YouTube:@ID (link)"]:
            val = row.get(key, "")
            if val and val.strip():
                handle = val.strip().lstrip("@").lower()
                master_by_handle[handle] = row
                break

    # Group brand tab rows by Creator ID (which is the handle in brand tabs)
    creator_campaigns = defaultdict(list)
    for brand, headers, rows in brand_tabs_data:
        for row in rows:
            cid = find_column(row, ["Creator ID", "Creator", "ID"])
            if cid:
                creator_campaigns[cid].append(row)

    # Build merged records (one per creator)
    merged = []
    for cid, campaigns in creator_campaigns.items():
        # Try matching master by: Internal ID (C-xxx), Creator ID (handle), or handle lookup
        internal_id = find_column(campaigns[0], ["Internal ID"])
        master = master_lookup.get(internal_id) or master_lookup.get(cid) or master_by_handle.get(cid.lower(), {})

        # Accumulate multi-select fields across all campaigns
        all_brands = set()
        all_platforms = set()
        all_products = set()
        all_fees = []
        earliest_flight = None
        latest_posted_date = None
        latest_content_link = ""
        latest_stage = ""
        latest_pic = ""

        for row in campaigns:
            # Brand
            brand_val = row.get("_brand", "")
            if brand_val:
                all_brands.add(brand_val)

            # Platform
            platform = find_column(row, ["Platform", "Platforms"])
            if platform:
                for p in platform.split(","):
                    if p.strip():
                        all_platforms.add(p.strip())

            # Deliverable Type -> add "Story"
            deliverable = find_column(row, ["Deliverable Type", "Deliverable", "Type"])
            if deliverable and "story" in deliverable.lower():
                all_platforms.add("Story")

            # Product
            product = find_column(row, ["Product", "Products"])
            if product:
                for p in product.split(","):
                    if p.strip():
                        all_products.add(p.strip())

            # Fee
            fee = find_column(row, ["Fee", "Rate", "Payment"])
            parsed_fee = parse_fee(fee)
            if parsed_fee:
                all_fees.append(parsed_fee)

            # Flight Period -> track earliest
            flight = parse_date(find_column(row, ["Flight Period", "Flight", "Period"]))
            if flight:
                earliest_flight = earlier_date(earliest_flight, flight)

            # Actual Upload -> track latest for posted date, content link, stage
            actual_upload = find_column(row, ["Actual Upload", "Upload Date", "Actual Upload Date"])
            posted = parse_date(actual_upload)
            if posted:
                if not latest_posted_date or posted > latest_posted_date:
                    latest_posted_date = posted
                    latest_content_link = find_column(row, ["Content 1", "Content Link", "Content URL"])
                    status = find_column(row, ["Status", "Campaign Status"])
                    latest_stage = STAGE_MAP.get(status.lower().strip(), "") if status else ""
                    latest_pic = find_column(row, ["PIC", "Person In Charge"])

        # Fallback: use last campaign row for stage/content/pic if no dated campaigns
        if not latest_stage:
            status = find_column(campaigns[-1], ["Status", "Campaign Status"])
            latest_stage = STAGE_MAP.get(status.lower().strip(), "") if status else ""
        if not latest_content_link:
            latest_content_link = find_column(campaigns[-1], ["Content 1", "Content Link", "Content URL"])
        if not latest_pic:
            latest_pic = find_column(campaigns[-1], ["PIC", "Person In Charge"])

        # Outreach Date: earlier of Date Discovered and earliest Flight Period
        date_discovered = parse_date(master.get("Date Discovered", ""))
        outreach = earlier_date(date_discovered, earliest_flight)

        # Creator handle (from Master DB social IDs)
        handle = ""
        for key in ["Instagram:@ID (link)", "@ID (link)", "Tiktok:@ID (link)", "YouTube:@ID (link)"]:
            val = master.get(key, "")
            if val and val.strip():
                handle = val.strip().lstrip("@")
                break
        if not handle:
            handle = cid

        # Collaboration Count
        count = len(campaigns)
        if count >= 4:
            collab = "4th+ time"
        else:
            collab = COLLAB_COUNT_MAP.get(count, "1st time")

        record = {
            "creator_id": cid,
            "full_name": master.get("Full Name", ""),
            "handle": handle,
            "brand": sorted(all_brands),
            "outreach_date": outreach,
            "platform": sorted(all_platforms),
            "product": sorted(all_products),
            "pic": latest_pic,
            "stage": latest_stage,
            "posted_date": latest_posted_date,
            "content_link": latest_content_link,
            "fee": sum(all_fees) if all_fees else None,
            "collab_count": collab,
            "notes": master.get("Recent Rate", ""),
        }

        merged.append(record)

    return merged


def merge_master_only(master_rows):
    """Create records from Master DB only (no brand tab data)."""
    merged = []
    for row in master_rows:
        cid = row.get("Creator ID", "").strip()
        if not cid:
            continue

        full_name = row.get("Full Name", "")
        handle = ""
        # Try to get a social handle (prefer Instagram, then TikTok, then YouTube)
        for key in ["Instagram:@ID (link)", "@ID (link)", "Tiktok:@ID (link)", "YouTube:@ID (link)"]:
            val = row.get(key, "")
            if val and val.strip():
                handle = val.strip().lstrip("@")
                break
        if not handle:
            handle = cid

        record = {
            "creator_id": cid,
            "full_name": full_name,
            "handle": handle,
            "brand": "",
            "outreach_date": parse_date(row.get("Date Discovered", "")),
            "platform": [],
            "product": [],
            "pic": row.get("PIC", ""),
            "stage": "",
            "posted_date": None,
            "content_link": "",
            "fee": None,
            "collab_count": "1st time",
            "notes": row.get("Recent Rate", ""),
        }

        # Detect platforms from available social IDs
        if row.get("Tiktok:@ID (link)", "").strip() or row.get("@ID (link)", "").strip():
            record["platform"].append("TikTok")
        if row.get("Instagram:@ID (link)", "").strip():
            record["platform"].append("Instagram")
        if row.get("YouTube:@ID (link)", "").strip():
            record["platform"].append("YouTube")

        merged.append(record)

    return merged


# ---------------------------------------------------------------------------
# Notion property builder
# ---------------------------------------------------------------------------


def build_notion_properties(record, user_map=None):
    """Convert a merged record to Notion page properties."""
    props = {}

    # Title: "handle - Full Name"
    title_text = record.get("handle", record.get("creator_id", ""))
    if record.get("full_name"):
        title_text += f" - {record['full_name']}"
    props["Influencer ID | Project"] = {
        "title": [{"text": {"content": title_text}}]
    }

    # Outreach Date
    if record.get("outreach_date"):
        props["Outreach Date"] = {"date": {"start": record["outreach_date"]}}

    # Brand (list of brand names)
    brands = record.get("brand", [])
    if isinstance(brands, str):
        brands = [brands] if brands else []
    if brands:
        props["Brand"] = {"multi_select": [{"name": b} for b in brands]}

    # Platform
    if record.get("platform"):
        props["Platform"] = {"multi_select": [{"name": p} for p in record["platform"]]}

    # Product
    if record.get("product"):
        props["Product"] = {"multi_select": [{"name": p} for p in record["product"]]}

    # Stage
    if record.get("stage"):
        props["Stage"] = {"status": {"name": record["stage"]}}

    # Owner (people) - requires user ID lookup
    if record.get("pic") and user_map:
        pic_lower = record["pic"].lower().strip()
        user_id = user_map.get(pic_lower)
        if not user_id:
            # Try partial match
            for name, uid in user_map.items():
                if pic_lower in name or name in pic_lower:
                    user_id = uid
                    break
        if user_id:
            props["Owner"] = {"people": [{"id": user_id}]}

    # Deadline / Posted Date
    if record.get("posted_date"):
        props["Deadline / Posted Date"] = {"date": {"start": record["posted_date"]}}

    # Posted Content Link
    if record.get("content_link") and record["content_link"].startswith("http"):
        props["Posted Content Link"] = {"url": record["content_link"]}

    # Content Rate
    if record.get("fee") is not None:
        props["Content Rate"] = {"number": record["fee"]}
        props["Total Paid Amount"] = {"number": record["fee"]}

    # Collaboration Count
    if record.get("collab_count"):
        props["Collaboration Count"] = {"select": {"name": record["collab_count"]}}

    # Notes
    if record.get("notes"):
        props["Notes"] = {"rich_text": [{"text": {"content": record["notes"]}}]}

    return props


# ---------------------------------------------------------------------------
# Sync logic
# ---------------------------------------------------------------------------


def extract_title(page_data):
    """Extract title text from Notion page data."""
    title_val = page_data.get("Influencer ID | Project", "")
    if isinstance(title_val, str):
        return title_val
    return ""


def match_records(merged_records, notion_pages):
    """Match merged records to existing Notion pages by handle or full name."""
    notion_lookup = {}
    for page in notion_pages:
        data = extract_notion_page_data(page)
        title = extract_title(data).lower().strip()
        if title:
            notion_lookup[title] = data

    to_create = []
    to_update = []

    for record in merged_records:
        title = record.get("handle", record.get("creator_id", ""))
        if record.get("full_name"):
            title += f" - {record['full_name']}"
        title_lower = title.lower().strip()
        handle_lower = record.get("handle", "").lower().strip()
        name_lower = record.get("full_name", "").lower().strip()

        # Try exact title match
        if title_lower in notion_lookup:
            to_update.append((record, notion_lookup[title_lower]["_page_id"]))
            del notion_lookup[title_lower]
            continue

        # Try handle-at-start match (Notion title = "handle - Full Name")
        found = False
        if handle_lower:
            for notion_title, notion_data in list(notion_lookup.items()):
                notion_handle = notion_title.split(" - ")[0].strip()
                if handle_lower == notion_handle:
                    to_update.append((record, notion_data["_page_id"]))
                    del notion_lookup[notion_title]
                    found = True
                    break

        # Try handle as substring in Notion title (for brand tabs using full names as Creator ID)
        if not found and handle_lower and len(handle_lower) > 5:
            for notion_title, notion_data in list(notion_lookup.items()):
                if handle_lower in notion_title:
                    to_update.append((record, notion_data["_page_id"]))
                    del notion_lookup[notion_title]
                    found = True
                    break

        # Try full name match (name appears in Notion title)
        if not found and name_lower and len(name_lower) > 3:
            for notion_title, notion_data in list(notion_lookup.items()):
                if name_lower in notion_title:
                    to_update.append((record, notion_data["_page_id"]))
                    del notion_lookup[notion_title]
                    found = True
                    break

        if not found:
            to_create.append(record)

    # Remaining Notion pages = orphans (in Notion but not in Sheets)
    orphans = list(notion_lookup.values())

    return to_create, to_update, orphans


def sync_to_notion(to_create, to_update, user_map, dry_run=False):
    """Execute creates and updates to Notion."""
    created_count = 0
    updated_count = 0
    errors = []

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Syncing to Notion...")
    print(f"  To create: {len(to_create)}")
    print(f"  To update: {len(to_update)}")

    # Create new pages
    for i, record in enumerate(to_create):
        title = record.get("handle", "") + (" - " + record.get("full_name", "") if record.get("full_name") else "")
        print(f"  [{i + 1}/{len(to_create)}] CREATE: {title}")

        if dry_run:
            created_count += 1
            continue

        props = build_notion_properties(record, user_map)
        body = {
            "parent": {"database_id": NOTION_DB_ID},
            "properties": props,
        }
        try:
            result = notion_api("POST", "pages", body)
            if result:
                created_count += 1
            time.sleep(0.35)
        except Exception as e:
            print(f"    ERROR: {e}")
            errors.append({"action": "create", "title": title, "error": str(e)})

        if (i + 1) % 10 == 0:
            print(f"    Progress: {i + 1}/{len(to_create)} created")

    # Update existing pages
    for i, (record, page_id) in enumerate(to_update):
        title = record.get("handle", "") + (" - " + record.get("full_name", "") if record.get("full_name") else "")
        print(f"  [{i + 1}/{len(to_update)}] UPDATE: {title}")

        if dry_run:
            updated_count += 1
            continue

        props = build_notion_properties(record, user_map)
        body = {"properties": props}
        try:
            result = notion_api("PATCH", f"pages/{page_id}", body)
            if result:
                updated_count += 1
            time.sleep(0.35)
        except Exception as e:
            print(f"    ERROR: {e}")
            errors.append({"action": "update", "title": title, "error": str(e)})

        if (i + 1) % 10 == 0:
            print(f"    Progress: {i + 1}/{len(to_update)} updated")

    return created_count, updated_count, errors


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------


def write_sync_report(to_create, to_update, orphans, errors, dry_run=False):
    """Write an Excel sync report."""
    prefix = "dryrun_sync_report" if dry_run else "sync_report"
    filepath = get_output_path("influencer", prefix)

    wb = openpyxl.Workbook()

    # Created sheet
    ws_create = wb.active
    ws_create.title = "To Create" if dry_run else "Created"
    headers = ["Creator ID", "Handle", "Full Name", "Brand", "Stage", "Fee"]
    ws_create.append(headers)
    for row in ws_create[1]:
        row.font = Font(bold=True)
    for rec in to_create:
        brand = rec.get("brand", "")
        if isinstance(brand, list):
            brand = ", ".join(brand)
        ws_create.append([
            rec.get("creator_id", ""),
            rec.get("handle", ""),
            rec.get("full_name", ""),
            brand,
            rec.get("stage", ""),
            rec.get("fee", ""),
        ])

    # Updated sheet
    ws_update = wb.create_sheet("To Update" if dry_run else "Updated")
    ws_update.append(headers)
    for row in ws_update[1]:
        row.font = Font(bold=True)
    for rec, page_id in to_update:
        brand = rec.get("brand", "")
        if isinstance(brand, list):
            brand = ", ".join(brand)
        ws_update.append([
            rec.get("creator_id", ""),
            rec.get("handle", ""),
            rec.get("full_name", ""),
            brand,
            rec.get("stage", ""),
            rec.get("fee", ""),
        ])

    # Orphans sheet
    ws_orphan = wb.create_sheet("Orphans (Notion only)")
    ws_orphan.append(["Page ID", "Title"])
    for row in ws_orphan[1]:
        row.font = Font(bold=True)
    for orphan in orphans:
        ws_orphan.append([
            orphan.get("_page_id", ""),
            extract_title(orphan),
        ])

    # Errors sheet
    if errors:
        ws_err = wb.create_sheet("Errors")
        ws_err.append(["Action", "Title", "Error"])
        for row in ws_err[1]:
            row.font = Font(bold=True)
        for err in errors:
            ws_err.append([err["action"], err["title"], err["error"]])

    wb.save(filepath)
    print(f"\n  Report saved: {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# Discover mode
# ---------------------------------------------------------------------------


def discover():
    """Discover and print schemas from both Google Sheets and Notion."""
    print("=" * 60)
    print("SCHEMA DISCOVERY")
    print("=" * 60)

    # --- Notion ---
    print("\n--- Notion Database ---")
    props = read_notion_db_schema()
    if props:
        for name, info in sorted(props.items()):
            options = info.get("options", [])
            opt_str = f" (options: {', '.join(options[:10])})" if options else ""
            print(f"  {name}: {info['type']}{opt_str}")
    else:
        print("  ERROR: Could not read Notion database schema")

    # --- Notion Users ---
    print("\n--- Notion Users ---")
    users = read_notion_users()
    if users:
        for name, uid in users.items():
            print(f"  {name}: {uid}")
    else:
        print("  No users found (or insufficient permissions)")

    # --- Google Sheets ---
    print("\n--- Google Sheets ---")
    gc = get_gspread_client()
    if gc:
        headers, master_rows = read_master_db_gspread(gc)
        print(f"\n  Master DB Headers ({len(headers)} columns):")
        for i, h in enumerate(headers, 1):
            if h:
                print(f"    {i}. {h}")
        print(f"  Master DB Rows: {len(master_rows)}")

        print("\n  Brand Tabs:")
        sh = gc.open_by_key(SHEET_ID)
        for ws in sh.worksheets():
            if str(ws.id) != MASTER_GID:
                print(f"    Tab: '{ws.title}' (GID: {ws.id})")
    else:
        print("  Falling back to CSV export (Master DB only)...")
        headers, master_rows = read_master_db_csv()
        print(f"  Master DB Headers ({len(headers)} columns):")
        for i, h in enumerate(headers, 1):
            if h:
                print(f"    {i}. {h}")
        print(f"  Master DB Rows: {len(master_rows)}")
        print("  WARNING: Brand tabs not accessible without Service Account")

    # --- Existing Notion Data ---
    print("\n--- Existing Notion Pages ---")
    pages = read_notion_pages()
    print(f"  Total pages: {len(pages)}")
    for i, page in enumerate(pages[:5]):
        data = extract_notion_page_data(page)
        title = extract_title(data)
        print(f"    {i + 1}. {title}")
    if len(pages) > 5:
        print(f"    ... and {len(pages) - 5} more")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Sync influencer data: Google Sheets -> Notion"
    )
    parser.add_argument("--discover", action="store_true",
                        help="Discover schemas from both systems")
    parser.add_argument("--dry-run", action="store_true",
                        help="Compare data without writing to Notion")
    parser.add_argument("--sync", action="store_true",
                        help="Execute sync to Notion")
    parser.add_argument("--master-only", action="store_true",
                        help="Sync Master DB tab only (no brand tabs)")
    args = parser.parse_args()

    if not any([args.discover, args.dry_run, args.sync]):
        parser.print_help()
        return

    if args.discover:
        discover()
        return

    # --- Step 1: Read Google Sheets ---
    print("=" * 60)
    print(f"{'[DRY RUN] ' if args.dry_run else ''}INFLUENCER SYNC")
    print("=" * 60)

    print("\nStep 1: Reading Google Sheets...")
    gc = get_gspread_client()

    if gc:
        headers, master_rows = read_master_db_gspread(gc)
    else:
        print("  No Service Account. Using CSV export for Master DB...")
        headers, master_rows = read_master_db_csv()
        if not args.master_only:
            print("  WARNING: Brand tabs not accessible. Use --master-only or set up Service Account.")
            args.master_only = True

    print(f"  Master DB: {len(master_rows)} rows")

    brand_tabs_data = []
    if not args.master_only and gc:
        brand_tabs_data = read_brand_tabs(gc)
        total_brand_rows = sum(len(rows) for _, _, rows in brand_tabs_data)
        print(f"  Brand tabs: {len(brand_tabs_data)} tabs, {total_brand_rows} total rows")

    # --- Step 2: Merge data ---
    print("\nStep 2: Merging data...")
    if args.master_only or not brand_tabs_data:
        merged = merge_master_only(master_rows)
        print(f"  Merged records (master only): {len(merged)}")
    else:
        merged = merge_data(master_rows, brand_tabs_data)
        print(f"  Merged records: {len(merged)}")

    if not merged:
        print("  No data to sync. Exiting.")
        return

    # --- Step 3: Read existing Notion data ---
    print("\nStep 3: Reading existing Notion pages...")
    notion_pages = read_notion_pages()
    print(f"  Existing pages: {len(notion_pages)}")

    # --- Step 4: Match records ---
    print("\nStep 4: Matching records...")
    to_create, to_update, orphans = match_records(merged, notion_pages)
    print(f"  To create: {len(to_create)}")
    print(f"  To update: {len(to_update)}")
    print(f"  Orphans (Notion only): {len(orphans)}")

    # --- Step 5: Get user map for Owner field ---
    print("\nStep 5: Loading Notion users...")
    user_map = read_notion_users()
    print(f"  Found {len(user_map)} users")

    # --- Step 6: Sync ---
    created, updated, errors = sync_to_notion(
        to_create, to_update, user_map, dry_run=args.dry_run
    )

    # --- Step 7: Report ---
    print("\nStep 7: Writing report...")
    report_path = write_sync_report(to_create, to_update, orphans, errors, dry_run=args.dry_run)

    # --- Summary ---
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Created: {created}")
    print(f"  Updated: {updated}")
    print(f"  Orphans: {len(orphans)}")
    print(f"  Errors:  {len(errors)}")
    print(f"  Report:  {report_path}")


if __name__ == "__main__":
    main()
