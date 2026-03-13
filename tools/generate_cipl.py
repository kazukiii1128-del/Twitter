"""
Generate CIPL (Commercial Invoice & Packing List) from packing info data.

Usage:
  python tools/generate_cipl.py [shipment_key]

  shipment_key: Key from SHIPMENTS config (default: "2601_2nd")
  Example: python tools/generate_cipl.py 2601_1st

Reads:
  - REFERENCE/ packing info file (per shipment config)
  - REFERENCE/2025_Ex Price_Grosmimi_20250930_미국_카톤당수량 업뎃.xlsx  (base price list)

Pricing:
  LFU sells to FLT at Ex Price * 1.05 (5% markup), rounded to 2 decimal places.

Outputs:
  - Data Storage/export/{naming convention}.xlsx
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ─── Configuration ───
MASTER_FILE = os.path.join(BASE_DIR, "REFERENCE",
                           "2025_Ex Price_Grosmimi_20250930_미국_카톤당수량 업뎃.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "Data Storage", "export")

# 파일명 양식: 년월_브랜드_운송방식_Exporter_Importer_FinalConsignee_Destination.ext
SHIPMENTS = {
    "2601_1st": {
        "packing_file": os.path.join(BASE_DIR, "REFERENCE", "미국 1월 1차_팩킹정보(제품)_0220.xlsx"),
        "output_file": os.path.join(OUTPUT_DIR, "2026-01-1st_Grosmimi_SEA_LFU_FLT_WBF_USA.xlsx"),
        "invoice_no": "USA-FLEETERS-26-01-1",
        "invoice_date": datetime(2026, 1, 15),
        "remark_code": "zzbb_sea_20260115",
        "consignee": "WBF",
    },
    "2601_2nd": {
        "packing_file": os.path.join(BASE_DIR, "REFERENCE", "미국 1월2차_팩킹정보.xlsx"),
        "output_file": os.path.join(OUTPUT_DIR, "2026-01-2nd_Grosmimi_SEA_LFU_FLT_WBF_USA.xlsx"),
        "invoice_no": "USA-FLEETERS-26-01-2",
        "invoice_date": datetime(2026, 1, 31),
        "remark_code": "zzbb_sea_20260131",
        "consignee": "WBF",
    },
}
DEFAULT_SHIPMENT = "2601_2nd"

# ─── Company Info ───
EXPORTER = {
    "name": "LittlefingerUSA Inc.",
    "address": "A-320, 3, Godeung-ro, Sujeong-gu, Seongnam-si, Gyeonggi-do, Republic of Korea, 13105",
    "email": "grosmimi.usa@gmail.com",
    "tel": "82-10-4803-4704",
}
IMPORTER = {
    "name": "Fleeters Inc.",
    "address": "A-320, 3, Godeung-ro, Sujeong-gu, Seongnam-si, Gyeonggi-do, Republic of Korea, 13105",
    "email": "grosmimi.usa@gmail.com",
    "tel": "Tel : 82-10-4803-4704",
}

# Per-shipment consignee configs
CONSIGNEES = {
    "WBF": {
        "name": "Walk by Faith",
        "address": "5900 Katella Ave, BLDG C, STE 100, Cypress, CA 90630, USA",
        "email": "",
        "tel": "714-403-8718",
    },
    "Shipbob": {
        "name": "Fleeters Inc C/O ShipBob, Inc",
        "address": "28010 Eucalyptus Ave, Moreno Valley, CA, 92555",
        "email": "support@shipbob.com",
        "tel": "(844) 474-4726",
    },
}

# Invoice date, invoice no, remark code are per-shipment (see SHIPMENTS dict)
DESTINATION = "USA"
PAYMENT = "100% by T/T advance before shipment, General Transaction"
PRICE_TERMS = "EXW"
ORIGIN = "Republic of Korea"
PACKING_TYPE = "Standard Export Packing Box"

BANK_INFO = {
    "company": "ORBITERS Co.,Ltd.",
    "remittee": "LITTLEFINGERUSA",
    "bank": "KEB Hana Bank",
    "swift": "KOEXKRSE",
    "account": "630-010399-748",
}

SIGNER = "Chung Hae Jung / LittlefingerUSA Inc."

# ─── Manual product data for items not in Ex Price master ───
# Strap products (older items, weight/HS data from reference CIPL)
_STRAP_PINK = {
    "desc": "Straw Cup Strap(Pink)",
    "price": 0, "ex_price": 0,
    "net_wt_g": 30, "qty_ctn": 360, "gw_ctn": 17.30,
    "dim_l": 550, "dim_w": 550, "dim_h": 520, "cbm": 0.1573,
    "barcode": "8809466585623", "hs6": "6307.90", "hs10": "6307.90-9000",
    "inner": None, "inner_qty": None, "is_foc": True,
}
_STRAP_SKYBLUE = {
    "desc": "Straw Cup Strap(Sky Blue)",
    "price": 0, "ex_price": 0,
    "net_wt_g": 30, "qty_ctn": 360, "gw_ctn": 17.30,
    "dim_l": 550, "dim_w": 550, "dim_h": 520, "cbm": 0.1573,
    "barcode": "8809466585678", "hs6": "6307.90", "hs10": "6307.90-9000",
    "inner": None, "inner_qty": None, "is_foc": True,
}
_STRAP_CHARCOAL = {
    "desc": "Straw Cup Strap(Charcoal)",
    "price": 0, "ex_price": 0,
    "net_wt_g": 30, "qty_ctn": 360, "gw_ctn": 17.30,
    "dim_l": 550, "dim_w": 550, "dim_h": 520, "cbm": 0.1573,
    "barcode": "8809466584848", "hs6": "6307.90", "hs10": "6307.90-9000",
    "inner": None, "inner_qty": None, "is_foc": True,
}
# Indexed by both barcode AND item code
MANUAL_PRODUCTS = {
    "8809466585623": _STRAP_PINK,
    "8809466585678": _STRAP_SKYBLUE,
    "8809466584848": _STRAP_CHARCOAL,
    "GP4-SCS004": _STRAP_PINK,
    "gp4-scs004": _STRAP_PINK,
    "GP4-SCS006": _STRAP_SKYBLUE,
    "gp4-scs006": _STRAP_SKYBLUE,
    "GP4-SCS011": _STRAP_CHARCOAL,
    "gp4-scs011": _STRAP_CHARCOAL,
}


def load_product_master(filepath):
    """Load product master from Ex Price file. Apply 5% markup for LFU->FLT pricing.

    Ex Price file columns (sheet: 'Order Sheet _ USA', data starts row 16):
      B(2):  Item Code
      C(3):  Description
      G(7):  Net Weight per item (g)
      H(8):  Ex Price Each (USD) - base price before markup
      I(9):  Total Quantity per CTN
      O(15): Inner boxes per CTN
      P(16): Items per inner box
      Q(17): CTN Dimension L (cm)
      R(18): CTN Dimension W (cm)
      S(19): CTN Dimension H (cm)
      T(20): CBM per CTN
      V(22): Gross Weight per CTN (kg)
      X(24): Barcode
      Y(25): HS Code (6-digit)
      Z(26): HS Code (10-digit)
    """
    MARKUP = 1.05  # LFU sells to FLT at Ex Price + 5%
    master = {}

    def to_float(v):
        """Safely convert to float, return None for non-numeric values like 'TBU'."""
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def to_int(v):
        if v is None:
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["Order Sheet _ USA"]
        for row_idx in range(16, ws.max_row + 1):
            code = ws.cell(row=row_idx, column=2).value  # B: Item Code
            if not code:
                continue
            code = str(code).strip()
            if code in ('', 'Item Code') or '\n' in code:
                continue

            try:
                desc = ws.cell(row=row_idx, column=3).value or ""
                ex_price = to_float(ws.cell(row=row_idx, column=8).value)  # H
                net_wt = to_float(ws.cell(row=row_idx, column=7).value)    # G
                qty_ctn = to_int(ws.cell(row=row_idx, column=9).value)     # I
                inner = to_int(ws.cell(row=row_idx, column=15).value)      # O
                inner_qty = to_int(ws.cell(row=row_idx, column=16).value)  # P
                dim_l = to_float(ws.cell(row=row_idx, column=17).value)    # Q
                dim_w = to_float(ws.cell(row=row_idx, column=18).value)    # R
                dim_h = to_float(ws.cell(row=row_idx, column=19).value)    # S
                cbm = to_float(ws.cell(row=row_idx, column=20).value)      # T
                gw_ctn = to_float(ws.cell(row=row_idx, column=22).value)   # V
                barcode = ws.cell(row=row_idx, column=24).value            # X
                hs6 = ws.cell(row=row_idx, column=25).value                # Y
                hs10 = ws.cell(row=row_idx, column=26).value               # Z

                # Apply 5% markup, round to 2 decimal places
                unit_price = round(ex_price * MARKUP, 2) if ex_price else None

                # Format barcode as string (remove .0)
                if barcode:
                    bc_str = str(barcode)
                    if bc_str.endswith('.0'):
                        bc_str = bc_str[:-2]
                    barcode_clean = bc_str
                else:
                    barcode_clean = ""

                master[code] = {
                    "desc": str(desc).strip(),
                    "ex_price": ex_price,
                    "price": unit_price,  # with 5% markup
                    "net_wt_g": net_wt,
                    "qty_ctn": qty_ctn,
                    "inner": inner,
                    "inner_qty": inner_qty,
                    "dim_l": dim_l,
                    "dim_w": dim_w,
                    "dim_h": dim_h,
                    "cbm": cbm,
                    "gw_ctn": gw_ctn,
                    "barcode": barcode_clean,
                    "hs6": str(hs6) if hs6 else "",
                    "hs10": str(hs10) if hs10 else "",
                }

                # Also store lowercase for case-insensitive matching
                master[code.lower()] = master[code]
            except Exception as e:
                print(f"  WARN: Row {row_idx} ({code}): {e}")
                continue
        wb.close()
    except Exception as e:
        print(f"  ERROR: Could not load master file: {e}")
        import traceback
        traceback.print_exc()
    return master


def match_product(item_code, master):
    """Find a product in the master data. Ex Price file should have exact codes."""
    code = str(item_code).strip()

    # Exact match
    if code in master:
        return master[code]

    # Case-insensitive
    lower = code.lower()
    if lower in master:
        return master[lower]

    # Prefix match fallback (same design + size)
    for mk in master:
        if isinstance(mk, str) and len(mk) > 8 and len(code) > 8:
            if mk[:6] == code[:6]:
                mk_size = '2' if '-2' in mk else ('3' if '-3' in mk else '')
                code_size = '2' if '-2' in code else ('3' if '-3' in code else '')
                if mk_size == code_size and mk_size:
                    return master[mk]

    return None


def match_product_or_manual(item_code, barcode, master):
    """Try master first, then fall back to MANUAL_PRODUCTS by item code or barcode."""
    product = match_product(item_code, master)
    if product:
        return product

    # Fallback: check MANUAL_PRODUCTS by item code
    code = str(item_code).strip()
    if code in MANUAL_PRODUCTS:
        return MANUAL_PRODUCTS[code]
    if code.lower() in MANUAL_PRODUCTS:
        return MANUAL_PRODUCTS[code.lower()]

    # Fallback: check MANUAL_PRODUCTS by barcode
    bc = str(barcode).strip() if barcode else ""
    if bc in MANUAL_PRODUCTS:
        return MANUAL_PRODUCTS[bc]

    return None


def load_packing_info(filepath):
    """Load packing info and return list of items."""
    items = []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_pallet = 1
    for row_idx in range(3, ws.max_row + 1):
        item_code = ws.cell(row=row_idx, column=2).value  # B: Item Code
        if not item_code:
            # Check if it's the total row
            a_val = ws.cell(row=row_idx, column=1).value
            if a_val and 'total' in str(a_val).lower():
                break
            continue

        desc = ws.cell(row=row_idx, column=3).value or ""  # C: Description
        qty_per_ctn = ws.cell(row=row_idx, column=4).value  # D: Quantity per carton
        order_ctn = ws.cell(row=row_idx, column=6).value  # F: Order Carton count
        order_box = ws.cell(row=row_idx, column=7).value  # G: Order Box (inner box count)
        total_qty = ws.cell(row=row_idx, column=8).value  # H: Total Qty
        inner_box = ws.cell(row=row_idx, column=9).value  # I: inner box per CTN
        inner_qty = ws.cell(row=row_idx, column=10).value  # J: inner QTY (items per CTN)
        ctn_no = ws.cell(row=row_idx, column=11).value  # K: CTN/No (carton number range)
        pallet_no = ws.cell(row=row_idx, column=12).value  # L: Pallet number
        note = ws.cell(row=row_idx, column=1).value  # A: Note (e.g., 샘플박스)

        if pallet_no and isinstance(pallet_no, (int, float)):
            current_pallet = int(pallet_no)

        items.append({
            "code": str(item_code).strip(),
            "desc": str(desc).strip().replace('\n', ' '),
            "qty_per_ctn": int(qty_per_ctn) if qty_per_ctn else 0,
            "ctn_qty": float(order_ctn) if order_ctn else 0,
            "inner_box_qty": int(order_box) if order_box else 0,
            "total_qty": int(total_qty) if total_qty else 0,
            "inner_per_ctn": int(inner_box) if inner_box else 0,
            "items_per_ctn": int(inner_qty) if inner_qty else 0,
            "ctn_no": str(ctn_no) if ctn_no else "",
            "pallet": current_pallet,
            "note": str(note).strip() if note else "",
        })

    wb.close()
    return items


def create_cipl(items, master, output_path, shipment_cfg):
    """Create the CIPL Excel document."""
    INVOICE_DATE = shipment_cfg["invoice_date"]
    INVOICE_NO = shipment_cfg["invoice_no"]
    REMARK_CODE = shipment_cfg["remark_code"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CI, PL"

    # ─── Styles ───
    title_font = Font(name="Calibri", size=14, bold=True)
    header_font = Font(name="Calibri", size=10, bold=True)
    label_font = Font(name="Calibri", size=9, bold=True)
    data_font = Font(name="Calibri", size=9)
    small_font = Font(name="Calibri", size=8)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='thin'))

    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # ─── Column widths ───
    col_widths = {
        'A': 3,    # spacer
        'B': 8,    # Pallet NO.
        'C': 6,    # CTN NO start
        'D': 3,    # dash
        'E': 6,    # CTN NO end
        'F': 42,   # Commodity & Description
        'G': 16,   # SKU/Barcode
        'H': 8,    # Q'ty/CTN
        'I': 8,    # CTN QTY
        'J': 8,    # EA
        'K': 10,   # Unit Price
        'L': 12,   # Amount
        'M': 8,    # PER NW
        'N': 10,   # N.W
        'O': 8,    # PER GW
        'P': 10,   # G.W
        'Q': 6,    # Dim L
        'R': 6,    # Dim W
        'S': 6,    # Dim H
        'T': 8,    # CBM
        'U': 8,    # ORIGIN
        'V': 12,   # HS CODE
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ─── Row 2: Title ───
    ws.merge_cells('B2:T2')
    ws['B2'] = "COMMERCIAL INVOICE & PACKING LIST"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center')

    # ─── Row 4-8: Exporter info ───
    ws['B4'] = "Shipper/Exporter"
    ws['B4'].font = label_font
    ws['O4'] = DESTINATION
    ws['O4'].font = label_font

    ws['B5'] = EXPORTER["name"]
    ws['B5'].font = data_font

    ws['B6'] = EXPORTER["address"]
    ws['B6'].font = data_font

    ws['B7'] = EXPORTER["email"]
    ws['B7'].font = data_font
    ws['O7'] = INVOICE_DATE
    ws['O7'].font = data_font
    ws['O7'].number_format = 'YYYY-MM-DD'

    ws['B8'] = "Tel :"
    ws['B8'].font = label_font
    ws['D8'] = EXPORTER["tel"]
    ws['D8'].font = data_font
    ws['O8'] = INVOICE_NO
    ws['O8'].font = data_font

    # ─── Row 9-13: Consignee info ───
    consignee_key = shipment_cfg.get("consignee", "WBF")
    consignee = CONSIGNEES.get(consignee_key, CONSIGNEES["WBF"])

    ws['B9'] = "Consignee / Importer"
    ws['B9'].font = label_font
    ws['H9'] = "Ultimate Consignee"
    ws['H9'].font = label_font
    ws['O9'] = "Final Destination"
    ws['O9'].font = label_font

    ws['B10'] = IMPORTER["name"]
    ws['B10'].font = data_font
    ws['H10'] = consignee["name"]
    ws['H10'].font = data_font
    ws['O10'] = consignee["name"]
    ws['O10'].font = data_font

    ws['B11'] = IMPORTER["address"]
    ws['B11'].font = data_font
    ws['H11'] = consignee["address"]
    ws['H11'].font = data_font
    ws['O11'] = consignee["address"]
    ws['O11'].font = data_font

    ws['B12'] = IMPORTER["email"]
    ws['B12'].font = data_font
    if consignee.get("email"):
        ws['H12'] = consignee["email"]
        ws['H12'].font = data_font
        ws['O12'] = consignee["email"]
        ws['O12'].font = data_font

    ws['B13'] = IMPORTER["tel"]
    ws['B13'].font = data_font
    ws['H13'] = f"Tel : {consignee['tel']}"
    ws['H13'].font = data_font
    ws['O13'] = consignee["tel"]
    ws['O13'].font = data_font

    # ─── Row 14-19: Trade terms ───
    terms = [
        ("Destination", DESTINATION),
        ("Payment", PAYMENT),
        ("Price Terms", PRICE_TERMS),
        ("Origin", ORIGIN),
        ("Packing", PACKING_TYPE),
    ]
    for i, (label, value) in enumerate(terms):
        r = 14 + i
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = label_font
        ws[f'H{r}'] = value
        ws[f'H{r}'].font = data_font

    # Row 19: Total PKG placeholder (filled after data)
    ws['B19'] = "Total PKG"
    ws['B19'].font = label_font
    ws['H19'] = "Total : "
    ws['H19'].font = label_font

    # ─── Row 22-23: Column headers ───
    headers = [
        ('B', "Pallet NO."),
        ('C', "CTN NO."),
        ('F', "Commodity & Description"),
        ('G', "SKU/Barcode"),
        ('H', "Q'ty / \nCTN"),
        ('I', "CTN QTY."),
        ('J', "EA"),
        ('K', "Unit Price \n(USD)"),
        ('L', "Amount \n(USD)"),
        ('M', "PER"),
        ('N', "N.W\n(KG)"),
        ('O', "PER"),
        ('P', "G.W\n(KG)"),
        ('Q', "Dimensions"),
        ('T', "CBM"),
        ('U', "ORIGIN"),
        ('V', "HS CODE"),
    ]
    for col, text in headers:
        cell = ws[f'{col}22']
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # Sub-headers row 23
    ws['M23'] = "CTN"
    ws['M23'].font = small_font
    ws['M23'].alignment = center
    ws['M23'].border = thin_border
    ws['O23'] = "CTN"
    ws['O23'].font = small_font
    ws['O23'].alignment = center
    ws['O23'].border = thin_border
    ws['Q23'] = "L x W x H (cm)"
    ws['Q23'].font = small_font
    ws['Q23'].alignment = center
    ws['Q23'].border = thin_border

    # Apply borders to all header cells row 22-23
    for col_idx in range(2, 23):  # B to V
        for r in [22, 23]:
            cell = ws.cell(row=r, column=col_idx)
            cell.border = thin_border
            if r == 22 and not cell.value:
                cell.fill = header_fill

    # ─── Data rows ───
    data_start = 24
    row = data_start
    total_ctn = 0
    total_ea = 0
    total_amount = 0
    total_nw = 0
    total_gw = 0
    total_cbm = 0
    unmatched_items = []

    for item in items:
        # Try master first, then MANUAL_PRODUCTS by barcode
        product = match_product_or_manual(item["code"], item.get("barcode", ""), master)

        # F.O.C (Free of Charge) items: unit price = $0
        is_foc = 'F.O.C' in item.get("note", '') or 'f.o.c' in item.get("note", '').lower()
        # Also check if manual product is marked F.O.C
        if product and product.get("is_foc"):
            is_foc = True

        # Get values from product master or use defaults
        if product:
            unit_price = 0 if is_foc else product["price"]  # F.O.C = $0
            # N.W per CTN = net weight per item (g) * total items per CTN / 1000
            items_in_ctn = item["qty_per_ctn"] or product.get("qty_ctn", 0) or 1
            nw_per_ctn = (product["net_wt_g"] or 0) * items_in_ctn / 1000 if product["net_wt_g"] else None
            gw_per_ctn = product["gw_ctn"]
            dim_l = product["dim_l"]
            dim_w = product["dim_w"]
            dim_h = product["dim_h"]
            cbm_per_ctn = product["cbm"]
            barcode = product["barcode"]
            hs_code = product.get("hs6", "")
        else:
            unit_price = 0 if is_foc else None
            nw_per_ctn = None
            gw_per_ctn = None
            dim_l = None
            dim_w = None
            dim_h = None
            cbm_per_ctn = None
            barcode = ""
            hs_code = ""
            if not is_foc:
                unmatched_items.append(item["code"])

        ctn_qty = item["ctn_qty"]
        ea = item["total_qty"]
        qty_per_ctn = item["qty_per_ctn"]  # total items per CTN (column D from packing)

        # Detect shared carton (fractional CTN, e.g. 0.5 means 2 SKUs share 1 carton)
        is_shared_ctn = isinstance(ctn_qty, float) and ctn_qty > 0 and ctn_qty < 1

        # Calculate amounts
        amount = round(unit_price * ea, 2) if unit_price is not None and ea else None
        total_nw_item = round(nw_per_ctn * ctn_qty, 2) if nw_per_ctn and ctn_qty else None
        total_gw_item = round(gw_per_ctn * ctn_qty, 2) if gw_per_ctn and ctn_qty else None
        total_cbm_item = round(cbm_per_ctn * ctn_qty, 4) if cbm_per_ctn and ctn_qty else None

        # Parse CTN NO range
        ctn_no = item["ctn_no"]
        has_ctn_no = bool(ctn_no and str(ctn_no).strip() and str(ctn_no).strip() != 'None')
        if has_ctn_no and '~' in str(ctn_no):
            parts = str(ctn_no).split('~')
            ctn_start = parts[0].strip()
            ctn_end = parts[1].strip()
        elif has_ctn_no:
            ctn_start = str(ctn_no).strip()
            ctn_end = str(ctn_no).strip()
        else:
            ctn_start = ""
            ctn_end = ""

        # Write row
        ws.cell(row=row, column=2, value=item["pallet"]).font = data_font
        ws.cell(row=row, column=2).alignment = center
        if has_ctn_no:
            ws.cell(row=row, column=3, value=ctn_start).font = data_font
            ws.cell(row=row, column=3).alignment = center
            if ctn_start != ctn_end:
                ws.cell(row=row, column=4, value='-').font = data_font
                ws.cell(row=row, column=4).alignment = center
                ws.cell(row=row, column=5, value=ctn_end).font = data_font
                ws.cell(row=row, column=5).alignment = center

        # Description - add note prefix if sample box or F.O.C
        desc = item["desc"]
        if is_foc:
            desc = f"[F.O.C] {desc}"
        elif item["note"] and "샘플" in item["note"]:
            desc = f"[SAMPLE] {desc}"
        ws.cell(row=row, column=6, value=desc).font = data_font
        ws.cell(row=row, column=6).alignment = left_wrap

        ws.cell(row=row, column=7, value=barcode).font = small_font
        ws.cell(row=row, column=7).alignment = center

        ws.cell(row=row, column=8, value=qty_per_ctn).font = data_font
        ws.cell(row=row, column=8).alignment = center

        ws.cell(row=row, column=9, value=ctn_qty).font = data_font
        ws.cell(row=row, column=9).alignment = center

        ws.cell(row=row, column=10, value=ea).font = data_font
        ws.cell(row=row, column=10).alignment = center
        ws.cell(row=row, column=10).number_format = '#,##0'

        tbd_font = Font(name="Calibri", size=9, color="FF0000")
        tbd_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        is_unmatched = not product and not is_foc

        if unit_price is not None:
            ws.cell(row=row, column=11, value=unit_price).font = data_font
            ws.cell(row=row, column=11).alignment = right_align
            ws.cell(row=row, column=11).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=11, value="TBD").font = tbd_font
            ws.cell(row=row, column=11).fill = tbd_fill
            ws.cell(row=row, column=11).alignment = center

        if amount is not None:
            ws.cell(row=row, column=12, value=amount).font = data_font
            ws.cell(row=row, column=12).alignment = right_align
            ws.cell(row=row, column=12).number_format = '#,##0.00'
            total_amount += amount
        elif is_unmatched:
            ws.cell(row=row, column=12, value="TBD").font = tbd_font
            ws.cell(row=row, column=12).fill = tbd_fill
            ws.cell(row=row, column=12).alignment = center

        # N.W per CTN
        if nw_per_ctn:
            ws.cell(row=row, column=13, value=round(nw_per_ctn, 2)).font = data_font
            ws.cell(row=row, column=13).alignment = right_align
            ws.cell(row=row, column=13).number_format = '#,##0.00'
        elif is_unmatched:
            ws.cell(row=row, column=13, value="TBD").font = tbd_font
            ws.cell(row=row, column=13).fill = tbd_fill
            ws.cell(row=row, column=13).alignment = center

        # Total N.W
        if total_nw_item:
            ws.cell(row=row, column=14, value=total_nw_item).font = data_font
            ws.cell(row=row, column=14).alignment = right_align
            ws.cell(row=row, column=14).number_format = '#,##0.00'
            total_nw += total_nw_item
        elif is_unmatched:
            ws.cell(row=row, column=14, value="TBD").font = tbd_font
            ws.cell(row=row, column=14).fill = tbd_fill
            ws.cell(row=row, column=14).alignment = center

        # G.W per CTN (skip for shared carton without own CTN NO)
        if gw_per_ctn and has_ctn_no:
            ws.cell(row=row, column=15, value=round(gw_per_ctn, 2)).font = data_font
            ws.cell(row=row, column=15).alignment = right_align
            ws.cell(row=row, column=15).number_format = '#,##0.00'
        elif is_unmatched and has_ctn_no:
            ws.cell(row=row, column=15, value="TBD").font = tbd_font
            ws.cell(row=row, column=15).fill = tbd_fill
            ws.cell(row=row, column=15).alignment = center

        # Total G.W
        if total_gw_item:
            ws.cell(row=row, column=16, value=total_gw_item).font = data_font
            ws.cell(row=row, column=16).alignment = right_align
            ws.cell(row=row, column=16).number_format = '#,##0.00'
            total_gw += total_gw_item
        elif is_unmatched:
            ws.cell(row=row, column=16, value="TBD").font = tbd_font
            ws.cell(row=row, column=16).fill = tbd_fill
            ws.cell(row=row, column=16).alignment = center

        # Dimensions (skip for shared carton without own CTN NO)
        if has_ctn_no:
            if dim_l:
                ws.cell(row=row, column=17, value=dim_l).font = data_font
                ws.cell(row=row, column=17).alignment = center
            elif is_unmatched:
                ws.cell(row=row, column=17, value="TBD").font = tbd_font
                ws.cell(row=row, column=17).fill = tbd_fill
                ws.cell(row=row, column=17).alignment = center
            if dim_w:
                ws.cell(row=row, column=18, value=dim_w).font = data_font
                ws.cell(row=row, column=18).alignment = center
            if dim_h:
                ws.cell(row=row, column=19, value=dim_h).font = data_font
                ws.cell(row=row, column=19).alignment = center

        # CBM (skip display for shared carton without own CTN NO, but still add to total)
        if total_cbm_item:
            if has_ctn_no:
                ws.cell(row=row, column=20, value=total_cbm_item).font = data_font
                ws.cell(row=row, column=20).alignment = right_align
                ws.cell(row=row, column=20).number_format = '#,##0.0000'
            total_cbm += total_cbm_item
        elif is_unmatched and has_ctn_no:
            ws.cell(row=row, column=20, value="TBD").font = tbd_font
            ws.cell(row=row, column=20).fill = tbd_fill
            ws.cell(row=row, column=20).alignment = center

        # ORIGIN
        ws.cell(row=row, column=21, value="Korea").font = data_font
        ws.cell(row=row, column=21).alignment = center

        # HS CODE
        if hs_code:
            ws.cell(row=row, column=22, value=hs_code).font = data_font
            ws.cell(row=row, column=22).alignment = center
        elif is_unmatched:
            ws.cell(row=row, column=22, value="TBD").font = tbd_font
            ws.cell(row=row, column=22).fill = tbd_fill
            ws.cell(row=row, column=22).alignment = center

        # Borders for all data cells
        for col_idx in range(2, 23):  # B to V
            ws.cell(row=row, column=col_idx).border = thin_border

        total_ctn += ctn_qty
        total_ea += ea
        row += 1

    # ─── Total row ───
    total_row = row
    ws.cell(row=total_row, column=2, value="TOTAL ").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=2).fill = total_fill

    ws.cell(row=total_row, column=9, value=total_ctn).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=9).alignment = center
    ws.cell(row=total_row, column=9).fill = total_fill

    ws.cell(row=total_row, column=10, value=total_ea).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=10).alignment = center
    ws.cell(row=total_row, column=10).number_format = '#,##0'
    ws.cell(row=total_row, column=10).fill = total_fill

    ws.cell(row=total_row, column=12, value=round(total_amount, 2)).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=12).alignment = right_align
    ws.cell(row=total_row, column=12).number_format = '#,##0.00'
    ws.cell(row=total_row, column=12).fill = total_fill

    ws.cell(row=total_row, column=14, value=round(total_nw, 2)).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=14).alignment = right_align
    ws.cell(row=total_row, column=14).number_format = '#,##0.00'
    ws.cell(row=total_row, column=14).fill = total_fill

    ws.cell(row=total_row, column=16, value=round(total_gw, 2)).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=16).alignment = right_align
    ws.cell(row=total_row, column=16).number_format = '#,##0.00'
    ws.cell(row=total_row, column=16).fill = total_fill

    ws.cell(row=total_row, column=20, value=round(total_cbm, 4)).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=20).alignment = right_align
    ws.cell(row=total_row, column=20).number_format = '#,##0.0000'
    ws.cell(row=total_row, column=20).fill = total_fill

    for col_idx in range(2, 23):  # B to V
        ws.cell(row=total_row, column=col_idx).border = thin_border
        if not ws.cell(row=total_row, column=col_idx).fill or ws.cell(row=total_row, column=col_idx).fill.start_color.rgb == '00000000':
            ws.cell(row=total_row, column=col_idx).fill = total_fill

    # Update Total PKG in header (formatted with units like reference)
    ws['I19'] = f"{int(total_ea)} PCS"
    ws['I19'].font = Font(name="Calibri", size=10, bold=True)
    ws['J19'] = f"{int(total_ctn)} CTN"
    ws['J19'].font = data_font
    ws['K19'] = f"{round(total_cbm, 4)} CBM"
    ws['K19'].font = data_font

    # ─── Shipping Mark & Bank info ───
    mark_row = total_row + 1
    ws[f'B{mark_row}'] = "Shipping Mark"
    ws[f'B{mark_row}'].font = label_font
    ws[f'F{mark_row}'] = f"REMARK : {INVOICE_NO}"
    ws[f'F{mark_row}'].font = data_font

    ws[f'B{mark_row + 2}'] = BANK_INFO["company"]
    ws[f'B{mark_row + 2}'].font = data_font
    ws[f'F{mark_row + 1}'] = "Bank Information"
    ws[f'F{mark_row + 1}'].font = label_font

    ws[f'B{mark_row + 4}'] = f"PO: {INVOICE_NO}"
    ws[f'B{mark_row + 4}'].font = data_font
    ws[f'F{mark_row + 2}'] = f"Remittess's Name : {BANK_INFO['remittee']}"
    ws[f'F{mark_row + 2}'].font = data_font

    ws[f'F{mark_row + 3}'] = f"Bank Name : {BANK_INFO['bank']}"
    ws[f'F{mark_row + 3}'].font = data_font

    ws[f'F{mark_row + 4}'] = f"Swift Code : {BANK_INFO['swift']}"
    ws[f'F{mark_row + 4}'].font = data_font

    ws[f'F{mark_row + 5}'] = f"Remittess's Account No. : {BANK_INFO['account']}"
    ws[f'F{mark_row + 5}'].font = data_font

    # Signature
    ws[f'P{mark_row + 9}'] = SIGNER
    ws[f'P{mark_row + 9}'].font = data_font
    ws[f'P{mark_row + 10}'] = "President / Company Name"
    ws[f'P{mark_row + 10}'].font = small_font

    # ─── Print settings ───
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # ─── Save ───
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return {
        "total_items": len(items),
        "total_ctn": total_ctn,
        "total_ea": total_ea,
        "total_amount": round(total_amount, 2),
        "total_nw": round(total_nw, 2),
        "total_gw": round(total_gw, 2),
        "total_cbm": round(total_cbm, 4),
        "unmatched": unmatched_items,
    }


def main():
    # Select shipment from CLI arg or default
    shipment_key = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SHIPMENT
    if shipment_key not in SHIPMENTS:
        print(f"ERROR: Unknown shipment '{shipment_key}'. Available: {', '.join(SHIPMENTS.keys())}")
        sys.exit(1)

    cfg = SHIPMENTS[shipment_key]
    packing_file = cfg["packing_file"]
    output_file = cfg["output_file"]

    print("=" * 80)
    print(f"CIPL Generator - LFU (Exporter) -> FLT (Importer)  [{shipment_key}]")
    print("=" * 80)

    # Step 1: Load product master
    print("\n[1] Loading product master data...")
    master = load_product_master(MASTER_FILE)
    print(f"    Loaded {len(master) // 2} products from master")

    # Step 2: Load packing info
    print("\n[2] Loading packing info...")
    items = load_packing_info(packing_file)
    print(f"    Loaded {len(items)} line items")

    # Step 3: Generate CIPL
    print("\n[3] Generating CIPL...")
    result = create_cipl(items, master, output_file, cfg)

    print(f"\n{'=' * 80}")
    print(f"DONE! CIPL saved to: {output_file}")
    print(f"{'=' * 80}")
    print(f"  Items:       {result['total_items']}")
    print(f"  Total CTN:   {result['total_ctn']}")
    print(f"  Total EA:    {result['total_ea']:,}")
    print(f"  Total Amount: USD {result['total_amount']:,.2f}")
    print(f"  Total N.W:   {result['total_nw']:,.2f} KG")
    print(f"  Total G.W:   {result['total_gw']:,.2f} KG")
    print(f"  Total CBM:   {result['total_cbm']:.4f}")

    if result['unmatched']:
        unique_unmatched = sorted(set(result['unmatched']))
        print(f"\n  WARNING: {len(unique_unmatched)} item codes not found in master (price marked TBD):")
        for code in unique_unmatched:
            print(f"    - {code}")
    print()


if __name__ == "__main__":
    main()
