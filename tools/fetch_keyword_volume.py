"""
DataForSEO Keyword Volume Fetcher
Fetches keyword search volume from Google Ads, Amazon APIs, and Google Search Console.

Usage:
    python fetch_keyword_volume.py --keywords "keyword1,keyword2" [--channel google|amazon|both|gsc] [--location US] [--output path.xlsx]

Env vars required:
    DATAFORSEO_LOGIN
    DATAFORSEO_PASSWORD
    GOOGLE_SERVICE_ACCOUNT_PATH  (for GSC)
"""

import os
import sys
import json
import argparse
import requests
from datetime import datetime, date
from dotenv import load_dotenv

load_dotenv()

API_BASE = "https://api.dataforseo.com/v3"

# Location codes for DataForSEO
LOCATION_CODES = {
    "US": 2840,
    "UK": 2826,
    "DE": 2276,
    "FR": 2250,
    "CA": 2124,
    "AU": 2036,
    "IN": 2356,
    "IT": 2380,
    "ES": 2724,
    "MX": 2484,
    "NL": 2528,
    "SG": 2702,
}

LANGUAGE_CODES = {
    "US": "en", "UK": "en", "CA": "en", "AU": "en", "IN": "en", "SG": "en",
    "DE": "de", "FR": "fr", "IT": "it", "ES": "es", "MX": "es", "NL": "nl",
}


def get_auth():
    login = os.getenv("DATAFORSEO_LOGIN")
    password = os.getenv("DATAFORSEO_PASSWORD")
    if not login or not password:
        print("ERROR: DATAFORSEO_LOGIN and DATAFORSEO_PASSWORD must be set in .env")
        sys.exit(1)
    return (login, password)


def fetch_google_volume(keywords: list[str], location: str = "US") -> list[dict]:
    """Fetch search volume from Google Ads API (Live endpoint)."""
    auth = get_auth()
    location_code = LOCATION_CODES.get(location, 2840)
    language_code = LANGUAGE_CODES.get(location, "en")

    payload = [{
        "keywords": keywords,
        "location_code": location_code,
        "language_code": language_code,
    }]

    url = f"{API_BASE}/keywords_data/google_ads/search_volume/live"
    resp = requests.post(url, auth=auth, json=payload, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    if data.get("status_code") != 20000:
        print(f"Google API error: {data.get('status_message')}")
        return []

    results = []
    for task in data.get("tasks", []):
        if task.get("status_code") != 20000:
            print(f"  Task error: {task.get('status_message')}")
            continue
        for item in (task.get("result") or []):
            results.append({
                "keyword": item.get("keyword"),
                "channel": "Google",
                "search_volume": item.get("search_volume"),
                "competition": item.get("competition"),
                "competition_index": item.get("competition_index"),
                "cpc": item.get("cpc"),
                "low_bid": item.get("low_top_of_page_bid"),
                "high_bid": item.get("high_top_of_page_bid"),
                "monthly_searches": item.get("monthly_searches"),
            })

    return results


def fetch_amazon_volume(keywords: list[str], location: str = "US") -> list[dict]:
    """Fetch search volume from Amazon Bulk Search Volume API."""
    auth = get_auth()
    location_code = LOCATION_CODES.get(location, 2840)
    language_code = LANGUAGE_CODES.get(location, "en")

    payload = [{
        "keywords": keywords,
        "location_code": location_code,
        "language_code": language_code,
    }]

    url = f"{API_BASE}/dataforseo_labs/amazon/bulk_search_volume/live"
    resp = requests.post(url, auth=auth, json=payload, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    if data.get("status_code") != 20000:
        print(f"Amazon API error: {data.get('status_message')}")
        return []

    results = []
    for task in data.get("tasks", []):
        if task.get("status_code") != 20000:
            print(f"  Task error: {task.get('status_message')}")
            continue
        for res in (task.get("result") or []):
            for item in (res.get("items") or []):
                results.append({
                    "keyword": item.get("keyword"),
                    "channel": "Amazon",
                    "search_volume": item.get("search_volume"),
                })

    return results


def fetch_google_volume_historical(keywords: list, location: str = "US",
                                    date_from: str = "2024-01-01", date_to: str = None) -> dict:
    """Fetch monthly historical search volume from Google Ads API.
    Returns: {keyword_lower: {"avg": int, "cpc": float, "monthly": {(year, month): int}}}
    """
    auth = get_auth()
    location_code = LOCATION_CODES.get(location, 2840)
    language_code = LANGUAGE_CODES.get(location, "en")
    if date_to is None:
        date_to = date.today().strftime("%Y-%m-%d")

    payload = [{
        "keywords": keywords,
        "location_code": location_code,
        "language_code": language_code,
        "date_from": date_from,
        "date_to": date_to,
    }]

    url = f"{API_BASE}/keywords_data/google_ads/search_volume/live"
    resp = requests.post(url, auth=auth, json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()

    results = {}
    for task in data.get("tasks", []):
        if task.get("status_code") != 20000:
            print(f"  Task error: {task.get('status_message')}")
            continue
        for item in (task.get("result") or []):
            kw = (item.get("keyword") or "").lower()
            monthly_raw = item.get("monthly_searches") or []
            results[kw] = {
                "avg": item.get("search_volume"),
                "cpc": item.get("cpc"),
                "monthly": {(m["year"], m["month"]): m["search_volume"] for m in monthly_raw},
            }
    return results


def fetch_gsc_monthly(site_url: str, start_date: str, end_date: str,
                       service_account_path: str = None) -> dict:
    """Fetch monthly impressions from GSC for all queries (full date range, one call).
    Returns: {query_lower: {(year, month): impressions}}
    """
    try:
        from google.oauth2 import service_account as sa_module
        from googleapiclient.discovery import build
    except ImportError:
        print("ERROR: Install google-auth and google-api-python-client")
        return {}

    sa_path = service_account_path or os.getenv(
        "GOOGLE_SERVICE_ACCOUNT_PATH", "credentials/google_service_account.json"
    )
    scopes = ["https://www.googleapis.com/auth/webmasters.readonly"]
    creds = sa_module.Credentials.from_service_account_file(sa_path, scopes=scopes)
    service = build("searchconsole", "v1", credentials=creds)

    results = {}
    start_row = 0
    row_limit = 25000

    while True:
        body = {
            "startDate": start_date,
            "endDate": end_date,
            "dimensions": ["query", "date"],
            "rowLimit": row_limit,
            "startRow": start_row,
        }
        resp = service.searchanalytics().query(siteUrl=site_url, body=body).execute()
        rows = resp.get("rows", [])
        if not rows:
            break
        for row in rows:
            query = row["keys"][0].lower()
            date_str = row["keys"][1]  # "2024-01-15"
            yr = int(date_str[:4])
            mo = int(date_str[5:7])
            impressions = int(row.get("impressions", 0))
            if query not in results:
                results[query] = {}
            ym = (yr, mo)
            results[query][ym] = results[query].get(ym, 0) + impressions
        if len(rows) < row_limit:
            break
        start_row += row_limit

    return results


def test_connection():
    """Test API connection by checking account balance."""
    auth = get_auth()
    url = f"{API_BASE}/appendix/user_data"
    resp = requests.get(url, auth=auth, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    if data.get("status_code") != 20000:
        print(f"Connection FAILED: {data.get('status_message')}")
        return False

    for task in data.get("tasks", []):
        result = task.get("result", [{}])
        if result:
            info = result[0]
            print(f"Connection OK!")
            print(f"  Login: {info.get('login')}")
            print(f"  Balance: ${info.get('money', {}).get('balance', 'N/A')}")
            print(f"  Total spent: ${info.get('money', {}).get('total_spent', 'N/A')}")
            return True

    print("Connection OK but no account data returned.")
    return True


def save_to_excel(google_results: list[dict], amazon_results: list[dict], output_path: str):
    """Save results to Excel with separate sheets per channel."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # Summary sheet — side-by-side comparison
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Build lookup dicts
    google_map = {r["keyword"]: r for r in google_results}
    amazon_map = {r["keyword"]: r for r in amazon_results}
    all_keywords = list(dict.fromkeys(
        [r["keyword"] for r in google_results] + [r["keyword"] for r in amazon_results]
    ))

    headers = ["Keyword", "Google Volume", "Amazon Volume", "Google CPC", "Google Competition"]
    ws_summary.append(headers)
    for kw in all_keywords:
        g = google_map.get(kw, {})
        a = amazon_map.get(kw, {})
        ws_summary.append([
            kw,
            g.get("search_volume", "N/A"),
            a.get("search_volume", "N/A"),
            g.get("cpc", "N/A"),
            g.get("competition", "N/A"),
        ])

    # Google detail sheet
    if google_results:
        ws_google = wb.create_sheet("Google Detail")
        g_headers = ["Keyword", "Search Volume", "Competition", "Competition Index", "CPC", "Low Bid", "High Bid"]
        ws_google.append(g_headers)
        for r in google_results:
            ws_google.append([
                r["keyword"], r["search_volume"], r["competition"],
                r["competition_index"], r["cpc"], r["low_bid"], r["high_bid"],
            ])

    # Amazon detail sheet
    if amazon_results:
        ws_amazon = wb.create_sheet("Amazon Detail")
        a_headers = ["Keyword", "Search Volume"]
        ws_amazon.append(a_headers)
        for r in amazon_results:
            ws_amazon.append([r["keyword"], r["search_volume"]])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Fetch keyword search volume from DataForSEO")
    parser.add_argument("--test", action="store_true", help="Test API connection only")
    parser.add_argument("--keywords", type=str, help="Comma-separated keywords")
    parser.add_argument("--keywords-file", type=str, help="File with one keyword per line")
    parser.add_argument("--channel", type=str, default="both", choices=["google", "amazon", "both"],
                        help="Which channel to query (default: both)")
    parser.add_argument("--location", type=str, default="US", help="Location code (default: US)")
    parser.add_argument("--output", type=str, help="Output Excel path")
    parser.add_argument("--json", action="store_true", help="Output raw JSON to stdout")
    args = parser.parse_args()

    if args.test:
        test_connection()
        return

    # Gather keywords
    keywords = []
    if args.keywords:
        keywords = [k.strip() for k in args.keywords.split(",") if k.strip()]
    elif args.keywords_file:
        with open(args.keywords_file, "r", encoding="utf-8") as f:
            keywords = [line.strip() for line in f if line.strip()]

    if not keywords:
        print("ERROR: Provide --keywords or --keywords-file")
        sys.exit(1)

    if len(keywords) > 1000:
        print(f"WARNING: {len(keywords)} keywords provided, max 1000 per request. Truncating.")
        keywords = keywords[:1000]

    print(f"Fetching volume for {len(keywords)} keywords (channel={args.channel}, location={args.location})...")

    google_results = []
    amazon_results = []

    if args.channel in ("google", "both"):
        print("  Querying Google Ads...")
        google_results = fetch_google_volume(keywords, args.location)
        print(f"  Google: {len(google_results)} results")

    if args.channel in ("amazon", "both"):
        print("  Querying Amazon...")
        amazon_results = fetch_amazon_volume(keywords, args.location)
        print(f"  Amazon: {len(amazon_results)} results")

    # Output
    if args.json:
        print(json.dumps({"google": google_results, "amazon": amazon_results}, indent=2, ensure_ascii=False))
    else:
        from output_utils import get_output_path
        output_path = args.output or get_output_path("marketing", "keyword_volume")
        save_to_excel(google_results, amazon_results, output_path)

    # Print cost summary
    total_cost = 0
    if google_results:
        total_cost += 0.075  # Live mode per task
    if amazon_results:
        total_cost += 0.01 + len(amazon_results) * 0.0001
    print(f"\nEstimated cost: ~${total_cost:.4f}")


if __name__ == "__main__":
    main()
