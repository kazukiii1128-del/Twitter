"""
scrape_job_listings.py
Scrapes remote job listings from DailyRemote and exports to Excel.

Usage:
    python tools/scrape_job_listings.py

Output:
    .tmp/job_listings_YYYY-MM-DD_HHMMSS.xlsx
"""

import os
import re
import sys
import json
import requests

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font

# --- Config ---
BASE_URL = "https://dailyremote.com/remote-support-jobs"
PARAMS = "employmentType=full-time&benefits=maternity"
MAX_WORKERS = 10  # parallel workers for individual job page scrapes

load_dotenv()
FIRECRAWL_API_KEY = os.getenv("FIRECRAWL_API_KEY")
if not FIRECRAWL_API_KEY:
    raise ValueError("FIRECRAWL_API_KEY not found in .env")

FIRECRAWL_SCRAPE_URL = "https://api.firecrawl.dev/v1/scrape"
HEADERS = {
    "Authorization": f"Bearer {FIRECRAWL_API_KEY}",
    "Content-Type": "application/json",
}


# --- Firecrawl helpers ---

def scrape_page(url: str) -> str:
    """Scrape a URL via Firecrawl and return its markdown content."""
    payload = {"url": url, "formats": ["markdown"]}
    resp = requests.post(FIRECRAWL_SCRAPE_URL, headers=HEADERS, json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    return data.get("data", {}).get("markdown", "")


# --- Listing page parser ---

JOB_BLOCK_RE = re.compile(
    r"## \[(.+?)\]\((https://dailyremote\.com/remote-job/[^)]+)\)(.*?)(?=\n## \[|\Z)",
    re.DOTALL,
)
SALARY_RE = re.compile(r"💵\s*\$?([\d,.]+ ?-? ?[\d,.]*(?:\s*per\s*\w+)?)")
SALARY_RE2 = re.compile(r"💵\s*([\d,.]+\s*-\s*[\d,.]+(?:\s*per\s*\w+)?)")
EXPERIENCE_RE = re.compile(r"⭐\s*([\w\-+ ]+(?:yrs?|years?)[^\[]*)")
LOCATION_RE = re.compile(r"🌎\s*\n+\s*([^\n💵⭐💼\[]+)")
POSTED_RE = re.compile(r"·\s*(.+?)(?:\n|$)")
EMPLOYMENT_RE = re.compile(r"(Full Time|Part Time|Contract|Internship)")
TAGS_RE = re.compile(r"\[([a-z][a-z0-9\-]+)\]\(https://dailyremote\.com/remote-[a-z\-]+-jobs\)")
CATEGORY_RE = re.compile(r"💼\s*\[?([^\]\n]+)\]?")


def parse_listing_page(markdown: str, page_num: int) -> list[dict]:
    """Extract all job listings from a listing page's markdown."""
    jobs = []
    for match in JOB_BLOCK_RE.finditer(markdown):
        title = match.group(1).strip()
        url = match.group(2).strip()
        block = match.group(3)

        # Employment type
        emp_match = EMPLOYMENT_RE.search(block)
        employment_type = emp_match.group(1) if emp_match else ""

        # Posted time
        posted_match = POSTED_RE.search(block)
        posted = posted_match.group(1).strip() if posted_match else ""

        # Location (text after 🌎, before salary/experience markers)
        loc_match = LOCATION_RE.search(block)
        location = loc_match.group(1).strip() if loc_match else ""
        # Clean location: remove trailing emoji/symbols
        location = re.sub(r'[💵⭐💼].*', '', location).strip()

        # Salary
        sal_match = SALARY_RE.search(block) or SALARY_RE2.search(block)
        # Try simpler pattern: grab text between 💵 and next ⭐ or 💼 or newline
        sal_raw = re.search(r'💵\s*(.*?)(?=⭐|💼|\n\n)', block, re.DOTALL)
        salary = ""
        if sal_raw:
            salary = sal_raw.group(1).strip().replace("\n", " ")
            # Remove markdown link cruft
            salary = re.sub(r'\[.*?\]\(.*?\)', '', salary).strip()

        # Experience
        exp_match = EXPERIENCE_RE.search(block)
        experience = exp_match.group(1).strip() if exp_match else ""

        # Category
        cat_match = CATEGORY_RE.search(block)
        category = cat_match.group(1).strip() if cat_match else ""
        category = re.sub(r'\]\(.*', '', category).strip()

        # Description: first substantial paragraph (not a link line)
        desc = ""
        for line in block.splitlines():
            line = line.strip()
            if (
                len(line) > 80
                and not line.startswith("[")
                and not line.startswith("!")
                and "💵" not in line
                and "🌎" not in line
                and "⭐" not in line
                and "💼" not in line
            ):
                desc = line
                break

        # Tags: deduplicated, comma-separated
        tags = list(dict.fromkeys(TAGS_RE.findall(block)))
        tags_str = ", ".join(tags)

        # Apply URL
        apply_match = re.search(r'\[APPLY\]\((https://dailyremote\.com/remote-job/[^)]+)\)', block)
        apply_url = apply_match.group(1) if apply_match else url

        jobs.append({
            "title": title,
            "company": "",  # filled in Step 2
            "url": url,
            "employment_type": employment_type,
            "posted": posted,
            "location": location,
            "salary": salary,
            "experience": experience,
            "category": category,
            "description": desc,
            "tags": tags_str,
            "apply_url": apply_url,
            "page": page_num,
        })

    return jobs


# --- Individual job page: company name extractor ---

COMPANY_RE = re.compile(
    r'\*\*([^*]+)\*\*\]\(https://dailyremote\.com/remote-company/'
)


def fetch_company(job: dict) -> dict:
    """Scrape the individual job page and add the company name."""
    try:
        md = scrape_page(job["url"])
        match = COMPANY_RE.search(md)
        job["company"] = match.group(1).strip() if match else ""
        print(f"  [job] {job['title'][:60]} -> {job['company'] or '(no company)'}")
    except Exception as e:
        print(f"  [job] WARN: could not fetch {job['url']}: {e}")
        job["company"] = ""
    return job


# --- Excel writer ---

COLUMNS = [
    ("Job Title", "title"),
    ("Company", "company"),
    ("Job URL", "url"),
    ("Employment Type", "employment_type"),
    ("Posted", "posted"),
    ("Location", "location"),
    ("Salary Range", "salary"),
    ("Experience Level", "experience"),
    ("Category", "category"),
    ("Description", "description"),
    ("Skills / Tags", "tags"),
    ("Apply URL", "apply_url"),
]


def write_excel(jobs: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    # Header row
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for job in jobs:
        row = [job.get(col[1], "") for col in COLUMNS]
        ws.append(row)

    # Auto-width (approximate)
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[col_letter].width = max_len + 2

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# --- Main ---

def main():
    all_jobs = []
    seen_urls: set[str] = set()

    # Step 1: scrape listing pages
    print("=== Step 1: Scraping listing pages ===")
    page = 1
    MAX_PAGES = 50  # safety ceiling
    while page <= MAX_PAGES:
        url = f"{BASE_URL}?{PARAMS}&page={page}"
        print(f"  Scraping page {page}: {url}")
        try:
            md = scrape_page(url)
        except Exception as e:
            print(f"  ERROR on page {page}: {e}. Stopping pagination.")
            break
        jobs = parse_listing_page(md, page)

        # Detect duplicate pages (site loops after last real page)
        new_jobs = [j for j in jobs if j["url"] not in seen_urls]
        print(f"  Found {len(jobs)} jobs on page {page} ({len(new_jobs)} new)")
        if not new_jobs:
            print("  All jobs on this page already seen — reached end of listings.")
            break

        for j in new_jobs:
            seen_urls.add(j["url"])
        all_jobs.extend(new_jobs)
        page += 1

    print(f"\nTotal jobs found across all pages: {len(all_jobs)}")

    if not all_jobs:
        print("No jobs found. Exiting.")
        return

    # Step 2: fetch company names in parallel
    print(f"\n=== Step 2: Fetching company names ({len(all_jobs)} individual pages) ===")
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_company, job): job for job in all_jobs}
        enriched = []
        for future in as_completed(futures):
            enriched.append(future.result())
    all_jobs = enriched

    # Step 3: write Excel
    from output_utils import get_output_path
    output_path = get_output_path("misc", "job_listings")
    print(f"\n=== Step 3: Writing Excel to {output_path} ===")
    write_excel(all_jobs, output_path)
    print(f"Done. {len(all_jobs)} jobs saved to {output_path}")


if __name__ == "__main__":
    main()
