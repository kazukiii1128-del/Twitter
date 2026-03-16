"""
調査マン: 競合・参考ブランドのTwitter運用を週次で調査しTeamsに報告する。

Flow:
  Firecrawlで各ブランドの最新ツイートを収集
  → Claudeで投稿傾向・カテゴリ・戦略を分析
  → Teams Webhookに調査レポートを送信
"""

import os
import sys
import json
import time
import logging
import subprocess
from pathlib import Path
from datetime import datetime, timezone, timedelta

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).parent))

PROJECT_ROOT = Path(__file__).parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
FIRECRAWL_DIR = PROJECT_ROOT / ".firecrawl"
JST = timezone(timedelta(hours=9))
MODEL = "claude-sonnet-4-20250514"

# ── 調査対象ブランド ──────────────────────────────────────────────────────────

BRANDS = {
    "競合": [
        {"name": "ピジョン",      "handle": "pigeon_official_jp", "query": "ピジョン 育児 site:x.com"},
        {"name": "コンビ",        "handle": "combi_jp",           "query": "コンビ ベビー site:x.com"},
        {"name": "リッチェル",    "handle": "richell_jp",         "query": "リッチェル マグ site:x.com"},
        {"name": "NUKジャパン",   "handle": "nuk_japan",          "query": "NUK 赤ちゃん site:x.com"},
        {"name": "マンチキン",    "handle": "munchkin_japan",     "query": "マンチキン ベビー site:x.com"},
    ],
    "参考": [
        {"name": "アカチャンホンポ", "handle": "akachanhonpo",     "query": "アカチャンホンポ 育児 site:x.com"},
        {"name": "西松屋",          "handle": "nishimatsuya_com", "query": "西松屋 赤ちゃん site:x.com"},
        {"name": "BABYBJORN",       "handle": "babybjorn",        "query": "BabyBjorn baby site:x.com"},
    ],
}


# ── Firecrawl検索 ─────────────────────────────────────────────────────────────

def search_brand_tweets(brand: dict, limit: int = 10) -> list[dict]:
    """Firecrawlでブランドの最新ツイートを検索する。"""
    from firecrawl import FirecrawlApp

    api_key = os.getenv("FIRECRAWL_API_KEY")
    if not api_key:
        logger.warning("FIRECRAWL_API_KEY not set")
        return []

    try:
        app = FirecrawlApp(api_key=api_key)
        result = app.search(brand["query"], limit=limit)
        items = result.get("data", [])

        tweets = []
        for item in items:
            url = item.get("url", "")
            if "x.com" in url and "/status/" in url:
                tweets.append({
                    "url": url,
                    "title": item.get("title", ""),
                    "description": item.get("description", ""),
                })
        return tweets
    except Exception as e:
        logger.warning(f"Search failed for {brand['name']}: {e}")

    return []


# ── Claude分析 ────────────────────────────────────────────────────────────────

def analyze_brand(brand: dict, tweets: list[dict]) -> str:
    """Claudeでブランドのツイート傾向を分析する。"""
    import anthropic

    if not tweets:
        return "ツイートが見つかりませんでした。"

    tweet_texts = "\n".join([
        f"- {t.get('description', t.get('title', ''))[:100]}"
        for t in tweets[:8]
    ])

    prompt = f"""以下は「{brand['name']}」(@{brand['handle']})の最近のツイートです。

{tweet_texts}

以下の観点で簡潔に分析してください（日本語、各項目1〜2行）：
1. **投稿テーマ**: どんな内容が多いか
2. **トーン**: 硬い/柔らかい、感情的/情報的など
3. **ハッシュタグ戦略**: よく使うタグのパターン
4. **グロミミへの示唆**: 参考にできる点・差別化できる点

箇条書きで簡潔に。"""

    try:
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        resp = client.messages.create(
            model=MODEL,
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        return resp.content[0].text.strip()
    except Exception as e:
        logger.warning(f"Claude analysis failed for {brand['name']}: {e}")
        return "分析できませんでした。"


# ── Excel生成 ─────────────────────────────────────────────────────────────────

def create_research_excel(results: dict, date_str: str) -> Path:
    """調査結果をExcelファイルに出力する。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "競合調査"

    # ヘッダー
    headers = ["カテゴリ", "ブランド名", "ツイート件数", "分析結果", "サンプルURL①", "サンプルURL②", "サンプルURL③"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # データ行
    row = 2
    fill_comp = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    fill_ref  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for category, brand_results in results.items():
        fill = fill_comp if category == "競合" else fill_ref
        for brand_name, data in brand_results.items():
            urls = data.get("sample_urls", [])
            ws.cell(row=row, column=1, value=category).fill = fill
            ws.cell(row=row, column=2, value=brand_name).fill = fill
            ws.cell(row=row, column=3, value=data.get("tweet_count", 0)).fill = fill
            analysis_cell = ws.cell(row=row, column=4, value=data.get("analysis", ""))
            analysis_cell.fill = fill
            analysis_cell.alignment = Alignment(wrap_text=True)
            for i, url in enumerate(urls[:3]):
                ws.cell(row=row, column=5 + i, value=url).fill = fill
            row += 1

    # 列幅調整
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    for col in ["E", "F", "G"]:
        ws.column_dimensions[col].width = 40

    TMP_DIR.mkdir(parents=True, exist_ok=True)
    path = TMP_DIR / f"twitter_research_{date_str}.xlsx"
    wb.save(path)
    logger.info(f"Research Excel saved: {path}")
    return path


# ── Teams報告 ─────────────────────────────────────────────────────────────────

def send_report(results: dict) -> None:
    """調査レポートをTeamsに送信する。"""
    import requests

    webhook_url = os.getenv("TEAMS_WEBHOOK_URL") or os.getenv("TEAMS_MASTER_WEBHOOK_URL")
    if not webhook_url:
        logger.warning("TEAMS_WEBHOOK_URL not set")
        return

    today = datetime.now(JST).strftime("%Y-%m-%d（%a）")
    lines = [f"🔍 **調査マン週次レポート** {today}\n"]

    for category, brand_results in results.items():
        lines.append(f"## {category}ブランド")
        for brand_name, data in brand_results.items():
            tweet_count = data.get("tweet_count", 0)
            analysis = data.get("analysis", "")
            lines.append(f"\n### {brand_name}（直近{tweet_count}件）")
            lines.append(analysis)
        lines.append("")

    message = "\n".join(lines)

    try:
        resp = requests.post(webhook_url, json={"text": message}, timeout=15)
        resp.raise_for_status()
        logger.info("Research report sent to Teams")
    except Exception as e:
        logger.error(f"Teams send failed: {e}")


# ── メイン ────────────────────────────────────────────────────────────────────

def run_research() -> None:
    """全ブランドを調査してTeamsに報告する。"""
    logger.info("=== 調査マン START ===")
    results = {}

    for category, brands in BRANDS.items():
        results[category] = {}
        for brand in brands:
            logger.info(f"Searching: {brand['name']}")
            tweets = search_brand_tweets(brand, limit=10)
            logger.info(f"  Found {len(tweets)} tweets")

            analysis = analyze_brand(brand, tweets)

            results[category][brand["name"]] = {
                "tweet_count": len(tweets),
                "analysis": analysis,
                "sample_urls": [t["url"] for t in tweets[:3]],
            }

            time.sleep(3)  # Firecrawl rate limit

    # ローカル保存
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now(JST).strftime("%Y-%m-%d")
    out_path = TMP_DIR / f"twitter_research_{date_str}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved: {out_path}")

    send_report(results)

    # Excel生成 → Teamsにアップロード
    try:
        excel_path = create_research_excel(results, date_str)
        from teams_upload import upload_file
        upload_file(
            str(excel_path),
            notify=True,
            message=f"📊 {date_str} 競合Twitter調査レポート（Excelで確認できます）",
        )
        logger.info("Research Excel uploaded to Teams")
    except Exception as e:
        logger.warning(f"Excel upload failed: {e}")

    logger.info("=== 調査マン DONE ===")


if __name__ == "__main__":
    run_research()
