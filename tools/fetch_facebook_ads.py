"""
Facebook Ads 캠페인 상세 데이터 수집 툴
- 캠페인별 성과 + 광고 URL(링크) 포함
- 결과를 .tmp/facebook_ads.xlsx 로 저장
"""

import os
import json
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from dotenv import load_dotenv

load_dotenv()

ACCESS_TOKEN = os.getenv("META_ACCESS_TOKEN")
AD_ACCOUNT_ID = os.getenv("META_AD_ACCOUNT_ID")
API_VERSION = "v18.0"
BASE_URL = f"https://graph.facebook.com/{API_VERSION}"


def api_get(path, params):
    params["access_token"] = ACCESS_TOKEN
    url = f"{BASE_URL}{path}?{urllib.parse.urlencode(params)}"
    try:
        with urllib.request.urlopen(url) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        error_body = json.loads(e.read())
        raise Exception(f"API 오류: {error_body.get('error', {}).get('message', str(e))}")


def get_campaigns():
    """캠페인 목록 가져오기"""
    data = api_get(f"/{AD_ACCOUNT_ID}/campaigns", {
        "fields": "id,name,status,objective,daily_budget,lifetime_budget,start_time,stop_time",
        "limit": 100
    })
    return data.get("data", [])


def get_campaign_insights(campaign_id, date_preset="last_30d"):
    """캠페인 성과 지표"""
    data = api_get(f"/{campaign_id}/insights", {
        "fields": "impressions,clicks,spend,ctr,cpc,cpm,reach,frequency,actions",
        "date_preset": date_preset
    })
    result = data.get("data", [{}])
    return result[0] if result else {}


def get_ads_for_campaign(campaign_id):
    """캠페인 하위 광고 목록 + 랜딩 URL"""
    data = api_get(f"/{AD_ACCOUNT_ID}/ads", {
        "fields": "id,name,status,creative{id,name,object_url,link_url,call_to_action}",
        "filtering": json.dumps([{"field": "campaign.id", "operator": "EQUAL", "value": campaign_id}]),
        "limit": 100
    })
    ads = data.get("data", [])

    results = []
    for ad in ads:
        creative = ad.get("creative", {})
        link_url = (
            creative.get("link_url") or
            creative.get("object_url") or
            creative.get("call_to_action", {}).get("value", {}).get("link", "") or
            ""
        )
        results.append({
            "ad_id": ad.get("id"),
            "ad_name": ad.get("name"),
            "ad_status": ad.get("status"),
            "link_url": link_url,
        })
    return results


def extract_conversions(actions):
    """actions 배열에서 구매/전환 추출"""
    if not actions:
        return 0, 0
    purchases = 0
    purchase_value = 0
    for a in actions:
        if a.get("action_type") in ("purchase", "offsite_conversion.fb_pixel_purchase"):
            purchases += int(a.get("value", 0))
        if a.get("action_type") == "offsite_conversion.fb_pixel_purchase":
            purchase_value += float(a.get("value", 0))
    return purchases, purchase_value


def main():
    print(f"📊 Facebook Ads 데이터 수집 시작 | 계정: {AD_ACCOUNT_ID}\n")

    campaigns = get_campaigns()
    print(f"캠페인 {len(campaigns)}개 발견\n")

    rows = []
    for camp in campaigns:
        cid = camp["id"]
        cname = camp.get("name", "")
        status = camp.get("status", "")
        objective = camp.get("objective", "")

        print(f"  처리중: {cname}")

        insights = get_campaign_insights(cid)
        purchases, purchase_value = extract_conversions(insights.get("actions"))

        ads = get_ads_for_campaign(cid)
        # 광고가 여러 개면 URL들을 콤마로 합침
        links = ", ".join(filter(None, [a["link_url"] for a in ads]))
        ad_names = ", ".join([a["ad_name"] for a in ads])

        rows.append({
            "캠페인 ID": cid,
            "캠페인명": cname,
            "상태": status,
            "목표": objective,
            "광고 수": len(ads),
            "광고명": ad_names,
            "랜딩 URL": links,
            "노출": insights.get("impressions", 0),
            "클릭": insights.get("clicks", 0),
            "지출(USD)": insights.get("spend", 0),
            "CTR(%)": insights.get("ctr", 0),
            "CPC(USD)": insights.get("cpc", 0),
            "CPM(USD)": insights.get("cpm", 0),
            "도달": insights.get("reach", 0),
            "구매 수": purchases,
        })

    # 엑셀 저장
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("\nopenpyxl 없음 — pip install openpyxl 실행 후 재시도하세요")
        return

    os.makedirs(".tmp", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facebook Ads"

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        # 헤더 스타일
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1877F2")
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([row[h] for h in headers])

        # 컬럼 너비 자동조정
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    from output_utils import get_intermediate_path
    out_path = get_intermediate_path("", "facebook_ads.xlsx")
    wb.save(out_path)
    print(f"\n✅ 완료! 저장 위치: {out_path}")
    print(f"   총 캠페인: {len(rows)}개")


if __name__ == "__main__":
    main()