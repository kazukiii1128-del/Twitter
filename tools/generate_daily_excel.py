"""
WAT Tool: Generate daily Twitter plan as Excel file for Teams upload.

Creates an Excel with two sheets:
1. 내 트윗 (My Tweets): time, theme, JP text, KR text, char count, approval dropdown
2. 리플 계획 (Replies): time, target account, target tweet, draft reply JP/KR, approval dropdown

Supports:
- AM/PM split (slots parameter)
- Weekend mode (tweets only, no replies)
- Multi-date (target_date for Friday → Sat/Sun plans)

Usage:
    python tools/generate_daily_excel.py                    # generate from saved plan
    python tools/generate_daily_excel.py --generate         # generate fresh + create Excel
    python tools/generate_daily_excel.py --slots 9,11,13,15 # AM only
    python tools/generate_daily_excel.py --weekend          # tweets only (no replies)
"""

import sys
import json
import argparse
import logging
from pathlib import Path
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
def _plan_path(date_str: str = None) -> Path:
    """Get date-specific plan file path."""
    if not date_str:
        jst = timezone(timedelta(hours=9))
        date_str = datetime.now(jst).strftime("%Y-%m-%d")
    return TMP_DIR / f"daily_tweet_plan_{date_str}.json"

ALL_SLOTS = [10, 19]
AM_SLOTS = [10, 19]
PM_SLOTS = []

SLOT_INFO = {
    10: {"theme_ko": "오전 (공감/일상)", "theme_jp": "朝の育児あるある"},
    19: {"theme_ko": "저녁 (일상/계절)", "theme_jp": "日常エピソード"},
}

# ── Styles ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="맑은 고딕")
CELL_FONT = Font(size=10, name="맑은 고딕")
CELL_FONT_JP = Font(size=10, name="Yu Gothic")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
EVEN_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ALT_TEXT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
SLOT_HEADER_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")


def get_jst_now() -> datetime:
    jst = timezone(timedelta(hours=9))
    return datetime.now(jst)


def load_plan(plan_file: str = None) -> dict:
    path = Path(plan_file) if plan_file else _plan_path()
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _make_approval_dv() -> DataValidation:
    """Create a Confirmed/Declined dropdown validator."""
    dv = DataValidation(
        type="list",
        formula1='"Confirmed,Declined"',
        allow_blank=True,
    )
    dv.error = "Confirmed 또는 Declined만 선택 가능합니다"
    dv.errorTitle = "입력 오류"
    dv.prompt = "Confirmed = 승인, Declined = 거절 (대안 텍스트 작성)"
    dv.promptTitle = "승인 여부"
    return dv


def create_daily_excel(
    plan: dict,
    output_path: str = None,
    slots: list[int] = None,
    include_replies: bool = True,
    label: str = "",
) -> str:
    """Create the daily tweet plan Excel file.

    Args:
        plan: plan dict with slots data
        output_path: output file path (auto-generated if None)
        slots: which slots to include (default: all 8)
        include_replies: False for weekend mode (tweets only)
        label: optional label for filename (e.g., "AM", "PM", "weekend_sat")
    """
    date_str = plan.get("date", get_jst_now().strftime("%Y-%m-%d"))
    if slots is None:
        slots = ALL_SLOTS

    if output_path is None:
        TMP_DIR.mkdir(parents=True, exist_ok=True)
        suffix = f"_{label}" if label else ""
        output_path = str(TMP_DIR / f"tweet_plan_{date_str}{suffix}.xlsx")

    wb = Workbook()

    # ── Sheet 1: 내 트윗 ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "내 트윗"

    period = label.upper() if label else "ALL"
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"@grosmimi_japan 트윗 플랜 — {date_str} ({period})"
    ws1["A1"].font = Font(bold=True, size=14, name="맑은 고딕")
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:H2")
    ws1["A2"] = "승인: Confirmed 선택 | 거절: Declined 선택 후 G열에 대안 텍스트 작성"
    ws1["A2"].font = Font(size=9, italic=True, color="666666", name="맑은 고딕")

    headers1 = ["시간", "테마", "트윗 (JP)", "번역 (KR)", "글자수", "승인", "대안 텍스트", "메모"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    approval_dv1 = _make_approval_dv()
    ws1.add_data_validation(approval_dv1)

    for idx, slot in enumerate(slots):
        row = idx + 4
        slot_data = plan.get("slots", {}).get(str(slot), {})
        info = SLOT_INFO.get(slot, {})

        tweet_jp = slot_data.get("tweet_jp", "")
        tweet_ko = slot_data.get("tweet_ko", "")
        chars = slot_data.get("chars", 0)

        ws1.cell(row=row, column=1, value=f"{slot}:00").font = Font(bold=True, size=11, name="맑은 고딕")
        ws1.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws1.cell(row=row, column=2, value=info.get("theme_ko", "")).font = CELL_FONT

        cell_jp = ws1.cell(row=row, column=3, value=tweet_jp)
        cell_jp.font = CELL_FONT_JP
        cell_jp.alignment = WRAP_ALIGN

        cell_kr = ws1.cell(row=row, column=4, value=tweet_ko)
        cell_kr.font = CELL_FONT
        cell_kr.alignment = WRAP_ALIGN

        ws1.cell(row=row, column=5, value=f"{chars}/280").font = CELL_FONT
        ws1.cell(row=row, column=5).alignment = CENTER_ALIGN

        approval_cell = ws1.cell(row=row, column=6, value="")
        approval_cell.font = Font(bold=True, size=11, name="맑은 고딕")
        approval_cell.alignment = CENTER_ALIGN
        approval_dv1.add(approval_cell)

        alt_cell = ws1.cell(row=row, column=7, value="")
        alt_cell.font = CELL_FONT_JP
        alt_cell.alignment = WRAP_ALIGN
        alt_cell.fill = ALT_TEXT_FILL

        ws1.cell(row=row, column=8, value="").font = CELL_FONT

        for c in range(1, 9):
            ws1.cell(row=row, column=c).border = THIN_BORDER
            if idx % 2 == 1 and c not in (6, 7):
                ws1.cell(row=row, column=c).fill = EVEN_ROW_FILL

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 50
    ws1.column_dimensions["D"].width = 50
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 14
    ws1.column_dimensions["G"].width = 50
    ws1.column_dimensions["H"].width = 20

    for row in range(4, 4 + len(slots)):
        ws1.row_dimensions[row].height = 80

    # ── Sheet 2: 리플 계획 ────────────────────────────────────────────
    if include_replies:
        ws2 = wb.create_sheet("리플 계획")

        ws2.merge_cells("A1:H1")
        ws2["A1"] = f"@grosmimi_japan 리플 계획 — {date_str} ({period})"
        ws2["A1"].font = Font(bold=True, size=14, name="맑은 고딕")
        ws2["A1"].alignment = Alignment(horizontal="center")

        ws2.merge_cells("A2:H2")
        ws2["A2"] = "승인: Confirmed 선택 | 거절: Declined 선택 후 G열에 대안 리플 작성"
        ws2["A2"].font = Font(size=9, italic=True, color="666666", name="맑은 고딕")

        headers2 = ["시간", "타겟 계정", "타겟 트윗 (요약)", "리플 (JP)", "번역 (KR)", "승인", "대안 텍스트", "메모"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        approval_dv2 = _make_approval_dv()
        ws2.add_data_validation(approval_dv2)

        current_row = 4
        total_replies = 0

        for slot in slots:
            slot_data = plan.get("slots", {}).get(str(slot), {})
            replies = slot_data.get("replies", [])

            if not replies:
                # No replies for this slot — show info row
                ws2.cell(row=current_row, column=1, value=f"{slot}:00").font = Font(bold=True, size=10, name="맑은 고딕")
                ws2.cell(row=current_row, column=1).alignment = CENTER_ALIGN
                ws2.cell(row=current_row, column=2, value="—").font = CELL_FONT
                ws2.cell(row=current_row, column=3, value="(리플 없음)").font = Font(size=10, italic=True, color="999999", name="맑은 고딕")
                for c in range(1, 9):
                    ws2.cell(row=current_row, column=c).border = THIN_BORDER
                    ws2.cell(row=current_row, column=c).fill = SLOT_HEADER_FILL
                current_row += 1
                continue

            for r_idx, reply in enumerate(replies):
                ws2.cell(row=current_row, column=1, value=f"{slot}:00").font = Font(bold=True, size=10, name="맑은 고딕")
                ws2.cell(row=current_row, column=1).alignment = CENTER_ALIGN

                ws2.cell(row=current_row, column=2, value=f"@{reply.get('target_username', '?')}").font = CELL_FONT

                target_text = reply.get("target_text", "")[:100]
                ws2.cell(row=current_row, column=3, value=target_text).font = CELL_FONT_JP
                ws2.cell(row=current_row, column=3).alignment = WRAP_ALIGN

                reply_jp = reply.get("reply_jp", "")
                ws2.cell(row=current_row, column=4, value=reply_jp).font = CELL_FONT_JP
                ws2.cell(row=current_row, column=4).alignment = WRAP_ALIGN

                reply_ko = reply.get("reply_ko", "")
                ws2.cell(row=current_row, column=5, value=reply_ko).font = CELL_FONT
                ws2.cell(row=current_row, column=5).alignment = WRAP_ALIGN

                approval_cell2 = ws2.cell(row=current_row, column=6, value="")
                approval_cell2.font = Font(bold=True, size=11, name="맑은 고딕")
                approval_cell2.alignment = CENTER_ALIGN
                approval_dv2.add(approval_cell2)

                alt_cell2 = ws2.cell(row=current_row, column=7, value="")
                alt_cell2.font = CELL_FONT_JP
                alt_cell2.alignment = WRAP_ALIGN
                alt_cell2.fill = ALT_TEXT_FILL

                ws2.cell(row=current_row, column=8, value="").font = CELL_FONT

                for c in range(1, 9):
                    ws2.cell(row=current_row, column=c).border = THIN_BORDER
                    if r_idx % 2 == 1 and c not in (6, 7):
                        ws2.cell(row=current_row, column=c).fill = EVEN_ROW_FILL

                ws2.row_dimensions[current_row].height = 60
                current_row += 1
                total_replies += 1

        # Summary row
        summary_row = current_row + 1
        ws2.cell(row=summary_row, column=1, value="합계").font = Font(bold=True, name="맑은 고딕")
        ws2.cell(row=summary_row, column=2, value=f"트윗 {len(slots)} + 리플 {total_replies} = {len(slots) + total_replies}개").font = Font(bold=True, name="맑은 고딕")

        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 40
        ws2.column_dimensions["D"].width = 45
        ws2.column_dimensions["E"].width = 45
        ws2.column_dimensions["F"].width = 14
        ws2.column_dimensions["G"].width = 45
        ws2.column_dimensions["H"].width = 15

    # ── Save ──────────────────────────────────────────────────────────
    wb.save(output_path)
    logger.info(f"Excel saved: {output_path}")
    return output_path


def create_weekly_excel(
    weekly_plans: dict,
    output_path: str = None,
    slots: list = None,
    include_replies: bool = True,
) -> str:
    """Create one Excel file with 7 sheets (one per day).

    Args:
        weekly_plans: dict mapping date_str → plan_dict (e.g., {"2026-03-06": {...}, ...})
        output_path: output file path (auto-generated if None)
        slots: which slots to include (default: ALL_SLOTS)
        include_replies: False for weekend mode
    Returns: output file path
    """
    if slots is None:
        slots = ALL_SLOTS

    # Sort dates
    sorted_dates = sorted(weekly_plans.keys())
    if not sorted_dates:
        raise ValueError("weekly_plans is empty")

    start_date = sorted_dates[0]
    if output_path is None:
        TMP_DIR.mkdir(parents=True, exist_ok=True)
        output_path = str(TMP_DIR / f"tweet_plan_weekly_{start_date}.xlsx")

    day_names_jp = ["月", "火", "水", "木", "金", "土", "日"]
    wb = Workbook()
    first_sheet = True

    for date_str in sorted_dates:
        plan = weekly_plans[date_str]
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            day_jp = day_names_jp[dt.weekday()]
            sheet_name = f"{dt.strftime('%m-%d')}({day_jp})"
        except Exception:
            sheet_name = date_str

        if first_sheet:
            ws1 = wb.active
            ws1.title = sheet_name
            first_sheet = False
        else:
            ws1 = wb.create_sheet(sheet_name)

        period = sheet_name
        ws1.merge_cells("A1:H1")
        ws1["A1"] = f"@grosmimi_japan トゥウィートプラン — {date_str} ({period})"
        ws1["A1"].font = Font(bold=True, size=14, name="맑은 고딕")
        ws1["A1"].alignment = Alignment(horizontal="center")

        ws1.merge_cells("A2:H2")
        ws1["A2"] = "승인: Confirmed 선택 | 거절: Declined 선택 후 G열에 대안 텍스트 작성"
        ws1["A2"].font = Font(size=9, italic=True, color="666666", name="맑은 고딕")

        headers1 = ["시간", "테마", "트윗 (JP)", "번역 (KR)", "글자수", "승인", "대안 텍스트", "메모"]
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=3, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        approval_dv = _make_approval_dv()
        ws1.add_data_validation(approval_dv)

        for idx, slot in enumerate(slots):
            row = idx + 4
            slot_data = plan.get("slots", {}).get(str(slot), {})
            info = SLOT_INFO.get(slot, {})

            tweet_jp = slot_data.get("tweet_jp", "")
            tweet_ko = slot_data.get("tweet_ko", "")
            chars = slot_data.get("chars", 0)

            ws1.cell(row=row, column=1, value=f"{slot}:00").font = Font(bold=True, size=11, name="맑은 고딕")
            ws1.cell(row=row, column=1).alignment = CENTER_ALIGN
            ws1.cell(row=row, column=2, value=info.get("theme_ko", "")).font = CELL_FONT

            cell_jp = ws1.cell(row=row, column=3, value=tweet_jp)
            cell_jp.font = CELL_FONT_JP
            cell_jp.alignment = WRAP_ALIGN

            cell_kr = ws1.cell(row=row, column=4, value=tweet_ko)
            cell_kr.font = CELL_FONT
            cell_kr.alignment = WRAP_ALIGN

            ws1.cell(row=row, column=5, value=f"{chars}/280").font = CELL_FONT
            ws1.cell(row=row, column=5).alignment = CENTER_ALIGN

            approval_cell = ws1.cell(row=row, column=6, value="")
            approval_cell.font = Font(bold=True, size=11, name="맑은 고딕")
            approval_cell.alignment = CENTER_ALIGN
            approval_dv.add(approval_cell)

            alt_cell = ws1.cell(row=row, column=7, value="")
            alt_cell.font = CELL_FONT_JP
            alt_cell.alignment = WRAP_ALIGN
            alt_cell.fill = ALT_TEXT_FILL

            ws1.cell(row=row, column=8, value="").font = CELL_FONT

            for c in range(1, 9):
                ws1.cell(row=row, column=c).border = THIN_BORDER
                if idx % 2 == 1 and c not in (6, 7):
                    ws1.cell(row=row, column=c).fill = EVEN_ROW_FILL

        ws1.column_dimensions["A"].width = 8
        ws1.column_dimensions["B"].width = 20
        ws1.column_dimensions["C"].width = 50
        ws1.column_dimensions["D"].width = 50
        ws1.column_dimensions["E"].width = 10
        ws1.column_dimensions["F"].width = 14
        ws1.column_dimensions["G"].width = 50
        ws1.column_dimensions["H"].width = 20

        for row in range(4, 4 + len(slots)):
            ws1.row_dimensions[row].height = 80

        # ── Reply sheet tab (append after tweet tab for this day) ─────
        if include_replies:
            reply_sheet_name = f"{sheet_name}_リプ"
            ws2 = wb.create_sheet(reply_sheet_name)

            ws2.merge_cells("A1:H1")
            ws2["A1"] = f"@grosmimi_japan リプ計画 — {date_str} ({period})"
            ws2["A1"].font = Font(bold=True, size=14, name="맑은 고딕")
            ws2["A1"].alignment = Alignment(horizontal="center")

            ws2.merge_cells("A2:H2")
            ws2["A2"] = "승인: Confirmed 선택 | 거절: Declined 선택 후 G열에 대안 리플 작성"
            ws2["A2"].font = Font(size=9, italic=True, color="666666", name="맑은 고딕")

            headers2 = ["시간", "타겟 계정", "타겟 트윗 (요약)", "리플 (JP)", "번역 (KR)", "승인", "대안 텍스트", "메모"]
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER

            approval_dv2 = _make_approval_dv()
            ws2.add_data_validation(approval_dv2)

            current_row = 4
            total_replies = 0

            for slot in slots:
                slot_data = plan.get("slots", {}).get(str(slot), {})
                replies = slot_data.get("replies", [])

                if not replies:
                    ws2.cell(row=current_row, column=1, value=f"{slot}:00").font = Font(bold=True, size=10, name="맑은 고딕")
                    ws2.cell(row=current_row, column=1).alignment = CENTER_ALIGN
                    ws2.cell(row=current_row, column=2, value="—").font = CELL_FONT
                    ws2.cell(row=current_row, column=3, value="(리플 없음)").font = Font(size=10, italic=True, color="999999", name="맑은 고딕")
                    for c in range(1, 9):
                        ws2.cell(row=current_row, column=c).border = THIN_BORDER
                        ws2.cell(row=current_row, column=c).fill = SLOT_HEADER_FILL
                    current_row += 1
                    continue

                for r_idx, reply in enumerate(replies):
                    ws2.cell(row=current_row, column=1, value=f"{slot}:00").font = Font(bold=True, size=10, name="맑은 고딕")
                    ws2.cell(row=current_row, column=1).alignment = CENTER_ALIGN
                    ws2.cell(row=current_row, column=2, value=f"@{reply.get('target_username', '?')}").font = CELL_FONT
                    target_text = reply.get("target_text", "")[:100]
                    ws2.cell(row=current_row, column=3, value=target_text).font = CELL_FONT_JP
                    ws2.cell(row=current_row, column=3).alignment = WRAP_ALIGN
                    reply_jp = reply.get("reply_jp", "")
                    ws2.cell(row=current_row, column=4, value=reply_jp).font = CELL_FONT_JP
                    ws2.cell(row=current_row, column=4).alignment = WRAP_ALIGN
                    reply_ko = reply.get("reply_ko", "")
                    ws2.cell(row=current_row, column=5, value=reply_ko).font = CELL_FONT
                    ws2.cell(row=current_row, column=5).alignment = WRAP_ALIGN

                    approval_cell2 = ws2.cell(row=current_row, column=6, value="")
                    approval_cell2.font = Font(bold=True, size=11, name="맑은 고딕")
                    approval_cell2.alignment = CENTER_ALIGN
                    approval_dv2.add(approval_cell2)

                    alt_cell2 = ws2.cell(row=current_row, column=7, value="")
                    alt_cell2.font = CELL_FONT_JP
                    alt_cell2.alignment = WRAP_ALIGN
                    alt_cell2.fill = ALT_TEXT_FILL

                    ws2.cell(row=current_row, column=8, value="").font = CELL_FONT

                    for c in range(1, 9):
                        ws2.cell(row=current_row, column=c).border = THIN_BORDER
                        if r_idx % 2 == 1 and c not in (6, 7):
                            ws2.cell(row=current_row, column=c).fill = EVEN_ROW_FILL

                    ws2.row_dimensions[current_row].height = 60
                    current_row += 1
                    total_replies += 1

            summary_row = current_row + 1
            ws2.cell(row=summary_row, column=1, value="합계").font = Font(bold=True, name="맑은 고딕")
            ws2.cell(row=summary_row, column=2, value=f"트윗 {len(slots)} + 리플 {total_replies} = {len(slots) + total_replies}개").font = Font(bold=True, name="맑은 고딕")

            ws2.column_dimensions["A"].width = 8
            ws2.column_dimensions["B"].width = 16
            ws2.column_dimensions["C"].width = 40
            ws2.column_dimensions["D"].width = 45
            ws2.column_dimensions["E"].width = 45
            ws2.column_dimensions["F"].width = 14
            ws2.column_dimensions["G"].width = 45
            ws2.column_dimensions["H"].width = 15

    wb.save(output_path)
    logger.info(f"Weekly Excel saved: {output_path} ({len(sorted_dates)} days)")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate daily tweet plan Excel")
    parser.add_argument("--generate", action="store_true", help="Generate fresh plan first")
    parser.add_argument("--output", type=str, help="Output file path")
    parser.add_argument("--slots", type=str, help="Comma-separated slots (e.g., 9,11,13,15)")
    parser.add_argument("--weekend", action="store_true", help="Weekend mode (no replies)")
    args = parser.parse_args()

    if args.generate:
        from teams_dashboard import generate_daily_plan
        plan = generate_daily_plan()
    else:
        plan = load_plan()
        if not plan:
            logger.error("No plan file found. Run with --generate or run teams_dashboard.py first")
            sys.exit(1)

    target_slots = [int(s.strip()) for s in args.slots.split(",")] if args.slots else None
    label = "weekend" if args.weekend else ""

    output = create_daily_excel(
        plan,
        output_path=args.output,
        slots=target_slots,
        include_replies=not args.weekend,
        label=label,
    )
    print(f"\nExcel file: {output}")
