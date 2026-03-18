"""
ハッシュタグ調査マン: 毎週金曜 09:00 JSTに日本の人気ハッシュタグを調査してTeamsに報告。

調査内容:
  - 全体トレンド: 日本で今週バズったハッシュタグ
  - ベビー・育児関連: ターゲット層（育児ママ）が使うハッシュタグ
  - グロミミ活用提案: 投稿に使えるタグをピックアップ
"""

import os
import sys
import json
import time
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).parent))

PROJECT_ROOT = Path(__file__).parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
JST = timezone(timedelta(hours=9))
MODEL = "claude-sonnet-4-20250514"

# ── 検索クエリ ────────────────────────────────────────────────────────────────

SEARCH_QUERIES = {
    "全体トレンド": [
        "日本 トレンド ハッシュタグ 今週 site:x.com",
        "日本 バズ ハッシュタグ ランキング site:x.com",
        "日本 人気 ハッシュタグ 2026 site:x.com",
    ],
    "ベビー・育児": [
        "育児 ハッシュタグ ママ site:x.com",
        "赤ちゃん 1歳 ハッシュタグ 育児ママ site:x.com",
        "イヤイヤ期 幼児食 ハッシュタグ site:x.com",
        "育児あるある ハッシュタグ ママ友 site:x.com",
        "ベビーグッズ コップ飲み ハッシュタグ site:x.com",
    ],
}


# ── Firecrawl検索 ─────────────────────────────────────────────────────────────

def search_hashtags(query: str, limit: int = 10) -> list[dict]:
    """Firecrawlでハッシュタグ関連ツイートを検索する。"""
    from firecrawl import FirecrawlApp

    api_key = os.getenv("FIRECRAWL_API_KEY")
    if not api_key:
        logger.warning("FIRECRAWL_API_KEY not set")
        return []

    try:
        app = FirecrawlApp(api_key=api_key)
        result = app.search(query, limit=limit)

        if hasattr(result, "web"):
            items = result.web or []
        elif hasattr(result, "data"):
            items = result.data or []
        else:
            items = result.get("web", result.get("data", [])) if isinstance(result, dict) else []

        results = []
        for item in (items if isinstance(items, list) else []):
            results.append({
                "url":         item.get("url", "")         if isinstance(item, dict) else getattr(item, "url", ""),
                "title":       item.get("title", "")       if isinstance(item, dict) else getattr(item, "title", ""),
                "description": item.get("description", "") if isinstance(item, dict) else getattr(item, "description", ""),
            })
        return results
    except Exception as e:
        logger.warning(f"Search failed for '{query}': {e}")
        return []


# ── Claude分析 ────────────────────────────────────────────────────────────────

def analyze_hashtags(category: str, raw_results: list[dict]) -> dict:
    """Claudeでハッシュタグを抽出・分析する。"""
    import anthropic

    if not raw_results:
        return {"hashtags": [], "summary": "データが取得できませんでした。"}

    texts = "\n".join([
        f"- {r.get('description', r.get('title', ''))[:120]}"
        for r in raw_results[:15]
    ])

    if category == "全体トレンド":
        instruction = """
以下のツイートやWebコンテンツから、**日本で今週人気・バズっているハッシュタグ**を抽出してください。

抽出結果を以下の形式で出力:
【人気ハッシュタグ TOP10】
1. #〇〇〇 — 使われている文脈・トレンドの理由（1行）
2. ...

【今週のトレンドまとめ】（2〜3行で要約）
"""
    else:
        instruction = """
以下のツイートやWebコンテンツから、**育児ママが使う人気ハッシュタグ**を抽出してください。

抽出結果を以下の形式で出力:
【育児・ベビー系 人気ハッシュタグ TOP10】
1. #〇〇〇 — 使われる場面・ターゲット層（1行）
2. ...

【グロミミ投稿への活用提案】（どのタグをどのコンテンツカテゴリに使うか、2〜3行）
"""

    prompt = f"""以下は最近のTwitter/X上のコンテンツです。

{texts}

{instruction}"""

    try:
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        resp = client.messages.create(
            model=MODEL,
            max_tokens=600,
            messages=[{"role": "user", "content": prompt}],
        )
        analysis_text = resp.content[0].text.strip()

        # ハッシュタグを抽出してリスト化
        import re
        hashtags = re.findall(r"#[\w\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]+", analysis_text)

        return {
            "hashtags": list(dict.fromkeys(hashtags)),  # 重複除去・順序保持
            "summary":  analysis_text,
        }
    except Exception as e:
        logger.warning(f"Claude analysis failed: {e}")
        return {"hashtags": [], "summary": "分析できませんでした。"}


# ── Excel生成 ─────────────────────────────────────────────────────────────────

def create_hashtag_excel(results: dict, date_str: str) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ハッシュタグ調査"

    headers = ["カテゴリ", "ハッシュタグ", "分析・活用提案"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fill_map = {
        "全体トレンド":   PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"),
        "ベビー・育児":   PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    }

    row = 2
    for category, data in results.items():
        fill = fill_map.get(category, PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"))
        hashtags_str = "  ".join(data.get("hashtags", []))
        ws.cell(row=row, column=1, value=category).fill = fill
        tag_cell = ws.cell(row=row, column=2, value=hashtags_str)
        tag_cell.fill = fill
        tag_cell.alignment = Alignment(wrap_text=True)
        summary_cell = ws.cell(row=row, column=3, value=data.get("summary", ""))
        summary_cell.fill = fill
        summary_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 120
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 70

    TMP_DIR.mkdir(parents=True, exist_ok=True)
    path = TMP_DIR / f"twitter_hashtag_{date_str}.xlsx"
    wb.save(path)
    logger.info(f"Hashtag Excel saved: {path}")
    return path


# ── Teams報告 ─────────────────────────────────────────────────────────────────

def send_report(results: dict) -> None:
    import requests

    webhook_url = os.getenv("TEAMS_WEBHOOK_URL") or os.getenv("TEAMS_MASTER_WEBHOOK_URL")
    if not webhook_url:
        logger.warning("TEAMS_WEBHOOK_URL not set")
        return

    today = datetime.now(JST).strftime("%Y-%m-%d（%a）")
    lines = [f"#️⃣ **ハッシュタグ調査マン週次レポート** {today}\n"]

    for category, data in results.items():
        hashtags = data.get("hashtags", [])
        summary  = data.get("summary", "")
        lines.append(f"## {category}")
        if hashtags:
            lines.append("**抽出タグ:** " + "  ".join(hashtags[:10]))
        lines.append(summary)
        lines.append("")

    message = "\n".join(lines)

    try:
        resp = requests.post(webhook_url, json={"text": message}, timeout=15)
        resp.raise_for_status()
        logger.info("Hashtag report sent to Teams")
    except Exception as e:
        logger.error(f"Teams send failed: {e}")


# ── メイン ────────────────────────────────────────────────────────────────────

def run_hashtag_research() -> None:
    logger.info("=== ハッシュタグ調査マン START ===")
    results = {}

    for category, queries in SEARCH_QUERIES.items():
        logger.info(f"Searching: {category}")
        all_items = []
        for q in queries:
            items = search_hashtags(q, limit=8)
            all_items.extend(items)
            time.sleep(2)

        logger.info(f"  Found {len(all_items)} results")
        results[category] = analyze_hashtags(category, all_items)
        time.sleep(2)

    # ローカル保存
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now(JST).strftime("%Y-%m-%d")
    out_path = TMP_DIR / f"twitter_hashtag_{date_str}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved: {out_path}")

    send_report(results)

    # Excel → Teams アップロード
    try:
        excel_path = create_hashtag_excel(results, date_str)
        from teams_upload import upload_file
        upload_file(
            str(excel_path),
            notify=True,
            message=f"#️⃣ {date_str} ハッシュタグ調査レポート（Excelで確認できます）",
        )
        logger.info("Hashtag Excel uploaded to Teams")
    except Exception as e:
        logger.warning(f"Excel upload failed: {e}")

    logger.info("=== ハッシュタグ調査マン DONE ===")


if __name__ == "__main__":
    run_hashtag_research()
