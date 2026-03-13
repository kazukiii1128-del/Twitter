"""Polar Dashboard Builder — reads Polar MCP JSON exports, generates Polar_Dashboard.xlsx.

Usage:
    python tools/polar_dashboard_builder.py

Reads JSON files from .tmp/polar_data/ and generates Data Storage/Polar_Dashboard.xlsx
with 6 tabs: Sales, AdSpend, Organic, Margin, Product, Data_Status.
"""
import json, os, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.join(SCRIPT_DIR, "..")
DATA_DIR = os.path.join(ROOT_DIR, ".tmp", "polar_data")
from output_utils import get_output_path
OUTPUT = get_output_path("polar", "dashboard")

# ── Constants ────────────────────────────────────────────────────────────────
BRANDS = [
    "Grosmimi", "Alpremio", "Easy Shower", "Comme Moi", "BabyRabbit",
    "Naeiae", "Bamboobebe", "Hattung", "CHA&MOM", "Beemymagic", "Nature Love Mere",
]
CHANNELS = [
    "D2C", "Target+", "B2B", "TikTok", "FBM", "PR",
    "Amazon - Grosmimi USA", "Amazon - Fleeters", "Amazon - Orbitool",
    "MCF/Removal Order",
]
MONTHS = ["2026-01-01", "2026-02-01"]
MONTH_LABELS = ["Jan 2026", "Feb 2026 (partial)"]
AD_PLATFORMS = ["Amazon Ads", "Facebook Ads", "Google Ads", "TikTok Ads"]

# ── Styles ───────────────────────────────────────────────────────────────────
BLUE = PatternFill("solid", fgColor="4472C4")
DARK = PatternFill("solid", fgColor="2F5496")
ORANGE = PatternFill("solid", fgColor="ED7D31")
GREEN = PatternFill("solid", fgColor="548235")
YELLOW = PatternFill("solid", fgColor="FFFF00")
GRAY = PatternFill("solid", fgColor="808080")
WF = Font(bold=True, size=10, color="FFFFFF")
HF = Font(bold=True, size=10)
BF = Font(bold=True)
NUM = "#,##0.00"
INT = "#,##0"
PCT = "0.0%"
DOL = "$#,##0.00"
THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)

# ── Data Loading ─────────────────────────────────────────────────────────────
def load(name):
    with open(os.path.join(DATA_DIR, name), encoding="utf-8") as f:
        return json.load(f)

# ── Campaign Parsing ─────────────────────────────────────────────────────────
# Priority-ordered rules: first match wins
AMAZON_BRAND_RULES = [
    ("cha&mom", "CHA&MOM"), ("naeiae", "Naeiae"),
    ("alpremio", "Alpremio"), ("comme", "Comme Moi"),
]
FB_BRAND_RULES = [
    ("alpremio", "Alpremio"), ("naeiae", "Naeiae"),
    ("cha&mom", "CHA&MOM"), ("love&care", "CHA&MOM"),
    ("| cm |", "CHA&MOM"), ("| cm_", "CHA&MOM"), ("_cm_", "CHA&MOM"),
    ("| gm |", "Grosmimi"), ("_gm_", "Grosmimi"),
    ("grosmimi", "Grosmimi"), ("dental mom", "Grosmimi"),
    ("dentalmom", "Grosmimi"), ("livfuselli", "Grosmimi"),
    ("tumbler", "Grosmimi"), ("stainless", "Grosmimi"),
    ("sls cup", "Grosmimi"), ("sls", "Grosmimi"),
    ("laurence", "Grosmimi"), ("lauren", "Grosmimi"),
    ("asc campaign", "Grosmimi"),
]
FB_LANDING_RULES = [
    ("amz_traffic", "Amazon"), ("shopify", "Shopify"), ("target", "Target+"),
]

def parse_brand(campaign, platform):
    c = campaign.lower()
    if platform == "Amazon Ads":
        for kw, brand in AMAZON_BRAND_RULES:
            if kw in c:
                return brand
        return "Grosmimi"
    elif platform == "Facebook Ads":
        for kw, brand in FB_BRAND_RULES:
            if kw in c:
                return brand
        return "Other"
    elif platform == "Google Ads":
        return "Grosmimi"
    return "Other"

def parse_landing(campaign, platform):
    c = campaign.lower()
    if platform == "Amazon Ads":
        return "Amazon"
    elif platform == "Google Ads":
        return "Shopify"
    elif platform == "Facebook Ads":
        for prefix, landing in FB_LANDING_RULES:
            if prefix in c:
                return landing
        return "Other"
    return "Other"

# ── Pivot Helpers ────────────────────────────────────────────────────────────
def pivot(rows, key_field, date_field, value_field):
    r = defaultdict(lambda: defaultdict(float))
    for row in rows:
        r[row.get(key_field, "?")][row.get(date_field, "")] += (row.get(value_field, 0) or 0)
    return r

def pivot2(rows, k1, k2, date_field, value_field):
    r = defaultdict(lambda: defaultdict(float))
    for row in rows:
        key = (row.get(k1, "?"), row.get(k2, "?"))
        r[key][row.get(date_field, "")] += (row.get(value_field, 0) or 0)
    return r

# ── Write Helpers ────────────────────────────────────────────────────────────
def bar(ws, row, text, fill=BLUE, cols=8):
    for c in range(2, 2 + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = WF
    ws.cell(row=row, column=2, value=text)
    return row + 1

def hdr(ws, row, labels, start=2):
    for i, lab in enumerate(labels):
        cell = ws.cell(row=row, column=start + i, value=lab)
        cell.font = HF
        cell.border = THIN
    return row + 1

def nc(ws, row, col, val, fmt=NUM, bold=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = fmt
    cell.border = THIN
    if bold:
        cell.font = BF
    return cell

def lc(ws, row, col, val, bold=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.border = THIN
    if bold:
        cell.font = BF
    return cell

def brand_table(ws, row, title, data, brands, fmt=NUM, fill=BLUE):
    """Standard brand x month table. data = {brand: {month: value}}."""
    row = bar(ws, row, title, fill=fill)
    row = hdr(ws, row, ["", "Brand"] + MONTH_LABELS)
    totals = [0.0] * len(MONTHS)
    for b in brands:
        lc(ws, row, 3, b)
        for j, m in enumerate(MONTHS):
            v = data.get(b, {}).get(m, 0)
            nc(ws, row, 4 + j, v, fmt=fmt)
            totals[j] += v
        row += 1
    lc(ws, row, 3, "TOTAL", bold=True)
    for j, t in enumerate(totals):
        nc(ws, row, 4 + j, t, fmt=fmt, bold=True)
    return row + 2

def set_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

STD_WIDTHS = {"B": 5, "C": 28, "D": 20, "E": 22}

# ── Tab 1: Sales_Monthly ────────────────────────────────────────────────────
def build_sales(wb, q1d, q2d, q3d):
    ws = wb.create_sheet("Sales_Monthly")
    ws.sheet_properties.tabColor = "4472C4"
    row = 1
    ws.cell(row=row, column=2, value="SALES OVERVIEW (Monthly)").font = Font(bold=True, size=14)
    ws.cell(row=row, column=5, value="Feb 2026 is partial (through Feb 18)").font = Font(italic=True, size=9, color="FF0000")
    row += 2
    q1 = q1d["tableData"]

    # A: Blended Sales by Brand
    s = pivot(q1, "custom_5036", "date", "blended_total_sales")
    row = brand_table(ws, row, "A. Blended Sales by Brand (All Channels)", s, BRANDS)

    # B: Sales by Channel x Brand
    s2 = pivot2(q1, "custom_5036", "custom_5005", "date", "blended_total_sales")
    row = bar(ws, row, "B. Sales by Sales Channel x Brand", fill=DARK)
    row += 1
    for ch in CHANNELS:
        ch_data = {b: s2.get((b, ch), {}) for b in BRANDS}
        row = brand_table(ws, row, f"  {ch}", ch_data, BRANDS)

    # C: Orders by Channel x Brand
    o2 = pivot2(q1, "custom_5036", "custom_5005", "date", "blended_total_orders")
    row = bar(ws, row, "C. Orders by Sales Channel x Brand", fill=DARK)
    row += 1
    for ch in CHANNELS:
        ch_data = {b: o2.get((b, ch), {}) for b in BRANDS}
        row = brand_table(ws, row, f"  {ch}", ch_data, BRANDS, fmt=INT)

    # D: AOV
    q2, q3 = q2d["tableData"], q3d["tableData"]
    row = bar(ws, row, "D. Average Order Value", fill=DARK)
    row += 1
    aov_s = pivot(q2, "custom_5036", "date", "shopify_sales_main.computed.avg_order_value")
    row = brand_table(ws, row, "  D-1: Shopify AOV", aov_s, BRANDS, fmt=DOL)
    aov_a = pivot(q3, "custom_5036", "date", "amazonsp_order_items.computed.avg_order_value_amazon")
    row = brand_table(ws, row, "  D-2: Amazon AOV", aov_a, BRANDS, fmt=DOL)

    # E: Discounts
    row = bar(ws, row, "E. Discounts", fill=DARK)
    row += 1
    disc_s = pivot(q2, "custom_5036", "date", "shopify_sales_main.raw.discounts")
    row = brand_table(ws, row, "  E-1: Shopify Discounts", disc_s, BRANDS)
    disc_a = pivot(q3, "custom_5036", "date", "amazonsp_order_items.raw.promotion_discounts_amazon")
    row = brand_table(ws, row, "  E-2: Amazon Promotion Discounts", disc_a, BRANDS)

    set_widths(ws, STD_WIDTHS)
    print(f"  Sales_Monthly: {row} rows")

# ── Tab 2: AdSpend_Monthly ──────────────────────────────────────────────────
def build_adspend(wb, q5d, q6d, q7d, q8d):
    ws = wb.create_sheet("AdSpend_Monthly")
    ws.sheet_properties.tabColor = "ED7D31"
    row = 1
    ws.cell(row=row, column=2, value="AD SPEND & SALES (Monthly)").font = Font(bold=True, size=14)
    ws.cell(row=row, column=5, value="Brand from campaign name parsing").font = Font(italic=True, size=9, color="666666")
    row += 2

    plat_rows = {
        "Amazon Ads": q5d["tableData"], "Facebook Ads": q6d["tableData"],
        "Google Ads": q7d["tableData"], "TikTok Ads": q8d["tableData"],
    }
    spend_k = {
        "Amazon Ads": "amazonads_campaign.raw.cost",
        "Facebook Ads": "facebookads_ad_platform_and_device.raw.spend",
        "Google Ads": "googleads_campaign_and_device.raw.cost",
        "TikTok Ads": "tiktokads_campaign_and_platform.raw.spend",
    }
    sales_k = {
        "Amazon Ads": "amazonads_campaign.raw.attributed_sales",
        "Facebook Ads": "facebookads_ad_platform_and_device.raw.purchases_conversion_value",
        "Google Ads": "googleads_campaign_and_device.raw.conversion_value",
        "TikTok Ads": "tiktokads_campaign_and_platform.raw.purchases_conversion_value",
    }
    click_k = {
        "Amazon Ads": "amazonads_campaign.raw.clicks",
        "Facebook Ads": "facebookads_ad_platform_and_device.raw.clicks",
        "Google Ads": "googleads_campaign_and_device.raw.clicks",
        "TikTok Ads": "tiktokads_campaign_and_platform.raw.clicks",
    }

    # Aggregate by parsed brand and landing channel
    spend_pb = {}  # {platform: {brand: {month: val}}}
    sales_pb = {}
    clicks_p = {}  # {platform: {month: val}}
    spend_lb = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    sales_lb = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for plat, rows in plat_rows.items():
        bs = defaultdict(lambda: defaultdict(float))
        bv = defaultdict(lambda: defaultdict(float))
        pc = defaultdict(float)
        for r in rows:
            camp = r.get("campaign", "")
            dt = r.get("date", "")
            sp = r.get(spend_k[plat], 0) or 0
            sl = r.get(sales_k[plat], 0) or 0
            cl = r.get(click_k[plat], 0) or 0
            brand = parse_brand(camp, plat)
            landing = parse_landing(camp, plat)
            bs[brand][dt] += sp
            bv[brand][dt] += sl
            pc[dt] += cl
            spend_lb[landing][brand][dt] += sp
            sales_lb[landing][brand][dt] += sl
        spend_pb[plat] = dict(bs)
        sales_pb[plat] = dict(bv)
        clicks_p[plat] = dict(pc)

    def brands_for(data):
        return [b for b in BRANDS if b in data] + \
               [b for b in sorted(data) if b not in BRANDS and b != "Other"] + \
               (["Other"] if "Other" in data else [])

    # A: Ad Spend by Platform x Brand
    row = bar(ws, row, "A. Ad Spend by Platform x Brand", fill=DARK)
    row += 1
    for plat in AD_PLATFORMS:
        d = spend_pb.get(plat, {})
        row = brand_table(ws, row, f"  {plat} Spend", d, brands_for(d) or ["(none)"], fill=ORANGE)

    # Platform totals
    row = bar(ws, row, "Ad Spend Totals by Platform", fill=DARK)
    row = hdr(ws, row, ["", "Platform"] + MONTH_LABELS)
    gt = [0.0] * len(MONTHS)
    for plat in AD_PLATFORMS:
        lc(ws, row, 3, plat)
        d = spend_pb.get(plat, {})
        for j, m in enumerate(MONTHS):
            t = sum(d.get(b, {}).get(m, 0) for b in d)
            nc(ws, row, 4 + j, t)
            gt[j] += t
        row += 1
    lc(ws, row, 3, "GRAND TOTAL", bold=True)
    for j, t in enumerate(gt):
        nc(ws, row, 4 + j, t, bold=True)
    row += 2

    # B: Ad Spend by Landing Channel x Brand
    row = bar(ws, row, "B. Ad Spend by Landing Channel x Brand", fill=DARK)
    row += 1
    for landing in ["Shopify", "Amazon", "Target+", "Other"]:
        d = spend_lb.get(landing, {})
        if d:
            row = brand_table(ws, row, f"  Landing: {landing}", d, brands_for(d), fill=ORANGE)

    # C: Ad Sales by Platform x Brand
    row = bar(ws, row, "C. Ad Sales (Attributed) by Platform x Brand", fill=DARK)
    row += 1
    for plat in AD_PLATFORMS:
        d = sales_pb.get(plat, {})
        row = brand_table(ws, row, f"  {plat} Sales", d, brands_for(d) or ["(none)"], fill=GREEN)

    # D: ROAS by Platform x Brand
    row = bar(ws, row, "D. ROAS by Platform x Brand (Sales / Spend)", fill=DARK)
    row += 1
    for plat in AD_PLATFORMS:
        sp = spend_pb.get(plat, {})
        sl = sales_pb.get(plat, {})
        all_b = set(list(sp.keys()) + list(sl.keys()))
        roas = {}
        for b in all_b:
            roas[b] = {}
            for m in MONTHS:
                s = sp.get(b, {}).get(m, 0)
                v = sl.get(b, {}).get(m, 0)
                roas[b][m] = round(v / s, 2) if s > 0 else 0
        row = brand_table(ws, row, f"  {plat} ROAS", roas, brands_for(roas) or ["(none)"], fmt="0.00", fill=GREEN)

    # E: CPC by Platform
    row = bar(ws, row, "E. CPC by Platform (Spend / Clicks)", fill=DARK)
    row = hdr(ws, row, ["", "Platform"] + MONTH_LABELS)
    for plat in AD_PLATFORMS:
        lc(ws, row, 3, plat)
        d = spend_pb.get(plat, {})
        cl = clicks_p.get(plat, {})
        for j, m in enumerate(MONTHS):
            s = sum(d.get(b, {}).get(m, 0) for b in d)
            c = cl.get(m, 0)
            nc(ws, row, 4 + j, s / c if c > 0 else 0, fmt=DOL)
        row += 1
    row += 1

    set_widths(ws, STD_WIDTHS)
    print(f"  AdSpend_Monthly: {row} rows")
    return spend_pb, sales_pb, spend_lb, sales_lb

# ── Tab 3: Organic_Monthly ──────────────────────────────────────────────────
def build_organic(wb, q1d, sales_lb):
    ws = wb.create_sheet("Organic_Monthly")
    ws.sheet_properties.tabColor = "548235"
    row = 1
    ws.cell(row=row, column=2, value="ORGANIC vs PAID SALES (Monthly)").font = Font(bold=True, size=14)
    row += 2
    q1 = q1d["tableData"]
    s2 = pivot2(q1, "custom_5036", "custom_5005", "date", "blended_total_sales")

    groups = {
        "Shopify (D2C+TikTok+FBM)": (["D2C", "TikTok", "FBM"], "Shopify"),
        "Amazon": (["Amazon - Grosmimi USA", "Amazon - Fleeters", "Amazon - Orbitool"], "Amazon"),
        "Target+": (["Target+"], "Target+"),
    }

    # A: Total Sales
    row = bar(ws, row, "A. Total Sales by Channel Group x Brand", fill=DARK)
    row += 1
    total_by_group = {}
    for gname, (channels, _) in groups.items():
        gd = defaultdict(lambda: defaultdict(float))
        for b in BRANDS:
            for ch in channels:
                for m in MONTHS:
                    gd[b][m] += s2.get((b, ch), {}).get(m, 0)
        total_by_group[gname] = gd
        row = brand_table(ws, row, f"  {gname}", gd, BRANDS)

    # B: Ad-Attributed Sales
    row = bar(ws, row, "B. Ad-Attributed Sales by Channel Group x Brand", fill=DARK)
    row += 1
    ad_by_group = {}
    for gname, (_, landing) in groups.items():
        d = sales_lb.get(landing, {})
        ad_by_group[gname] = d
        bl = [b for b in BRANDS if b in d] + [b for b in sorted(d) if b not in BRANDS]
        row = brand_table(ws, row, f"  Ad Sales: {gname}", d, bl or BRANDS, fill=ORANGE)

    # C: Organic = Total - Ad
    row = bar(ws, row, "C. Organic Sales (Total - Ad-Attributed)", fill=DARK)
    row += 1
    for gname in groups:
        od = defaultdict(lambda: defaultdict(float))
        for b in BRANDS:
            for m in MONTHS:
                total = total_by_group[gname].get(b, {}).get(m, 0)
                ad = ad_by_group[gname].get(b, {}).get(m, 0)
                od[b][m] = total - ad
        row = brand_table(ws, row, f"  Organic: {gname}", od, BRANDS, fill=GREEN)

    # D: Organic %
    row = bar(ws, row, "D. Organic % (Organic / Total)", fill=DARK)
    row += 1
    for gname in groups:
        pd = defaultdict(lambda: defaultdict(float))
        for b in BRANDS:
            for m in MONTHS:
                total = total_by_group[gname].get(b, {}).get(m, 0)
                ad = ad_by_group[gname].get(b, {}).get(m, 0)
                pd[b][m] = (total - ad) / total if total > 0 else 0
        row = brand_table(ws, row, f"  Organic %: {gname}", pd, BRANDS, fmt=PCT, fill=GREEN)

    set_widths(ws, STD_WIDTHS)
    print(f"  Organic_Monthly: {row} rows")

# ── Tab 4: Margin_Monthly ───────────────────────────────────────────────────
def build_margin(wb, q2d, q3d, spend_pb):
    ws = wb.create_sheet("Margin_Monthly")
    ws.sheet_properties.tabColor = "FF0000"
    row = 1
    ws.cell(row=row, column=2, value="MARGIN ANALYSIS (Monthly)").font = Font(bold=True, size=14)
    ws.cell(row=row, column=5, value="Yellow = manual input needed").font = Font(italic=True, size=9, color="FF0000")
    row += 2
    q2, q3 = q2d["tableData"], q3d["tableData"]

    # A: Revenue
    row = bar(ws, row, "A. Revenue (Net Sales)", fill=DARK)
    row += 1
    ss = pivot(q2, "custom_5036", "date", "shopify_sales_main.computed.total_sales")
    row = brand_table(ws, row, "  A-1: Shopify Net Sales", ss, BRANDS)
    sa = pivot(q3, "custom_5036", "date", "amazonsp_order_items.computed.total_sales_amazon")
    row = brand_table(ws, row, "  A-2: Amazon Net Sales", sa, BRANDS)

    # B: COGS
    row = bar(ws, row, "B. COGS", fill=DARK)
    row += 1
    cs = pivot(q2, "custom_5036", "date", "shopify_sales_main.raw.cost_of_products_custom")
    row = brand_table(ws, row, "  B-1: Shopify COGS (GSheet)", cs, BRANDS)
    ca = pivot(q3, "custom_5036", "date", "amazonsp_order_items.raw.cost_of_products_amazon")
    row = brand_table(ws, row, "  B-2: Amazon COGS", ca, BRANDS)

    # C: Channel Fees
    row = bar(ws, row, "C. Channel Fees", fill=DARK)
    row += 1
    fa = pivot(q3, "custom_5036", "date", "amazonsp_order_items.raw.total_fees_amazon")
    row = brand_table(ws, row, "  C-1: Amazon Total Fees", fa, BRANDS)
    fs = pivot(q2, "custom_5036", "date", "shopify_sales_main.raw.transaction_fees")
    row = brand_table(ws, row, "  C-2: Shopify Transaction Fees", fs, BRANDS)

    # D: CM1, CM2 from Polar
    row = bar(ws, row, "D. Contribution Margins (Polar Computed)", fill=DARK)
    row += 1
    cm1 = pivot(q2, "custom_5036", "date", "shopify_sales_main.computed.contribution_margin_1")
    row = brand_table(ws, row, "  D-1: Shopify CM1", cm1, BRANDS)
    cm2 = pivot(q2, "custom_5036", "date", "shopify_sales_main.computed.contribution_margin_2")
    row = brand_table(ws, row, "  D-2: Shopify CM2", cm2, BRANDS)

    # E: Ad Spend by Brand (for CM after Ads calc)
    row = bar(ws, row, "E. Total Ad Spend by Brand (all platforms)", fill=DARK)
    row = hdr(ws, row, ["", "Brand"] + MONTH_LABELS)
    for b in BRANDS:
        lc(ws, row, 3, b)
        for j, m in enumerate(MONTHS):
            t = sum(spend_pb.get(p, {}).get(b, {}).get(m, 0) for p in AD_PLATFORMS)
            nc(ws, row, 4 + j, t)
        row += 1
    row += 1

    # F: Manual Input Skeleton
    row = bar(ws, row, "F. Manual Input Items (fill yellow cells)", fill=ORANGE)
    row = hdr(ws, row, ["", "Cost Item"] + MONTH_LABELS)
    manual = [
        "Tariffs (15%/20%/0% by product)",
        "Korea to US Shipping",
        "US Fulfillment (FBA/Shipbob)",
        "Influencer Costs",
        "Other Operating Costs",
    ]
    for item in manual:
        lc(ws, row, 3, item)
        for j in range(len(MONTHS)):
            cell = ws.cell(row=row, column=4 + j, value=0)
            cell.fill = YELLOW
            cell.border = THIN
            cell.number_format = NUM
        row += 1
    row += 1

    set_widths(ws, {"B": 5, "C": 35, "D": 20, "E": 22})
    print(f"  Margin_Monthly: {row} rows")

# ── Tab 5: Product_Monthly ──────────────────────────────────────────────────
def build_product(wb, q4d):
    ws = wb.create_sheet("Product_Monthly")
    ws.sheet_properties.tabColor = "7030A0"
    row = 1
    ws.cell(row=row, column=2, value="PRODUCT VARIANT BREAKDOWN (Monthly)").font = Font(bold=True, size=14)
    row += 2
    q4 = q4d["tableData"]

    # A: Sales by Product
    ps = pivot(q4, "custom_5037", "date", "blended_total_sales")
    prods = sorted(ps, key=lambda p: ps[p].get("2026-01-01", 0), reverse=True)
    row = bar(ws, row, "A. Sales by Product Variant")
    row = hdr(ws, row, ["#", "Product Variant"] + MONTH_LABELS)
    ts = [0.0] * len(MONTHS)
    for i, p in enumerate(prods):
        lc(ws, row, 2, i + 1)
        lc(ws, row, 3, p)
        for j, m in enumerate(MONTHS):
            v = ps[p].get(m, 0)
            nc(ws, row, 4 + j, v)
            ts[j] += v
        row += 1
    lc(ws, row, 3, "TOTAL", bold=True)
    for j, t in enumerate(ts):
        nc(ws, row, 4 + j, t, bold=True)
    row += 2

    # B: Orders by Product
    po = pivot(q4, "custom_5037", "date", "blended_total_orders")
    prods_o = sorted(po, key=lambda p: po[p].get("2026-01-01", 0), reverse=True)
    row = bar(ws, row, "B. Orders by Product Variant")
    row = hdr(ws, row, ["#", "Product Variant"] + MONTH_LABELS)
    to = [0.0] * len(MONTHS)
    for i, p in enumerate(prods_o):
        lc(ws, row, 2, i + 1)
        lc(ws, row, 3, p)
        for j, m in enumerate(MONTHS):
            v = po[p].get(m, 0)
            nc(ws, row, 4 + j, v, fmt=INT)
            to[j] += v
        row += 1
    lc(ws, row, 3, "TOTAL", bold=True)
    for j, t in enumerate(to):
        nc(ws, row, 4 + j, t, fmt=INT, bold=True)
    row += 2

    set_widths(ws, {"B": 5, "C": 45, "D": 20, "E": 22})
    print(f"  Product_Monthly: {row} rows")

# ── Tab 6: Data_Status ──────────────────────────────────────────────────────
def build_status(wb):
    ws = wb.create_sheet("Data_Status")
    ws.sheet_properties.tabColor = "808080"
    row = 1
    ws.cell(row=row, column=2, value="DATA STATUS & GAPS").font = Font(bold=True, size=14)
    row += 2

    row = bar(ws, row, "Connector Status")
    row = hdr(ws, row, ["", "Connector", "Status", "Notes"])
    connectors = [
        ("Shopify", "Incremental", "Live and synced"),
        ("Amazon Selling Partner", "Incremental", "Live and synced"),
        ("Amazon Ads", "Incremental", "Live and synced"),
        ("Facebook Ads", "BUILDING", "Data may be incomplete"),
        ("Google Ads", "Incremental", "Live and synced"),
        ("TikTok Ads", "Incremental", "$0 spend - may not be active"),
    ]
    for name, status, note in connectors:
        lc(ws, row, 3, name)
        cell = lc(ws, row, 4, status)
        if "BUILD" in status:
            cell.font = Font(bold=True, color="FF0000")
        lc(ws, row, 5, note)
        row += 1
    row += 1

    row = bar(ws, row, "Known Data Gaps", fill=ORANGE)
    gaps = [
        "1. FB Ads connector still building - ad spend/sales data may be incomplete",
        "2. Feb 2026 is partial (through Feb 18)",
        "3. Grosmimi: Polar ~$657K vs existing ~$708K (Jan) - Gross vs Net difference",
        "4. Ad spend brand breakdown uses campaign name parsing - some misclassification possible",
        "5. TikTok Ads = $0 (not connected or not running)",
        "6. Organic Sales = Total - Ad-attributed (approximate)",
        "7. 'ETC' and 'B2B wholesale' brands excluded from standard brand tables (~$8K total)",
    ]
    for g in gaps:
        ws.cell(row=row, column=2, value=g).font = Font(size=9)
        row += 1
    row += 1

    row = bar(ws, row, "Manual Input Required", fill=ORANGE)
    manual = [
        "1. Tariffs (15%/20%/0% by product type)",
        "2. Korea to US Shipping costs",
        "3. US Fulfillment costs (FBA / Shipbob / D-Trans)",
        "4. Influencer costs by campaign/brand",
        "5. Other operating costs",
    ]
    for m in manual:
        ws.cell(row=row, column=2, value=m).font = Font(size=9)
        row += 1
    row += 1

    row = bar(ws, row, "Future Modules (TBD)")
    tbd = [
        "1. Weekly granularity tabs - add after monthly validation",
        "2. Inventory tracking - separate module",
        "3. Search volume - Helium10 / Google Ads Keyword Planner",
        "4. Content pipeline - Airtable integration",
    ]
    for t in tbd:
        ws.cell(row=row, column=2, value=t).font = Font(size=9)
        row += 1

    set_widths(ws, {"B": 5, "C": 30, "D": 20, "E": 60})
    print(f"  Data_Status: {row} rows")

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("Loading Polar data from JSON...")
    q1 = load("q1_blended_brand_channel.json")
    q2 = load("q2_shopify_brand.json")
    q3 = load("q3_amazon_brand.json")
    q4 = load("q4_product_variant.json")
    q5 = load("q5_amazon_ads_campaign.json")
    q6 = load("q6_facebook_ads_campaign.json")
    q7 = load("q7_google_ads_campaign.json")
    q8 = load("q8_tiktok_ads_campaign.json")

    print("Building dashboard tabs...")
    wb = Workbook()
    wb.remove(wb.active)

    build_sales(wb, q1, q2, q3)
    spend_pb, sales_pb, spend_lb, sales_lb = build_adspend(wb, q5, q6, q7, q8)
    build_organic(wb, q1, sales_lb)
    build_margin(wb, q2, q3, spend_pb)
    build_product(wb, q4)
    build_status(wb)

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"Tabs: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()
