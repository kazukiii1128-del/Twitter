"""
Parse CI/PL export documents from various brands and organize them.
Extracts: 년월, 브랜드, 운송방식, 금액, Exporter, Importer, Destination, 파일출처
Handles both Excel (.xlsx/.xls) and PDF files.
"""

import os
import re
import json
import openpyxl
import pdfplumber
from datetime import datetime
from pathlib import Path

# ─── Brand mapping from folder names ───
FOLDER_BRAND_MAP = {
    'LFU발주': 'Grosmimi',
    'LFU JP 발주': 'Grosmimi JP',
    '꼬메모이': 'Commemoi',
    '내아이애': 'Naeiae',
    '네이쳐러브메레': 'Nature Love Mere',
    '밤부베베': 'BambooBebe',
    '비마이매직': 'BeeMyMagic',
    '알프레미오 발주': 'Alpremio',
    '오비터스 직접수출': 'Orbiters Direct',
    '차앤맘': 'Cha&Mom',
    '코니스': 'Conys',
    '하뚱': 'Hattung',
    'Grosmimi': 'Grosmimi',
    'Alpremio': 'Alpremio',
    'BabyRabbit': 'BabyRabbit',
    'Commemoi': 'Commemoi',
    'Nature Love Mere': 'Nature Love Mere',
    'BambooBebe': 'BambooBebe',
    'beemymagic': 'BeeMyMagic',
    'New Folder With Items': 'Mixed',
}

# Known exporter → brand
EXPORTER_BRAND_MAP = {
    'littlefingerusa': 'Grosmimi',
    'naeiae': 'Naeiae',
    'commemoi': 'Commemoi',
    'alpremio': 'Alpremio',
    'jtomorrow': 'Alpremio',
    'babyrabbit': 'BabyRabbit',
    'thebamboo': 'BambooBebe',
    'the bamboo': 'BambooBebe',
    'klemarang': 'Nature Love Mere',
    'orbiters': 'Orbiters',
    'beemymagic': 'BeeMyMagic',
    '코니코프': 'Conys',
}


def detect_shipping_method(filepath, text_content=""):
    """Detect AIR vs SEA from filename or content."""
    fn = os.path.basename(filepath).lower()
    full_path = filepath.lower()
    combined = fn + " " + full_path + " " + text_content.lower()

    if 'air' in fn or '항공' in fn or '항공' in full_path:
        return 'AIR'
    if 'sea' in fn or '해상' in fn or '해상' in full_path or 'ocean' in combined:
        return 'SEA'
    if 'flight' in fn:
        return 'AIR'
    # Check content
    if 'air' in text_content.lower():
        return 'AIR'
    if 'sea' in text_content.lower() or 'ocean cargo' in text_content.lower():
        return 'SEA'
    return 'UNKNOWN'


def extract_date_from_filename(filepath):
    """Try to extract date from filename patterns like 20250814, 2024.03.28, etc."""
    fn = os.path.basename(filepath)

    # Pattern: YYYYMMDD
    m = re.search(r'20(\d{2})(\d{2})(\d{2})', fn)
    if m:
        try:
            y, mo, d = int('20' + m.group(1)), int(m.group(2)), int(m.group(3))
            if 1 <= mo <= 12 and 1 <= d <= 31:
                return f"{y:04d}-{mo:02d}"
        except:
            pass

    # Pattern: YYYY.MM.DD or YYYY.MMDD
    m = re.search(r'(20\d{2})[.\-_](\d{2})[.\-_]?(\d{2})?', fn)
    if m:
        try:
            y = int(m.group(1))
            mo = int(m.group(2))
            if 1 <= mo <= 12:
                return f"{y:04d}-{mo:02d}"
        except:
            pass

    # Pattern: YYMM in folder names like 2412월
    m = re.search(r'(\d{2})(\d{2})월', filepath)
    if m:
        y = int('20' + m.group(1))
        mo = int(m.group(2))
        if 1 <= mo <= 12:
            return f"{y:04d}-{mo:02d}"

    return None


def detect_brand_from_path(filepath):
    """Detect brand from folder path."""
    parts = filepath.replace('\\', '/').split('/')
    for part in parts:
        if part in FOLDER_BRAND_MAP:
            return FOLDER_BRAND_MAP[part]
    # Check filename
    fn = os.path.basename(filepath).lower()
    if 'grosmimi' in fn or 'lfu' in fn:
        return 'Grosmimi'
    if 'naeiae' in fn or '내아이애' in fn:
        return 'Naeiae'
    if 'commemoi' in fn or '꼬메모이' in fn:
        return 'Commemoi'
    if 'alpremio' in fn:
        return 'Alpremio'
    if 'babyrabbit' in fn or 'baby rabbit' in fn:
        return 'BabyRabbit'
    if 'bamboo' in fn or '밤부베베' in fn:
        return 'BambooBebe'
    if 'beemymagic' in fn or 'bee' in fn:
        return 'BeeMyMagic'
    if 'conys' in fn or '코니스' in fn or '코니코프' in fn:
        return 'Conys'
    if 'hattung' in fn or '하뚱' in fn or '처음교육' in fn:
        return 'Hattung'
    if 'nature' in fn or '네이쳐' in fn or 'klemarang' in fn:
        return 'Nature Love Mere'
    return None


def detect_brand_from_exporter(exporter_name):
    """Detect brand from exporter name."""
    if not exporter_name:
        return None
    lower = exporter_name.lower()
    for key, brand in EXPORTER_BRAND_MAP.items():
        if key in lower:
            return brand
    return None


def clean_amount(val):
    """Clean amount string to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    # Remove currency symbols and spaces
    s = re.sub(r'[US$₩,\s]', '', s)
    # Handle PDF spacing issues like "4 ,830.00" → "4830.00"
    s = re.sub(r'\s+', '', s)
    try:
        return float(s)
    except:
        return None


def normalize_shipment_key(filepath, date_str):
    """Create a key to identify the same shipment across file formats."""
    fn = os.path.basename(filepath).lower()

    # Pattern 1: zzbb (optionally with air/sea) + date
    m = re.search(r'zzbb[_\s]*(?:air|sea)?[_\s]*(\d{8})', fn)
    if m:
        return f"zzbb_{m.group(1)}"

    # Pattern 3: date in filename like _20250430 or _240701
    m = re.search(r'(\d{8})', fn)
    if m:
        date_code = m.group(1)
        # Get a brand hint from the filename
        brand_hint = ''
        for kw in ['grosmimi', '그로미미', 'naeiae', '내아이애', 'commemoi', '꼬메모이',
                    'alpremio', 'babyrabbit', 'bamboo', '밤부베베', 'bee', 'conys', '코니스',
                    'hattung', '하뚱', '처음교육', 'nature', '네이쳐', 'klemarang']:
            if kw in fn or kw in filepath.lower():
                brand_hint = kw[:5]
                break
        return f"{brand_hint}_{date_code}"

    return None


def parse_excel_cipl(filepath):
    """Parse a CI/PL Excel file and extract header info."""
    result = {
        'date': None,
        'brand': None,
        'shipping': detect_shipping_method(filepath),
        'total_amount': None,
        'currency': None,
        'exporter': None,
        'importer': None,
        'destination': None,
        'items_summary': [],
        'total_qty': None,
        'total_ctn': None,
        'cbm': None,
        'price_terms': None,
        'invoice_no': None,
        'source': filepath,
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        result['error'] = str(e)
        return result

    # Use first sheet or specific named sheets
    ws = wb.active
    if not ws:
        ws = wb[wb.sheetnames[0]]

    # Check if this sheet looks like a CI/PL
    is_cipl = False
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for cell in row:
            if cell and 'COMMERCIAL INVOICE' in str(cell).upper():
                is_cipl = True
                break
        if is_cipl:
            break

    # Also check named sheets
    if not is_cipl:
        for name in wb.sheetnames:
            if 'CI' in name.upper() or 'PL' in name.upper() or 'INVOICE' in name.upper() or 'Updated' in name:
                ws = wb[name]
                is_cipl = True
                break

    if not is_cipl:
        # Check if any sheet has CI content
        for name in wb.sheetnames:
            test_ws = wb[name]
            for row in test_ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if cell and 'COMMERCIAL INVOICE' in str(cell).upper():
                        ws = test_ws
                        is_cipl = True
                        break
                if is_cipl:
                    break
            if is_cipl:
                break

    if not is_cipl:
        # This might be an order sheet, not a CI/PL - try to extract what we can
        wb.close()
        result['error'] = 'Not a CI/PL document'
        return result

    # Read all cells into a dict for flexible access
    cells = {}
    max_row = min(ws.max_row or 100, 100)
    max_col = min(ws.max_column or 30, 30)

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cells[cell.coordinate] = cell.value

    # Helper to get value from any cell
    def find_value_near(keyword, search_range=30):
        """Find a value in the cell to the right or below a keyword."""
        for coord, val in cells.items():
            if val and keyword.lower() in str(val).lower():
                # Get row and col
                col_letter = re.match(r'([A-Z]+)', coord).group(1)
                row_num = int(re.search(r'(\d+)', coord).group(1))
                # Check cells to the right
                for offset in range(1, 15):
                    next_col = chr(ord(col_letter[-1]) + offset) if len(col_letter) == 1 else col_letter
                    next_coord = f"{next_col}{row_num}"
                    if next_coord in cells and cells[next_coord]:
                        return cells[next_coord]
        return None

    # Extract date
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val)
        if isinstance(val, datetime):
            result['date'] = val.strftime('%Y-%m')
            break
        # Check for date patterns
        m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
        if m:
            result['date'] = f"{m.group(1)}-{m.group(2)}"
            break

    if not result['date']:
        result['date'] = extract_date_from_filename(filepath)

    # Extract exporter
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        row_num = int(re.search(r'(\d+)', coord).group(1))

        if 'Shipper' in s or 'Seller' in s or 'Exporter' in s:
            # Look at row below
            for next_row in range(row_num + 1, row_num + 3):
                for col in ['A', 'B', 'C']:
                    nc = f"{col}{next_row}"
                    if nc in cells and cells[nc]:
                        v = str(cells[nc]).strip()
                        if v and len(v) > 3 and 'tel' not in v.lower() and 'email' not in v.lower() and '@' not in v:
                            result['exporter'] = v
                            break
                if result['exporter']:
                    break
            break

    # Extract importer/consignee (first match only, avoid Ultimate Consignee)
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        row_num = int(re.search(r'(\d+)', coord).group(1))

        if ('Consignee' in s or 'Buyer' in s or 'Importer' in s) and not result.get('importer'):
            # Look at row below, only in columns A-F (avoid Ultimate Consignee in G+)
            for next_row in range(row_num + 1, row_num + 3):
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    nc = f"{col}{next_row}"
                    if nc in cells and cells[nc]:
                        v = str(cells[nc]).strip()
                        if v and len(v) > 3 and 'tel' not in v.lower() and '@' not in v and 'walk by' not in v.lower() and 'same as' not in v.lower():
                            result['importer'] = v
                            break
                if result['importer']:
                    break
            break

    # Extract destination
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        if s == 'Destination' or (s.startswith('Destination') and 'Final' not in s):
            row_num = int(re.search(r'(\d+)', coord).group(1))
            for col in ['E', 'F', 'G', 'H', 'I', 'N']:
                nc = f"{col}{row_num}"
                if nc in cells and cells[nc]:
                    v = str(cells[nc]).strip()
                    if v and v != 'Destination' and 'Ultimate' not in v and len(v) >= 2:
                        result['destination'] = v
                        break
            break

    # Fallback: check for country mentions
    if not result['destination']:
        all_text = ' '.join(str(v) for v in cells.values() if v)
        if 'USA' in all_text:
            result['destination'] = 'USA'
        elif 'Japan' in all_text or 'JP' in all_text:
            result['destination'] = 'Japan'

    # Extract Price Terms
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        if 'Price Terms' in s or 'Price Term' in s:
            row_num = int(re.search(r'(\d+)', coord).group(1))
            for col in ['E', 'F', 'G', 'H']:
                nc = f"{col}{row_num}"
                if nc in cells and cells[nc]:
                    v = str(cells[nc]).strip()
                    if v and 'Price' not in v:
                        result['price_terms'] = v
                        break
            break

    # Extract Invoice No
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        if 'Invoice No' in s:
            row_num = int(re.search(r'(\d+)', coord).group(1))
            for col in ['J', 'K', 'L', 'M', 'N', 'O']:
                nc = f"{col}{row_num}"
                if nc in cells and cells[nc]:
                    v = str(cells[nc]).strip()
                    if v and 'Invoice' not in v:
                        result['invoice_no'] = v
                        break
            break

    # Detect currency from column headers or amounts
    currency_detected = False
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val)
        if 'USD' in s or '($)' in s:
            result['currency'] = 'USD'
            currency_detected = True
            break
        if 'KRW' in s or '(KRW)' in s or '원' in s:
            result['currency'] = 'KRW'
            currency_detected = True
            break

    if not currency_detected:
        result['currency'] = 'USD'  # default

    # Find the Amount column header first
    amount_col = None
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        if 'Amount' in s and ('$' in s or 'USD' in s or 'KRW' in s):
            col_letter = re.match(r'([A-Z]+)', coord).group(1)
            amount_col = col_letter
            if 'KRW' in s:
                result['currency'] = 'KRW'
            elif 'USD' in s or '$' in s:
                result['currency'] = 'USD'
            break

    # Find TOTAL row - skip "Total PKG" rows, look for actual sum total
    total_rows = []
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip().upper()
        row_num = int(re.search(r'(\d+)', coord).group(1))
        # Match "TOTAL" but skip "TOTAL PKG"
        if (s == 'TOTAL' or s == 'TOTAL :' or s.startswith('TOTAL ')) and 'PKG' not in s:
            total_rows.append(row_num)

    for row_num in total_rows:
        # If we know the amount column, use it directly
        if amount_col:
            nc = f"{amount_col}{row_num}"
            if nc in cells:
                amt = clean_amount(cells[nc])
                if amt and amt > 10:
                    result['total_amount'] = amt
                    break

        # Otherwise scan the total row for the largest numeric value
        amounts = []
        for col_idx in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            nc = f"{col_letter}{row_num}"
            if nc in cells:
                v = cells[nc]
                amt = clean_amount(v)
                if amt and amt > 100:
                    amounts.append(amt)

        if amounts:
            result['total_amount'] = max(amounts)
            break

    # Also check Total PKG row for qty/ctn/cbm
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        if 'Total PKG' in s or 'Total:' in s:
            row_num = int(re.search(r'(\d+)', coord).group(1))
            # Scan for numbers
            for col_idx in range(1, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                nc = f"{col_letter}{row_num}"
                if nc in cells:
                    v = cells[nc]
                    if isinstance(v, str) and 'PCS' in v:
                        result['total_qty'] = v
                    elif isinstance(v, str) and 'CTN' in v:
                        result['total_ctn'] = v
                    elif isinstance(v, (int, float)):
                        fv = float(v)
                        if fv < 100:  # Likely CBM
                            result['cbm'] = fv
                        elif fv < 10000:  # Likely CTN count
                            if not result['total_ctn']:
                                result['total_ctn'] = int(fv)
            break

    # Extract item summaries from data rows
    items = []
    data_start = None
    for coord, val in cells.items():
        if val is None:
            continue
        s = str(val).strip()
        if 'Commodity' in s or 'Description' in s:
            row_num = int(re.search(r'(\d+)', coord).group(1))
            data_start = row_num + 1
            # Check if next row is also header (sub-header)
            for col in ['A', 'B', 'G', 'H']:
                nc = f"{col}{data_start}"
                if nc in cells:
                    v = str(cells[nc]).strip()
                    if v in ('CTN', 'Box', '($)', 'CTN Q'):
                        data_start += 1
                        break
            break

    # Non-item keywords to filter out
    skip_item_kw = ['total', 'remark', 'bank', 'remittee', 'remittess', 'swift',
                    'account', 'signed', 'branch', 'authorized', 'stamp', 'seal']

    if data_start:
        for r in range(data_start, min(data_start + 50, max_row + 1)):
            desc = None
            for col in ['D', 'E', 'F']:
                nc = f"{col}{r}"
                if nc in cells and cells[nc]:
                    v = str(cells[nc]).strip()
                    if len(v) > 5 and not v.startswith('-'):
                        # Check it's not a non-item line
                        if not any(kw in v.lower() for kw in skip_item_kw):
                            desc = v
                            break
            if desc:
                items.append(desc)

    result['items_summary'] = items[:20]

    wb.close()
    return result


def parse_pdf_cipl(filepath):
    """Parse a CI/PL PDF file and extract header info."""
    result = {
        'date': None,
        'brand': None,
        'shipping': detect_shipping_method(filepath),
        'total_amount': None,
        'currency': None,
        'exporter': None,
        'importer': None,
        'destination': None,
        'items_summary': [],
        'total_qty': None,
        'price_terms': None,
        'invoice_no': None,
        'source': filepath,
    }

    try:
        with pdfplumber.open(filepath) as pdf:
            full_text = ""
            for page in pdf.pages[:3]:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
    except Exception as e:
        result['error'] = str(e)
        return result

    if not full_text.strip():
        result['error'] = 'No text extracted from PDF'
        return result

    # Update shipping method with content
    result['shipping'] = detect_shipping_method(filepath, full_text)

    lines = full_text.split('\n')

    # Check if this is a CI/PL
    is_cipl = any('COMMERCIAL INVOICE' in line.upper() or 'PACKING LIST' in line.upper()
                   for line in lines[:10])
    if not is_cipl:
        # Check for export cert (수출필증) - different doc type
        if any('수출필증' in line or 'EXP_' in line for line in lines[:5]):
            result['error'] = 'Export certificate, not CI/PL'
            return result
        result['error'] = 'Not a CI/PL document'
        return result

    # Extract date
    for line in lines:
        # Pattern: YYYY-MM-DD
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', line)
        if m:
            result['date'] = f"{m.group(1)}-{m.group(2)}"
            break
        # Pattern: Month Day, Year
        m = re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),?\s*(\d{4})', line)
        if m:
            months = {'January': 1, 'February': 2, 'March': 3, 'April': 4,
                      'May': 5, 'June': 6, 'July': 7, 'August': 8,
                      'September': 9, 'October': 10, 'November': 11, 'December': 12}
            mo = months[m.group(1)]
            y = int(m.group(3))
            result['date'] = f"{y:04d}-{mo:02d}"
            break

    if not result['date']:
        result['date'] = extract_date_from_filename(filepath)

    # Extract exporter (first company name after Shipper/Seller)
    for i, line in enumerate(lines):
        if 'Shipper' in line or 'Seller' in line:
            for j in range(i + 1, min(i + 5, len(lines))):
                candidate = lines[j].strip()
                if candidate and len(candidate) > 3:
                    if not any(kw in candidate.lower() for kw in ['tel', 'fax', 'email', '@', 'address', 'consignee', 'buyer']):
                        result['exporter'] = candidate.split('  ')[0].strip()
                        break
            break

    # Clean exporter - if it looks like a date, it's wrong
    if result.get('exporter'):
        exp = result['exporter']
        if re.match(r'^(January|February|March|April|May|June|July|August|September|October|November|December)\s', exp):
            result['exporter'] = None
        # If it includes an invoice number, clean it
        if 'EXP' in exp:
            result['exporter'] = re.split(r'\s+EXP', exp)[0].strip()

    # If no shipper section found, check first few lines for company name
    if not result['exporter']:
        for line in lines[:5]:
            line = line.strip()
            if line and ('Inc' in line or 'Ltd' in line or 'Co.' in line):
                result['exporter'] = line.split('  ')[0].strip()
                break

    # Extract importer/consignee
    for i, line in enumerate(lines):
        if ('Consignee' in line or 'Buyer' in line) and 'Ultimate' not in line:
            for j in range(i + 1, min(i + 5, len(lines))):
                candidate = lines[j].strip()
                if candidate and len(candidate) > 3:
                    if not any(kw in candidate.lower() for kw in ['tel', 'fax', 'email', '@', 'walk by faith', 'same as']):
                        # Take only the first entity (before double space or long gap)
                        parts = re.split(r'\s{2,}', candidate)
                        result['importer'] = parts[0].strip()
                        break
            break

    # Extract destination
    for line in lines:
        if 'Final Destination' in line or 'Destination' in line:
            # Check same line for country
            if 'USA' in line:
                result['destination'] = 'USA'
            elif 'Japan' in line or 'JP' in line:
                result['destination'] = 'Japan'
            else:
                m = re.search(r'Destination\s+(.+)', line)
                if m:
                    result['destination'] = m.group(1).strip()
            break

    # Extract price terms
    for line in lines:
        if 'Price Terms' in line or 'Incoterms' in line:
            terms = re.search(r'(EXW|FOB|CIF|DAP|DDP|CFR|FCA)', line)
            if terms:
                result['price_terms'] = terms.group(1)
            break

    # Extract invoice number
    for line in lines:
        m = re.search(r'Invoice No\.?\s*:?\s*([A-Z0-9\-_]+)', line)
        if m:
            result['invoice_no'] = m.group(1)
            break

    # Extract total amount
    for line in lines:
        if 'Total' in line:
            # Find USD amounts
            amounts = re.findall(r'US?\$?([\d,]+\.?\d*)', line)
            if amounts:
                try:
                    result['total_amount'] = max(float(a.replace(',', '')) for a in amounts)
                    result['currency'] = 'USD'
                except:
                    pass
            # Find KRW amounts
            if not result['total_amount']:
                amounts = re.findall(r'([\d,]+)', line)
                if amounts:
                    vals = []
                    for a in amounts:
                        try:
                            v = float(a.replace(',', ''))
                            if v > 1000:
                                vals.append(v)
                        except:
                            pass
                    if vals:
                        result['total_amount'] = max(vals)

    # Detect currency
    if 'USD' in full_text or 'US$' in full_text or '($)' in full_text:
        result['currency'] = 'USD'
    elif 'KRW' in full_text or '원' in full_text:
        result['currency'] = 'KRW'

    return result


def collect_cipl_files():
    """Collect all CI/PL related files from known locations."""
    search_roots = [
        (r"Z:\Orbiters\CI, PL, BL", True),
        (r"Z:\Orbiters\발주 서류 관리", True),
        (r"Z:\Orbiters\JH\리틀핑거스 cipl 및 서류", True),
        (r"C:\Users\wjcho\Downloads", False),
        (r"C:\Users\wjcho\OneDrive - Orbiters\Microsoft Teams 채팅 파일", False),
    ]

    files = []
    seen_names = {}  # For deduplication

    for root_path, deep_search in search_roots:
        if not os.path.exists(root_path):
            continue

        for dirpath, dirnames, filenames in os.walk(root_path):
            # Skip recycle folders
            if '#recycle' in dirpath:
                continue

            for f in filenames:
                # Skip temp files
                if f.startswith('~$'):
                    continue
                ext = os.path.splitext(f)[1].lower()
                if ext not in ('.xlsx', '.xls', '.pdf'):
                    continue

                full_path = os.path.join(dirpath, f)
                fl = f.lower()

                # Filter for CI/PL related files
                is_cipl_file = False

                # Check filename for CI/PL keywords
                if any(kw in fl for kw in ['cipl', 'ci pl', 'ci_pl', 'ci&pl', 'ci^0pl',
                                             'commercial invoice', 'packing list', 'ci_', 'pl_']):
                    is_cipl_file = True

                # If in CI, PL, BL folder, include all
                if 'CI, PL, BL' in dirpath:
                    is_cipl_file = True

                # If in 발주 서류 관리 and is xlsx/pdf with CI/PL content
                if '발주 서류 관리' in dirpath:
                    if any(kw in fl for kw in ['cipl', 'ci pl', 'ci&pl', 'ci^0pl', 'ci_', 'pl_',
                                                'receipt_import']):
                        is_cipl_file = True
                    # Also include files that look like order/shipment docs
                    if ext == '.xlsx' and any(kw in fl for kw in ['_zzbb_', '_sea_', '_air_']):
                        is_cipl_file = True

                # Check Downloads/OneDrive for specific CI/PL files
                if not deep_search:
                    if any(kw in fl for kw in ['cipl', 'ci pl', 'ci&pl', 'ci_pl']):
                        is_cipl_file = True

                if not is_cipl_file:
                    continue

                # Skip non-CI/PL files that matched loosely
                skip_keywords = ['라벨', 'label', '파렛트', 'pallet', 'carton', '카톤',
                                 'receipt_import', 'ex price', 'ex_price', 'quotation',
                                 '수출필증', '발주_', '발주 ', 'order', 'wholesale',
                                 '영문성적서', '사업자등록', '환급', '보험', 'insurance',
                                 '청구서', '계약서', 'agreement', '매뉴얼']
                if any(kw in fl for kw in skip_keywords) and not any(kw in fl for kw in ['cipl', 'ci pl', 'ci&pl', 'ci^0pl', 'ci_', 'pl_']):
                    continue

                # Get modification time for dedup
                try:
                    mtime = os.path.getmtime(full_path)
                except:
                    mtime = 0

                # Dedup: normalize filename for comparison
                base_name = re.sub(r'\s*\(\d+\)\s*', '', f)  # Remove (1), (2) etc
                base_name = re.sub(r'_\d{6}$', '', os.path.splitext(base_name)[0])  # Remove timestamp suffix
                dedup_key = base_name.lower()

                if dedup_key in seen_names:
                    # Keep the more recent one
                    if mtime > seen_names[dedup_key]['mtime']:
                        seen_names[dedup_key] = {'path': full_path, 'mtime': mtime}
                else:
                    seen_names[dedup_key] = {'path': full_path, 'mtime': mtime}

    # Return deduplicated file list
    return [v['path'] for v in seen_names.values()]


def main():
    print("=" * 80)
    print("Export Document Parser - Collecting CI/PL files...")
    print("=" * 80)

    files = collect_cipl_files()
    print(f"\nFound {len(files)} CI/PL files after deduplication\n")

    results = []
    errors = []

    for fp in sorted(files):
        ext = os.path.splitext(fp)[1].lower()
        print(f"  Parsing: {os.path.basename(fp)[:60]}...", end=" ")

        if ext in ('.xlsx', '.xls'):
            data = parse_excel_cipl(fp)
        elif ext == '.pdf':
            data = parse_pdf_cipl(fp)
        else:
            continue

        # Enrich brand detection
        if not data.get('brand'):
            data['brand'] = detect_brand_from_path(fp)
        if not data.get('brand') and data.get('exporter'):
            data['brand'] = detect_brand_from_exporter(data['exporter'])

        # Fix "Mixed" brand by checking exporter or content
        if data.get('brand') in ('Mixed', None) and data.get('exporter'):
            better = detect_brand_from_exporter(data['exporter'])
            if better:
                data['brand'] = better

        # Clean importer: remove "CGETC INC." prefix if followed by actual buyer
        if data.get('importer'):
            imp = data['importer']
            # Split on common patterns
            if 'CGETC' in imp and ('Fleeters' in imp or 'Orbiters' in imp):
                data['importer'] = imp  # Keep as is, shows the chain
            # Remove duplicate company names
            parts = re.split(r'\s{2,}', imp)
            if len(parts) > 1:
                data['importer'] = parts[0].strip()

        # Default destination
        if not data.get('destination') or data.get('destination') == 'Ultimate Consignee':
            data['destination'] = 'USA'

        if data.get('error'):
            print(f"SKIP ({data['error']})")
            errors.append({'file': fp, 'error': data['error']})
        else:
            print(f"OK - {data.get('brand', '?')} | {data.get('date', '?')} | {data.get('total_amount', '?')}")
            results.append(data)

    # ─── Cross-format deduplication ───
    # Group by shipment key (date + brand + shipping method)
    seen_shipments = {}
    deduped = []
    for r in results:
        # Create shipment key from file
        ship_key = normalize_shipment_key(r['source'], r.get('date'))
        if not ship_key:
            # Fallback: date + brand + shipping
            ship_key = f"{r.get('date', '')}_{r.get('brand', '')}_{r.get('shipping', '')}"

        if ship_key in seen_shipments:
            existing = seen_shipments[ship_key]
            # Prefer Excel over PDF (more data), and prefer file with more info
            existing_score = sum([
                bool(existing.get('total_amount')),
                bool(existing.get('items_summary')),
                len(existing.get('items_summary', [])) > 0,
                existing['source'].endswith('.xlsx'),
            ])
            new_score = sum([
                bool(r.get('total_amount')),
                bool(r.get('items_summary')),
                len(r.get('items_summary', [])) > 0,
                r['source'].endswith('.xlsx'),
            ])
            if new_score > existing_score:
                # Replace with better version
                idx = deduped.index(existing)
                deduped[idx] = r
                seen_shipments[ship_key] = r
            # If same score, keep existing (skip new)
        else:
            seen_shipments[ship_key] = r
            deduped.append(r)

    print(f"\n  Deduplication: {len(results)} → {len(deduped)} unique shipments")
    results = deduped

    # Sort by date descending
    results.sort(key=lambda x: x.get('date') or '0000-00', reverse=True)

    # Save results to Excel
    from output_utils import get_output_path, DATA_STORAGE
    output_dir = os.path.join(DATA_STORAGE, "export")
    os.makedirs(output_dir, exist_ok=True)
    output_path = get_output_path("export", "export_summary")

    wb_out = openpyxl.Workbook()

    # ─── Sheet 1: Summary ───
    ws = wb_out.active
    ws.title = "Export Summary"

    headers = ['년월', '브랜드', '운송방식', '금액', '통화', 'Exporter', 'Importer',
               'Destination', 'Price Terms', 'Invoice No', 'Total Qty', 'CBM',
               'Items Count', '파일출처']

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, data in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=data.get('date', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=data.get('brand', '')).border = thin_border
        ws.cell(row=row_idx, column=3, value=data.get('shipping', '')).border = thin_border
        amt_cell = ws.cell(row=row_idx, column=4, value=data.get('total_amount'))
        amt_cell.border = thin_border
        if data.get('total_amount'):
            amt_cell.number_format = '#,##0'
        ws.cell(row=row_idx, column=5, value=data.get('currency', '')).border = thin_border
        ws.cell(row=row_idx, column=6, value=data.get('exporter', '')).border = thin_border
        ws.cell(row=row_idx, column=7, value=data.get('importer', '')).border = thin_border
        ws.cell(row=row_idx, column=8, value=data.get('destination', '')).border = thin_border
        ws.cell(row=row_idx, column=9, value=data.get('price_terms', '')).border = thin_border
        ws.cell(row=row_idx, column=10, value=data.get('invoice_no', '')).border = thin_border
        ws.cell(row=row_idx, column=11, value=data.get('total_qty', '')).border = thin_border
        cbm_cell = ws.cell(row=row_idx, column=12, value=data.get('cbm'))
        cbm_cell.border = thin_border
        if data.get('cbm'):
            cbm_cell.number_format = '#,##0.00'
        ws.cell(row=row_idx, column=13, value=len(data.get('items_summary', []))).border = thin_border
        ws.cell(row=row_idx, column=14, value=data.get('source', '')).border = thin_border

    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['N'].width = 60

    # ─── Sheet 2: Reference (detailed items per shipment) ───
    ws2 = wb_out.create_sheet("Reference - Items Detail")
    ref_headers = ['년월', '브랜드', '운송방식', 'Invoice No', 'Item Description']
    for col, h in enumerate(ref_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ref_row = 2
    for data in results:
        for item in data.get('items_summary', []):
            ws2.cell(row=ref_row, column=1, value=data.get('date', '')).border = thin_border
            ws2.cell(row=ref_row, column=2, value=data.get('brand', '')).border = thin_border
            ws2.cell(row=ref_row, column=3, value=data.get('shipping', '')).border = thin_border
            ws2.cell(row=ref_row, column=4, value=data.get('invoice_no', '')).border = thin_border
            ws2.cell(row=ref_row, column=5, value=item).border = thin_border
            ref_row += 1

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 70

    # ─── Sheet 3: Errors / Skipped ───
    ws3 = wb_out.create_sheet("Skipped Files")
    ws3.cell(row=1, column=1, value="File Path").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Reason").font = Font(bold=True)
    for i, err in enumerate(errors, 2):
        ws3.cell(row=i, column=1, value=err['file'])
        ws3.cell(row=i, column=2, value=err['error'])
    ws3.column_dimensions['A'].width = 80
    ws3.column_dimensions['B'].width = 30

    # Save
    wb_out.save(output_path)
    print(f"\n{'=' * 80}")
    print(f"DONE! Saved to: {output_path}")
    print(f"  - {len(results)} shipments extracted")
    print(f"  - {len(errors)} files skipped")
    print(f"  - Sheet 1: Export Summary (main data)")
    print(f"  - Sheet 2: Reference - Items Detail")
    print(f"  - Sheet 3: Skipped Files")
    print(f"{'=' * 80}")

    # Also save raw JSON for Claude reference
    json_path = os.path.join(DATA_STORAGE, "export", "Export_Document_Data.json")
    json_data = []
    for r in results:
        entry = {k: v for k, v in r.items() if k != 'items_summary'}
        entry['items_count'] = len(r.get('items_summary', []))
        entry['items'] = r.get('items_summary', [])
        json_data.append(entry)

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    print(f"  - JSON reference saved to: {json_path}")


if __name__ == '__main__':
    main()
