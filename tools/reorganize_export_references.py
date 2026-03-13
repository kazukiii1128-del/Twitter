"""
Reorganize export document reference files.
- Abbreviate entity names: LFU, ORBI, FLT, etc.
- Extract and add Final Consignee (WBF, CGETC, Shipbob, DTrans)
- Remove amount from filename
- Organize into folders by buyer/importer (LFU/ vs FLT/)

Note: buyer = importer (same meaning in this context)

Usage:
  python tools/reorganize_export_references.py

Requires: Export_Document_Data.json (from tools/parse_export_documents.py)
"""
import json
import os
import shutil
import re
import openpyxl
import pdfplumber

# ─── Entity abbreviations ───
ENTITY_ABBREV = {
    'littlefinger': 'LFU',
    'lfu': 'LFU',
    'orbiters': 'ORBI',
    'orbi': 'ORBI',
    'fleeters': 'FLT',
    'flt': 'FLT',
    'cgetc': 'CGETC',
    'naeiae': 'Naeiae',
    'commemoi': 'Commemoi',
    'jtomorrow': 'Alpremio',
    'alpremio': 'Alpremio',
    'babyrabbit': 'BabyRabbit',
    'bamboo': 'BambooBebe',
    'conys': 'Conys',
    '코니코프': 'Conys',
    'klemarang': 'NLM',
}

CONSIGNEE_ABBREV = {
    'walk by faith': 'WBF',
    'cgetc': 'CGETC',
    'shipbob': 'Shipbob',
    'dtrans': 'DTrans',
    'dream tran': 'DTrans',
}

CONSIGNEE_BAD_KEYWORDS = [
    'terms of', 'yongin', 'if other', 'incoterms',
    'fob', 'cif', 'exw', 'payment',
]


def abbrev(name):
    """Abbreviate entity name (exporter/importer/buyer)."""
    if not name:
        return 'Unknown'
    n = name.lower()
    for keyword, short in ENTITY_ABBREV.items():
        if keyword in n:
            return short
    return name[:15].replace(' ', '_')


def abbrev_consignee(name):
    """Abbreviate Final Consignee name."""
    if not name:
        return ''
    n = name.lower()
    if any(b in n for b in CONSIGNEE_BAD_KEYWORDS):
        return ''
    for keyword, short in CONSIGNEE_ABBREV.items():
        if keyword in n:
            return short
    return name[:10].replace(' ', '_')


def extract_final_consignee_excel(filepath):
    """Extract Ultimate Consignee from Excel CI/PL."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        for name in wb.sheetnames:
            test_ws = wb[name]
            for row in test_ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if cell and 'COMMERCIAL INVOICE' in str(cell).upper():
                        ws = test_ws
                        break
            if 'Updated' in name or 'CI' in name.upper():
                ws = wb[name]

        cells = {}
        for row in ws.iter_rows(min_row=1, max_row=25, max_col=20):
            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = cell.value

        for coord, val in cells.items():
            if val and 'Ultimate Consignee' in str(val):
                row_num = int(re.search(r'(\d+)', coord).group(1))
                col_letter = re.match(r'([A-Z]+)', coord).group(1)
                for next_row in range(row_num + 1, row_num + 3):
                    for col in [col_letter, chr(ord(col_letter[0])+1) if len(col_letter)==1 else col_letter]:
                        nc = f"{col}{next_row}"
                        if nc in cells and cells[nc]:
                            v = str(cells[nc]).strip()
                            if v and len(v) > 2 and 'tel' not in v.lower() and '@' not in v:
                                wb.close()
                                return v
                break

        for coord, val in cells.items():
            if val and 'Final Destination' in str(val):
                row_num = int(re.search(r'(\d+)', coord).group(1))
                col_letter = re.match(r'([A-Z]+)', coord).group(1)
                for next_row in range(row_num + 1, row_num + 3):
                    for col_off in range(3):
                        try:
                            col = chr(ord(col_letter[0]) + col_off) if len(col_letter)==1 else col_letter
                        except:
                            continue
                        nc = f"{col}{next_row}"
                        if nc in cells and cells[nc]:
                            v = str(cells[nc]).strip()
                            if v and len(v) > 2 and 'tel' not in v.lower() and '@' not in v:
                                wb.close()
                                return v
                break

        wb.close()
    except:
        pass
    return None


def extract_final_consignee_pdf(filepath):
    """Extract Ultimate/Final Consignee from PDF."""
    try:
        with pdfplumber.open(filepath) as pdf:
            text = ""
            for page in pdf.pages[:2]:
                t = page.extract_text()
                if t:
                    text += t + "\n"

        lines = text.split('\n')
        for i, line in enumerate(lines):
            if 'Ultimate Consignee' in line or 'Final Destination' in line:
                for j in range(i+1, min(i+4, len(lines))):
                    candidate = lines[j].strip()
                    if candidate and len(candidate) > 3:
                        if not any(kw in candidate.lower() for kw in ['tel', 'fax', 'email', '@', 'usa', 'japan']):
                            parts = re.split(r'\s{2,}', candidate)
                            for p in parts:
                                p = p.strip()
                                if len(p) > 3 and not p.startswith('5900') and not p.startswith('18400') and not p.startswith('30 N'):
                                    return p
                break

        for line in lines:
            if 'Walk by Faith' in line:
                return 'Walk by Faith'
            if 'CGETC' in line:
                return 'CGETC Inc.'
    except:
        pass
    return None


def safe_fn(s, max_len=25):
    if not s:
        return 'Unknown'
    s = re.sub(r'[<>:"/\\|?*]', '', str(s))
    s = re.sub(r'[\s,]+', '_', s).strip('._')
    return s[:max_len]


def main():
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(base_path, 'Data Storage', 'Export Document')
    json_path = os.path.join(data_dir, 'Export_Document_Data.json')

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    ref_dir = os.path.join(data_dir, 'reference')

    # Clean old reference folder
    if os.path.exists(ref_dir):
        def onerror(func, path, exc_info):
            import stat
            os.chmod(path, stat.S_IWRITE)
            try:
                func(path)
            except:
                print(f"  WARN: Could not delete {os.path.basename(path)}")
        shutil.rmtree(ref_dir, onexc=onerror)

    # Create buyer/importer folders
    lfu_dir = os.path.join(ref_dir, 'LFU')
    flt_dir = os.path.join(ref_dir, 'FLT')
    os.makedirs(lfu_dir, exist_ok=True)
    os.makedirs(flt_dir, exist_ok=True)

    copied = 0
    for d in data:
        src = d.get('source', '')
        if not os.path.exists(src):
            print(f"  SKIP (missing): {os.path.basename(src)}")
            continue

        ext = os.path.splitext(src)[1]

        # Extract Final Consignee
        if ext in ('.xlsx', '.xls'):
            fc = extract_final_consignee_excel(src)
        else:
            fc = extract_final_consignee_pdf(src)

        fc_short = abbrev_consignee(fc) if fc else ''

        # Build parts
        date = d.get('date') or 'Unknown'
        brand = safe_fn(d.get('brand'), 15)
        shipping = d.get('shipping') or 'UNKNOWN'
        exporter = abbrev(d.get('exporter'))
        importer = abbrev(d.get('importer'))

        dest_raw = d.get('destination') or 'USA'
        if any(kw in dest_raw.lower() for kw in ['terms of', 'if other', 'incoterms', 'unknown']):
            dest_raw = 'USA'
        destination = safe_fn(dest_raw, 10)

        # Route to folder by buyer/importer
        if importer == 'LFU':
            target_dir = lfu_dir
        else:
            target_dir = flt_dir

        # Filename: 년월_브랜드_운송방식_Exporter_Importer_FinalConsignee_Destination
        parts = [date, brand, shipping, exporter, importer]
        if fc_short:
            parts.append(fc_short)
        parts.append(destination)

        new_name = '_'.join(parts) + ext

        dest_path = os.path.join(target_dir, new_name)
        counter = 1
        while os.path.exists(dest_path):
            base_n, ext2 = os.path.splitext(new_name)
            dest_path = os.path.join(target_dir, f"{base_n}_{counter}{ext2}")
            counter += 1

        try:
            shutil.copy2(src, dest_path)
            folder_label = 'LFU' if target_dir == lfu_dir else 'FLT'
            fc_label = f" [{fc_short}]" if fc_short else ""
            print(f"  [{folder_label}] {new_name}{fc_label}")
            copied += 1
        except Exception as e:
            print(f"  ERROR: {e}")

    lfu_count = len(os.listdir(lfu_dir)) if os.path.exists(lfu_dir) else 0
    flt_count = len(os.listdir(flt_dir)) if os.path.exists(flt_dir) else 0
    print(f"\nDone! {copied} files")
    print(f"  LFU/ : {lfu_count} files")
    print(f"  FLT/ : {flt_count} files")


if __name__ == '__main__':
    main()
