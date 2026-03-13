"""
WAT Framework Summary Excel Generator
======================================
Generates a 5-tab Excel summarizing all 17 workflows, 18 tools,
13 connected platforms, output methods, and data flows.

Usage:
    python tools/generate_wat_summary.py
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "Data Storage" / "misc"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH = OUTPUT_DIR / "WAT_Framework_Summary.xlsx"

# ── styling ────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

TAB_COLORS = {
    "Workflows": "2F5496",
    "Tools": "548235",
    "Platforms": "BF8F00",
    "Outputs": "C55A11",
    "Data Flows": "7030A0",
}


# ── data ───────────────────────────────────────────────────────────────

WORKFLOWS = [
    {
        "name": "polar_financial_model",
        "summary": "Build a 26-month financial model Excel from Polar data",
        "detail": "Pulls revenue, ads, and product data from Polar Analytics to build a comprehensive financial model by brand/channel/product. Covers 26 months (Jan 2024 - Feb 2026) in a hierarchical structure (Channel > Brand > Product) with SUM formulas for automatic rollup. Includes ad spend, margins, and influencer cost tracking.",
        "tools": "polar_financial_model.py\n(optional) fetch_meta_campaign_ids.py\n(optional) fetch_influencer_orders.py\n(optional) fetch_paypal_transactions.py",
        "platforms": "Polar Analytics, Meta Graph API, Shopify, PayPal",
        "output": "Data Storage/Polar_Financial_Model.xlsx (8 tabs)",
        "inputs": ".tmp/polar_data/ 8-11 JSON files (generated via Polar MCP)",
        "cost": "Free (Polar MCP, no API fees)",
    },
    {
        "name": "polar_dashboard_builder",
        "summary": "Build a 6-tab analytics dashboard Excel from Polar data",
        "detail": "Takes 8 Polar MCP query outputs and organizes them into a 6-tab dashboard: Sales, Ad Spend, Organic, Margin, Product, and Data Status. Shows key metrics like ROAS, CPC, AOV by brand and channel at a glance.",
        "tools": "polar_dashboard_builder.py",
        "platforms": "Polar Analytics",
        "output": "Data Storage/Polar_Dashboard.xlsx (6 tabs)",
        "inputs": ".tmp/polar_data/ 8 JSON files",
        "cost": "Free",
    },
    {
        "name": "weekly_performance_notion",
        "summary": "Auto-generate weekly performance report in Notion",
        "detail": "Every week, pulls ad performance (Meta/Google/TikTok/Amazon), D2C sales, GA4, and Klaviyo data from Polar, calculates OKR achievement rates, and auto-creates a weekly report page in Notion. Targets: ROAS 3.0, CAC $25, Email Open Rate 50%. Auto-flags On Track / At Risk / Behind status.",
        "tools": "weekly_performance_notion.py",
        "platforms": "Polar Analytics, Notion",
        "output": "Notion database page (weekly report)",
        "inputs": ".tmp/weekly_polar_data/ 7 JSON files",
        "cost": "Free",
    },
    {
        "name": "klaviyo_email_dashboard",
        "summary": "Analyze Klaviyo email performance in a 5-tab dashboard",
        "detail": "Pulls Klaviyo Flow and Campaign data from Polar and analyzes performance by flow type (Welcome, Cart Recovery, Post-Purchase, etc.) and campaign type (Promo, Product Launch, Seasonal, etc.). Benchmarks: Open Rate >= 50%, Click Rate >= 3%, Order Rate >= 2%.",
        "tools": "klaviyo_email_dashboard.py",
        "platforms": "Polar Analytics (Klaviyo data)",
        "output": ".tmp/klaviyo_email_dashboard_{date}.xlsx (5 tabs)",
        "inputs": ".tmp/polar_data/ 2 JSON files (Flow, Campaign)",
        "cost": "Free",
    },
    {
        "name": "classify_polar_ads",
        "summary": "Auto-classify Facebook ads by brand and product category",
        "detail": "Classifies 718 rows of Polar ad data into brands (Grosmimi, CHA&MOM, etc. — 11 brands) and categories. Uses priority-based keyword matching: Ad Name > Adset Name > Landing URL > Campaign Name. Also does Shopify product handle-to-SKU lookup for URL-based classification.",
        "tools": "classify_polar_ads.py",
        "platforms": "Shopify (product lookup), Facebook Ads (cached data)",
        "output": "Overwrites columns I (Brand) and J (Category) in source Excel",
        "inputs": "Polar Ads Raw Excel + Product Variant Excel + .tmp/facebook_ads.xlsx",
        "cost": "Free (Shopify API only)",
    },
    {
        "name": "fetch_keyword_volume",
        "summary": "Fetch keyword search volume from Google Ads + Amazon",
        "detail": "Uses DataForSEO API to retrieve monthly search volume, CPC, and competition for Google Ads, plus Amazon search volume — all in one run. Handles up to 1,000 keywords at once. Supports 12+ countries (US, UK, DE, etc.).",
        "tools": "fetch_keyword_volume.py",
        "platforms": "DataForSEO API (Google Ads + Amazon)",
        "output": ".tmp/fetch_keyword_volume/keyword_volume_{date}.xlsx (3 tabs)",
        "inputs": "Keyword list (comma-separated or file)",
        "cost": "~$0.09 per 100 keywords (pay-as-you-go)",
    },
    {
        "name": "process_influencer_order",
        "summary": "Process influencer gifting form into Shopify customer + draft order",
        "detail": "When an influencer submits a gifting form, this creates (or updates) a Shopify customer account and auto-generates a free draft order with selected products. Stores Instagram/TikTok handles and baby birthdays as customer metafields.",
        "tools": "process_influencer_order.py",
        "platforms": "Shopify (customer + orders), n8n (webhook)",
        "output": "Shopify customer account + Draft Order created",
        "inputs": "JSON payload (influencer info, products, shipping address)",
        "cost": "Free",
    },
    {
        "name": "influencer_typeform",
        "summary": "Auto-generate Typeform influencer gifting application form",
        "detail": "Creates a multi-step Typeform with age-based logic jumps that recommend different products by baby age range. Uploads product images to Typeform CDN. Steps: Personal Info > Baby Info > Product Selection > Shipping > Terms.",
        "tools": "create_typeform_influencer.py",
        "platforms": "Typeform API, Shopify CDN (product images)",
        "output": "Typeform survey URL",
        "inputs": "Product data (hardcoded in script)",
        "cost": "Free (within Typeform plan)",
    },
    {
        "name": "deploy_influencer_page",
        "summary": "Deploy influencer gifting page on Shopify storefront",
        "detail": "Deploys a custom influencer gifting application page directly to the Shopify store. Built with Liquid + CSS + JavaScript as a 5-step form (Personal > Baby > Product > Shipping > Terms). Submits data to n8n webhook.",
        "tools": "deploy_influencer_page.py",
        "platforms": "Shopify (theme assets + pages), n8n (webhook)",
        "output": "https://onzenna.com/pages/influencer-gifting",
        "inputs": "Shopify theme ID, product data, webhook URL",
        "cost": "Free",
    },
    {
        "name": "sync_influencer_notion",
        "summary": "Sync influencer Google Sheet to Notion database",
        "detail": "Syncs influencer master data from Google Sheets to a Notion database. Matches existing pages by handle or name, then creates new records or updates existing ones. Auto-tracks brands, platforms, and collaboration count.",
        "tools": "sync_influencer_notion.py",
        "platforms": "Google Sheets, Notion",
        "output": "Notion DB updated + .tmp/ sync report Excel",
        "inputs": "Google Sheet ID, Notion DB ID",
        "cost": "Free",
    },
    {
        "name": "gorgias_cs_template_builder",
        "summary": "Analyze Gorgias tickets with AI to generate CS reply templates",
        "detail": "Fetches last 6 months of closed Gorgias tickets, classifies them with Claude AI (Shipping, Refunds, Product Issues, etc. — 9 categories), and auto-generates standardized reply templates per category. Outputs to Google Sheets with macro name, recommended reply, checklist, and forbidden phrases.",
        "tools": "gorgias_cs_template_builder.py",
        "platforms": "Gorgias API, Claude API (Haiku + Sonnet), Google Sheets",
        "output": "Google Sheets (CS_Template_Library) + .tmp/ Excel backup",
        "inputs": "Gorgias credentials, --months parameter",
        "cost": "~$0.25 (Claude API, ~100 tickets)",
    },
    {
        "name": "gmail_affiliate_faq_builder",
        "summary": "Analyze affiliate Gmail inquiries to auto-generate FAQ",
        "detail": "Reads last 3 months of emails from affiliates@onzenna.com, classifies them with Claude AI (How to Join, Commission, Shipping, etc. — 11 categories), and generates FAQ entries per category. High Touch / Low Touch partnership distinctions are kept in internal notes only.",
        "tools": "gmail_affiliate_faq_builder.py",
        "platforms": "Gmail API (OAuth), Claude API (Haiku + Sonnet), Google Sheets",
        "output": "Google Sheets (Affiliate_FAQ) + .tmp/ Excel backup",
        "inputs": "Gmail OAuth credentials, --months parameter",
        "cost": "~$0.25 (Claude API)",
    },
    {
        "name": "scrape_job_listings",
        "summary": "Scrape DailyRemote job listings to Excel",
        "detail": "Scrapes remote support job listings from DailyRemote (full-time, maternity benefits filter). Extracts job title, company, salary, experience, skills, etc. (12 columns). Parallelized detail page crawling with 10 workers.",
        "tools": "scrape_job_listings.py",
        "platforms": "Firecrawl API, DailyRemote.com",
        "output": ".tmp/job_listings_{date}.xlsx",
        "inputs": "URL + filter parameters",
        "cost": "Firecrawl credits",
    },
    {
        "name": "export_document_management",
        "summary": "Parse export CI/PL documents and reorganize folders",
        "detail": "Auto-parses Commercial Invoice (CI) and Packing List (PL) PDF/Excel documents scattered across Z: drive folders. Extracts date, brand, shipping method, amount, exporter/importer, etc. Generates an Excel summary and reorganizes files into buyer-specific folders.",
        "tools": "parse_export_documents.py\nreorganize_export_references.py",
        "platforms": "File system (Z: drive)",
        "output": "Data Storage/export/Export_Summary.xlsx (3 tabs)\n+ reference/ folder structure",
        "inputs": "Z:\\Orbiters\\CI,PL,BL source folders",
        "cost": "Free",
    },
    {
        "name": "data_output_management",
        "summary": "Reference guide for data output folder/naming conventions",
        "detail": "A reference document defining standard storage rules for all workflow outputs. Final files go to Data Storage/{workflow}/, intermediate files to .tmp/. File naming: {base}_{date}_v{N}.xlsx with auto version increment.",
        "tools": "output_utils.py (shared utility)",
        "platforms": "File system",
        "output": "N/A (reference document)",
        "inputs": "N/A",
        "cost": "N/A",
    },
    {
        "name": "gorgias_analyze",
        "summary": "Analyze high-message Gorgias tickets with AI keyword extraction",
        "detail": "Fetches top N Gorgias tickets with the most messages. Extracts order numbers, full conversation transcripts, and uses Claude AI to identify top 5 complaint keywords. Primarily used for analyzing high-touch / wholesale inquiries.",
        "tools": "gorgias_analyze.py (in workflows/)",
        "platforms": "Gorgias API, Claude API (Haiku)",
        "output": ".tmp/gorgias_analysis_{date}.xlsx",
        "inputs": "--min-messages, --top, --status parameters",
        "cost": "~$0.05 (Claude Haiku)",
    },
    {
        "name": "gorgias_fetch",
        "summary": "Query Gorgias tickets/customer data and export to Excel",
        "detail": "Queries Gorgias REST API for ticket lists (by status) or customer details and exports to Excel. Ticket output includes 12 columns: ID, Status, Subject, Channel, Assignee, Customer, Tags, Message Count, etc. Customer search by email or ID.",
        "tools": "gorgias_fetch.py (in workflows/)",
        "platforms": "Gorgias API",
        "output": ".tmp/gorgias_tickets_{status}_{date}.xlsx or gorgias_customer_{date}.xlsx",
        "inputs": "--tickets or --customer + filter options",
        "cost": "Free",
    },
]

TOOLS = [
    {
        "name": "polar_financial_model.py",
        "summary": "Generate 26-month financial model Excel (revenue/ads/margin/influencer)",
        "apis": "None (reads local JSON files)",
        "inputs": ".tmp/polar_data/ 8-11 JSON files",
        "outputs": "Data Storage/Polar_Financial_Model.xlsx (8 tabs, with formulas)",
        "notes": "149KB large script. Handles 11 brands, 6 channels, 4 ad platforms",
    },
    {
        "name": "polar_dashboard_builder.py",
        "summary": "Build 6-tab analytics dashboard Excel from Polar data",
        "apis": "None (reads local JSON files)",
        "inputs": ".tmp/polar_data/ 8 JSON files",
        "outputs": "Data Storage/Polar_Dashboard.xlsx (6 tabs)",
        "notes": "Tabs: Sales, Ad Spend, Organic, Margin, Product, Data Status",
    },
    {
        "name": "weekly_performance_notion.py",
        "summary": "Generate weekly OKR report in Notion from Polar data",
        "apis": "Notion API",
        "inputs": ".tmp/weekly_polar_data/ 7 JSON files",
        "outputs": "Notion database page",
        "notes": "OKR targets: ROAS 3.0, CAC $25, Email Open Rate 50%",
    },
    {
        "name": "klaviyo_email_dashboard.py",
        "summary": "Build 5-tab Klaviyo Flow/Campaign performance dashboard",
        "apis": "None (reads local JSON files)",
        "inputs": ".tmp/polar_data/ 2 JSON files (kl1, kl2)",
        "outputs": ".tmp/klaviyo_email_dashboard_{date}.xlsx (5 tabs)",
        "notes": "Benchmarks: Open Rate >= 50%, Click Rate >= 3%, Order Rate >= 2%",
    },
    {
        "name": "classify_polar_ads.py",
        "summary": "Auto-classify Facebook ads by brand and product category",
        "apis": "Shopify REST API (product handle > SKU lookup)",
        "inputs": "Polar Ads Raw Excel + Product Variant Excel + facebook_ads.xlsx",
        "outputs": "Overwrites columns I/J in source Excel (718 rows)",
        "notes": "4-level priority matching: Ad Name > Adset > Landing URL > Campaign",
    },
    {
        "name": "fetch_keyword_volume.py",
        "summary": "Fetch Google + Amazon keyword search volume",
        "apis": "DataForSEO v3 (Google Ads + Amazon)",
        "inputs": "Keyword list (up to 1,000 keywords)",
        "outputs": ".tmp/fetch_keyword_volume/keyword_volume_{date}.xlsx (3 tabs)",
        "notes": "Cost ~$0.09/100 keywords. 12 countries supported. Rate limit: 12 req/min",
    },
    {
        "name": "fetch_facebook_ads.py",
        "summary": "Fetch Facebook campaign performance + landing URLs",
        "apis": "Meta Graph API v18.0",
        "inputs": "META_ACCESS_TOKEN, META_AD_ACCOUNT_ID",
        "outputs": ".tmp/facebook_ads.xlsx",
        "notes": "Includes impressions, clicks, spend, CTR, CPC, CPM, purchases",
    },
    {
        "name": "fetch_meta_campaign_ids.py",
        "summary": "Fetch Meta campaign name-to-ID mapping for direct ad links",
        "apis": "Meta Graph API v18.0",
        "inputs": "META_ACCESS_TOKEN, META_AD_ACCOUNT_ID",
        "outputs": ".tmp/polar_data/q9_meta_campaign_ids.json",
        "notes": "Auto-handles pagination. Used to generate direct campaign links",
    },
    {
        "name": "fetch_influencer_orders.py",
        "summary": "Extract influencer/PR/sample orders from Shopify",
        "apis": "Shopify REST API 2024-01",
        "inputs": "SHOPIFY_SHOP, SHOPIFY_ACCESS_TOKEN",
        "outputs": ".tmp/polar_data/q10_influencer_orders.json",
        "notes": "2-pass: tag-based queries (fast) + full order scan (thorough). 8 tags",
    },
    {
        "name": "fetch_paypal_transactions.py",
        "summary": "Fetch PayPal transactions for influencer payment matching",
        "apis": "PayPal v1 API (OAuth2)",
        "inputs": "PAYPAL_CLIENT_ID, PAYPAL_SECRET",
        "outputs": ".tmp/polar_data/q11_paypal_transactions.json",
        "notes": "Iterates in 31-day windows (API limit). Jan 2024 to present",
    },
    {
        "name": "process_influencer_order.py",
        "summary": "Process influencer gifting submission into Shopify customer + draft order",
        "apis": "Shopify REST API 2024-01 (customers/orders)",
        "inputs": "JSON payload (personal info, products, address)",
        "outputs": "Shopify customer + order creation result JSON",
        "notes": "Duplicate email detection, stores metafields (Instagram/TikTok/birthdays)",
    },
    {
        "name": "influencer_customer_lookup.py",
        "summary": "Search Shopify customers by name for influencer lookup",
        "apis": "Shopify REST API 2024-01",
        "inputs": "--name parameter",
        "outputs": "JSON (customer profile + metafields + address)",
        "notes": "Used by n8n for influencer gifting page login flow",
    },
    {
        "name": "create_typeform_influencer.py",
        "summary": "Create Typeform influencer gifting survey with logic jumps",
        "apis": "Typeform API (forms + images + logic jumps)",
        "inputs": "TYPEFORM_API_KEY, product data (hardcoded)",
        "outputs": "Typeform survey URL + .tmp/typeform/last_form.json",
        "notes": "Age-based logic jumps. Product images uploaded via base64",
    },
    {
        "name": "deploy_influencer_page.py",
        "summary": "Deploy influencer gifting page to Shopify storefront",
        "apis": "Shopify Theme Asset API + Pages API",
        "inputs": "Theme ID, product data, n8n webhook URL",
        "outputs": "Shopify page (onzenna.com/pages/influencer-gifting)",
        "notes": "48KB. Includes Liquid + CSS + JS. Supports --dry-run, --rollback",
    },
    {
        "name": "sync_influencer_notion.py",
        "summary": "Sync influencer data from Google Sheets to Notion DB",
        "apis": "Google Sheets API (gspread), Notion API 2022-06-28",
        "inputs": "Sheet ID, Notion DB ID, Service Account JSON",
        "outputs": "Notion DB updated + .tmp/ sync report Excel",
        "notes": "Handle > name matching. 0.35s delay + retry. Handles 100+ pages",
    },
    {
        "name": "gorgias_cs_template_builder.py",
        "summary": "AI-analyze Gorgias tickets and generate CS reply templates",
        "apis": "Gorgias API, Claude API (Haiku + Sonnet), Google Sheets",
        "inputs": "Gorgias credentials, --months, ANTHROPIC_API_KEY",
        "outputs": "Google Sheets (CS_Template_Library) + .tmp/ Excel backup",
        "notes": "200KB. 9 problem / 7 resolution categories. ~$0.25 per 100 tickets",
    },
    {
        "name": "scrape_job_listings.py",
        "summary": "Scrape DailyRemote job listings to Excel",
        "apis": "Firecrawl v1 API",
        "inputs": "FIRECRAWL_API_KEY, URL + filters",
        "outputs": ".tmp/job_listings_{date}.xlsx (12 columns)",
        "notes": "10 parallel workers for detail page crawling. Duplicate detection",
    },
    {
        "name": "parse_export_documents.py\n+ reorganize_export_references.py",
        "summary": "Parse CI/PL export documents + reorganize folders",
        "apis": "None (file system only)",
        "inputs": "Z: drive CI/PL PDF/Excel files",
        "outputs": "Export_Summary.xlsx + reference/ folders",
        "notes": "pdfplumber for PDF parsing. Organizes by buyer (LFU/FLT folders)",
    },
]

PLATFORMS = [
    {
        "name": "Shopify",
        "purpose": "E-commerce store — product, customer, order, and theme management",
        "auth": "OAuth 2.0 (Access Token)",
        "env_keys": "SHOPIFY_SHOP, SHOPIFY_ACCESS_TOKEN",
        "workflows": "process_influencer_order, classify_polar_ads, deploy_influencer_page, polar_financial_model",
        "tools": "process_influencer_order.py, classify_polar_ads.py, deploy_influencer_page.py, influencer_customer_lookup.py, fetch_influencer_orders.py, shopify_oauth.py",
    },
    {
        "name": "Polar Analytics",
        "purpose": "Unified e-commerce analytics — blended revenue, ads, email data",
        "auth": "API Key (MCP connection)",
        "env_keys": "POLAR_API_KEY",
        "workflows": "polar_financial_model, polar_dashboard_builder, weekly_performance_notion, klaviyo_email_dashboard",
        "tools": "polar_financial_model.py, polar_dashboard_builder.py, weekly_performance_notion.py, klaviyo_email_dashboard.py (indirect — JSON consumer)",
    },
    {
        "name": "Notion",
        "purpose": "Project management — influencer DB, weekly reports, workspace",
        "auth": "Integration Token (Bearer)",
        "env_keys": "NOTION_API_TOKEN, INFLUENCER_NOTION_DB_ID",
        "workflows": "sync_influencer_notion, weekly_performance_notion",
        "tools": "sync_influencer_notion.py, weekly_performance_notion.py",
    },
    {
        "name": "Google Sheets",
        "purpose": "Cloud data storage/sharing — CS templates, FAQ, influencer master sheet",
        "auth": "Service Account (JSON key file)",
        "env_keys": "GOOGLE_SERVICE_ACCOUNT_PATH, AFFILIATE_FAQ_SHEET_ID, INFLUENCER_SHEET_ID",
        "workflows": "gorgias_cs_template_builder, gmail_affiliate_faq_builder, sync_influencer_notion",
        "tools": "gorgias_cs_template_builder.py, gmail_affiliate_faq_builder.py, sync_influencer_notion.py",
    },
    {
        "name": "Meta / Facebook Ads",
        "purpose": "Facebook/Instagram ad performance data retrieval, campaign ID mapping",
        "auth": "Graph API Access Token",
        "env_keys": "META_ACCESS_TOKEN, META_AD_ACCOUNT_ID",
        "workflows": "classify_polar_ads, polar_financial_model",
        "tools": "fetch_facebook_ads.py, fetch_meta_campaign_ids.py",
    },
    {
        "name": "Gorgias",
        "purpose": "Customer service — ticket analysis, template generation, customer lookup",
        "auth": "HTTP Basic Auth (email:API key)",
        "env_keys": "GORGIAS_DOMAIN, GORGIAS_EMAIL, GORGIAS_API_KEY",
        "workflows": "gorgias_cs_template_builder, gorgias_analyze, gorgias_fetch",
        "tools": "gorgias_cs_template_builder.py, gorgias_analyze.py, gorgias_fetch.py",
    },
    {
        "name": "Gmail",
        "purpose": "Affiliate inquiry email analysis (affiliates@onzenna.com)",
        "auth": "OAuth 2.0 (Desktop App)",
        "env_keys": "GMAIL_OAUTH_CREDENTIALS_PATH",
        "workflows": "gmail_affiliate_faq_builder",
        "tools": "gmail_affiliate_faq_builder.py",
    },
    {
        "name": "Claude API (Anthropic)",
        "purpose": "AI text analysis — ticket classification (Haiku), template generation (Sonnet)",
        "auth": "API Key",
        "env_keys": "ANTHROPIC_API_KEY",
        "workflows": "gorgias_cs_template_builder, gmail_affiliate_faq_builder, gorgias_analyze",
        "tools": "gorgias_cs_template_builder.py, gmail_affiliate_faq_builder.py, gorgias_analyze.py",
    },
    {
        "name": "Typeform",
        "purpose": "Influencer gifting application form creation",
        "auth": "API Key (Bearer)",
        "env_keys": "TYPEFORM_API_KEY",
        "workflows": "influencer_typeform",
        "tools": "create_typeform_influencer.py",
    },
    {
        "name": "n8n",
        "purpose": "Workflow automation — webhook-based form submission processing",
        "auth": "API Key + Webhook URL",
        "env_keys": "N8N_API_KEY, N8N_BASE_URL, N8N_INFLUENCER_WEBHOOK",
        "workflows": "process_influencer_order, deploy_influencer_page",
        "tools": "process_influencer_order.py, deploy_influencer_page.py",
    },
    {
        "name": "PayPal",
        "purpose": "Influencer payment history retrieval (paid vs unpaid classification)",
        "auth": "OAuth 2.0 (Client Credentials)",
        "env_keys": "PAYPAL_CLIENT_ID, PAYPAL_SECRET",
        "workflows": "polar_financial_model",
        "tools": "fetch_paypal_transactions.py",
    },
    {
        "name": "DataForSEO",
        "purpose": "Keyword search volume retrieval (Google Ads + Amazon)",
        "auth": "HTTP Basic Auth (ID:Password)",
        "env_keys": "DATAFORSEO_LOGIN, DATAFORSEO_PASSWORD",
        "workflows": "fetch_keyword_volume",
        "tools": "fetch_keyword_volume.py",
    },
    {
        "name": "Firecrawl",
        "purpose": "Web scraping — job listings and other web page data collection",
        "auth": "API Key",
        "env_keys": "FIRECRAWL_API_KEY",
        "workflows": "scrape_job_listings",
        "tools": "scrape_job_listings.py",
    },
]

OUTPUTS = [
    {
        "workflow": "polar_financial_model",
        "location": "Data Storage/",
        "format": "Excel (.xlsx)",
        "filename": "Polar_Financial_Model.xlsx",
        "temp_files": ".tmp/polar_data/ 8-11 JSON files",
        "cloud": "Local only",
    },
    {
        "workflow": "polar_dashboard_builder",
        "location": "Data Storage/",
        "format": "Excel (.xlsx)",
        "filename": "Polar_Dashboard.xlsx",
        "temp_files": ".tmp/polar_data/ 8 JSON files",
        "cloud": "Local only",
    },
    {
        "workflow": "weekly_performance_notion",
        "location": "Notion (cloud)",
        "format": "Notion database page",
        "filename": "Notion weekly report page",
        "temp_files": ".tmp/weekly_polar_data/ 7 JSON files",
        "cloud": "Cloud (Notion)",
    },
    {
        "workflow": "klaviyo_email_dashboard",
        "location": ".tmp/",
        "format": "Excel (.xlsx)",
        "filename": "klaviyo_email_dashboard_{date}.xlsx",
        "temp_files": ".tmp/polar_data/ 2 JSON files",
        "cloud": "Local only",
    },
    {
        "workflow": "classify_polar_ads",
        "location": "Data Storage/",
        "format": "Excel (.xlsx) — overwrites source",
        "filename": "Polar Ads Raw.xlsx (columns I/J)",
        "temp_files": ".tmp/facebook_ads.xlsx",
        "cloud": "Local only",
    },
    {
        "workflow": "fetch_keyword_volume",
        "location": ".tmp/fetch_keyword_volume/",
        "format": "Excel (.xlsx)",
        "filename": "keyword_volume_{date}.xlsx",
        "temp_files": "None",
        "cloud": "Local only",
    },
    {
        "workflow": "process_influencer_order",
        "location": "Shopify (cloud)",
        "format": "Shopify Customer + Draft Order",
        "filename": "N/A (created in Shopify)",
        "temp_files": ".tmp/process_influencer_order/order_result_{date}.json",
        "cloud": "Cloud (Shopify)",
    },
    {
        "workflow": "influencer_typeform",
        "location": "Typeform (cloud)",
        "format": "Typeform survey URL",
        "filename": "N/A (Typeform URL)",
        "temp_files": ".tmp/typeform/last_form.json",
        "cloud": "Cloud (Typeform)",
    },
    {
        "workflow": "deploy_influencer_page",
        "location": "Shopify (cloud)",
        "format": "Shopify web page",
        "filename": "onzenna.com/pages/influencer-gifting",
        "temp_files": "None",
        "cloud": "Cloud (Shopify)",
    },
    {
        "workflow": "sync_influencer_notion",
        "location": "Notion (cloud)",
        "format": "Notion DB page updates",
        "filename": "N/A (Notion DB)",
        "temp_files": ".tmp/sync_influencer_notion/sync_report_{date}.xlsx",
        "cloud": "Cloud (Notion)",
    },
    {
        "workflow": "gorgias_cs_template_builder",
        "location": "Google Sheets (cloud)",
        "format": "Google Sheets worksheet",
        "filename": "CS_Template_Library (sheet name)",
        "temp_files": ".tmp/gorgias_cs_template_builder/cs_templates_{date}.xlsx",
        "cloud": "Cloud (Google Sheets)",
    },
    {
        "workflow": "gmail_affiliate_faq_builder",
        "location": "Google Sheets (cloud)",
        "format": "Google Sheets worksheet",
        "filename": "Affiliate_FAQ (sheet name)",
        "temp_files": ".tmp/gmail_affiliate_faq/affiliate_faq_{date}.xlsx",
        "cloud": "Cloud (Google Sheets)",
    },
    {
        "workflow": "scrape_job_listings",
        "location": ".tmp/",
        "format": "Excel (.xlsx)",
        "filename": "job_listings_{date}.xlsx",
        "temp_files": "None",
        "cloud": "Local only",
    },
    {
        "workflow": "export_document_management",
        "location": "Data Storage/export/",
        "format": "Excel (.xlsx) + JSON + folders",
        "filename": "Export_Summary.xlsx + reference/ folder",
        "temp_files": "None",
        "cloud": "Local only",
    },
    {
        "workflow": "gorgias_analyze",
        "location": ".tmp/",
        "format": "Excel (.xlsx)",
        "filename": "gorgias_analysis_{date}.xlsx",
        "temp_files": "None",
        "cloud": "Local only",
    },
    {
        "workflow": "gorgias_fetch",
        "location": ".tmp/",
        "format": "Excel (.xlsx)",
        "filename": "gorgias_tickets_{status}_{date}.xlsx",
        "temp_files": "None",
        "cloud": "Local only",
    },
]

DATA_FLOWS = [
    {
        "order": 1,
        "step": "Polar > Financial Model",
        "source": "Polar Analytics MCP",
        "tool": "polar_financial_model.py",
        "temp": ".tmp/polar_data/ (Q1-Q8 JSON)",
        "final": "Data Storage/Polar_Financial_Model.xlsx",
        "link": "Optionally connects to Meta/Shopify/PayPal for additional data",
    },
    {
        "order": 2,
        "step": "Polar > Dashboard",
        "source": "Polar Analytics MCP",
        "tool": "polar_dashboard_builder.py",
        "temp": ".tmp/polar_data/ (Q1-Q8 JSON)",
        "final": "Data Storage/Polar_Dashboard.xlsx",
        "link": "Shares same source data as Financial Model",
    },
    {
        "order": 3,
        "step": "Polar > Weekly Report",
        "source": "Polar Analytics MCP",
        "tool": "weekly_performance_notion.py",
        "temp": ".tmp/weekly_polar_data/ (7 JSON)",
        "final": "Notion weekly report page",
        "link": "Runs weekly on a recurring basis. Auto-writes to Notion",
    },
    {
        "order": 4,
        "step": "Polar > Email Dashboard",
        "source": "Polar Analytics MCP (Klaviyo data)",
        "tool": "klaviyo_email_dashboard.py",
        "temp": ".tmp/polar_data/ (kl1, kl2 JSON)",
        "final": ".tmp/klaviyo_email_dashboard_{date}.xlsx",
        "link": "Accesses Klaviyo connector data through Polar",
    },
    {
        "order": 5,
        "step": "Facebook > Ad Classification",
        "source": "Meta Graph API > .tmp/facebook_ads.xlsx",
        "tool": "fetch_facebook_ads.py > classify_polar_ads.py",
        "temp": ".tmp/facebook_ads.xlsx (intermediate cache)",
        "final": "Polar Ads Raw Excel (columns I/J updated)",
        "link": "2-step chain: first extract FB data, then run classification",
    },
    {
        "order": 6,
        "step": "Export Docs > Organize",
        "source": "Z: drive CI/PL files",
        "tool": "parse_export_documents.py > reorganize_export_references.py",
        "temp": "None",
        "final": "Export_Summary.xlsx + reference/ folders",
        "link": "2-step chain: first parse documents, then reorganize folders",
    },
    {
        "order": 7,
        "step": "Influencer App > Order",
        "source": "Typeform / Shopify page > n8n webhook",
        "tool": "process_influencer_order.py",
        "temp": ".tmp/process_influencer_order/ JSON",
        "final": "Shopify Customer + Draft Order",
        "link": "Form submission > n8n > Shopify auto-processing",
    },
    {
        "order": 8,
        "step": "Influencer Sheet > Notion",
        "source": "Google Sheets (influencer master)",
        "tool": "sync_influencer_notion.py",
        "temp": ".tmp/sync_influencer_notion/ Excel",
        "final": "Notion Influencer DB",
        "link": "Google Sheets is source of truth (read-only), Notion is management layer",
    },
    {
        "order": 9,
        "step": "Gorgias > CS Templates",
        "source": "Gorgias API (closed tickets)",
        "tool": "gorgias_cs_template_builder.py",
        "temp": ".tmp/gorgias_cs_template_builder/ Excel",
        "final": "Google Sheets (CS_Template_Library)",
        "link": "Claude Haiku classifies > Sonnet synthesizes templates > uploads to Sheets",
    },
    {
        "order": 10,
        "step": "Gmail > Affiliate FAQ",
        "source": "Gmail API (affiliates@ inbox)",
        "tool": "gmail_affiliate_faq_builder.py",
        "temp": ".tmp/gmail_affiliate_faq/ Excel",
        "final": "Google Sheets (Affiliate_FAQ)",
        "link": "Claude Haiku classifies > Sonnet synthesizes FAQ > uploads to Sheets",
    },
]


# ── Excel builder ──────────────────────────────────────────────────────

def style_header(ws, headers, col_widths):
    """Apply header styling and set column widths."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_rows(ws, data, keys, start_row=2):
    """Write data rows with alternating fills."""
    for row_idx, item in enumerate(data, start_row):
        for col_idx, key in enumerate(keys, 1):
            val = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if (row_idx - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def build_excel():
    wb = Workbook()

    # ── Tab 1: Workflows ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Workflows"
    ws1.sheet_properties.tabColor = TAB_COLORS["Workflows"]
    headers1 = ["#", "Workflow Name", "Summary", "Detailed Description", "Tools Used", "Connected Platforms", "Final Output", "Required Inputs", "Cost"]
    widths1 = [4, 24, 38, 58, 30, 30, 38, 38, 20]
    style_header(ws1, headers1, widths1)

    for i, wf in enumerate(WORKFLOWS, 1):
        row_data = {"#": i, **wf}
        keys = ["#", "name", "summary", "detail", "tools", "platforms", "output", "inputs", "cost"]
        for col_idx, key in enumerate(keys, 1):
            val = row_data.get(key, "")
            cell = ws1.cell(row=i + 1, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL

    ws1.freeze_panes = "C2"
    ws1.auto_filter.ref = f"A1:I{len(WORKFLOWS) + 1}"

    # ── Tab 2: Tools ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Tools")
    ws2.sheet_properties.tabColor = TAB_COLORS["Tools"]
    headers2 = ["#", "Tool Name", "Summary", "Connected APIs / Services", "Inputs", "Outputs", "Notes"]
    widths2 = [4, 30, 42, 32, 38, 42, 38]
    style_header(ws2, headers2, widths2)

    for i, tool in enumerate(TOOLS, 1):
        row_data = {"#": i, **tool}
        keys = ["#", "name", "summary", "apis", "inputs", "outputs", "notes"]
        for col_idx, key in enumerate(keys, 1):
            val = row_data.get(key, "")
            cell = ws2.cell(row=i + 1, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL

    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = f"A1:G{len(TOOLS) + 1}"

    # ── Tab 3: Platforms ───────────────────────────────────────────
    ws3 = wb.create_sheet("Platforms")
    ws3.sheet_properties.tabColor = TAB_COLORS["Platforms"]
    headers3 = ["Platform", "Purpose", "Auth Method", ".env Keys", "Used By Workflows", "Used By Tools"]
    widths3 = [20, 42, 24, 38, 42, 48]
    style_header(ws3, headers3, widths3)

    keys3 = ["name", "purpose", "auth", "env_keys", "workflows", "tools"]
    for i, pf in enumerate(PLATFORMS, 1):
        for col_idx, key in enumerate(keys3, 1):
            val = pf.get(key, "")
            cell = ws3.cell(row=i + 1, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL

    ws3.freeze_panes = "B2"
    ws3.auto_filter.ref = f"A1:F{len(PLATFORMS) + 1}"

    # ── Tab 4: Outputs ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Outputs")
    ws4.sheet_properties.tabColor = TAB_COLORS["Outputs"]
    headers4 = ["Workflow", "Final Output Location", "File Format", "Filename Pattern", "Intermediate Files (.tmp/)", "Cloud / Local"]
    widths4 = [24, 26, 24, 38, 38, 20]
    style_header(ws4, headers4, widths4)

    keys4 = ["workflow", "location", "format", "filename", "temp_files", "cloud"]
    for i, out in enumerate(OUTPUTS, 1):
        for col_idx, key in enumerate(keys4, 1):
            val = out.get(key, "")
            cell = ws4.cell(row=i + 1, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL

    ws4.freeze_panes = "B2"
    ws4.auto_filter.ref = f"A1:F{len(OUTPUTS) + 1}"

    # ── Tab 5: Data Flows ──────────────────────────────────────────
    ws5 = wb.create_sheet("Data Flows")
    ws5.sheet_properties.tabColor = TAB_COLORS["Data Flows"]
    headers5 = ["#", "Step", "Data Source", "Tool(s)", "Intermediate (.tmp/)", "Final Output", "Notes / Connections"]
    widths5 = [4, 22, 32, 38, 32, 38, 44]
    style_header(ws5, headers5, widths5)

    keys5 = ["order", "step", "source", "tool", "temp", "final", "link"]
    for i, flow in enumerate(DATA_FLOWS, 1):
        for col_idx, key in enumerate(keys5, 1):
            val = flow.get(key, "")
            cell = ws5.cell(row=i + 1, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL

    ws5.freeze_panes = "C2"
    ws5.auto_filter.ref = f"A1:G{len(DATA_FLOWS) + 1}"

    # ── Save ───────────────────────────────────────────────────────
    wb.save(str(OUTPUT_PATH))
    print(f"[OK] Excel saved: {OUTPUT_PATH}")
    print(f"     Tab 1: Workflows ({len(WORKFLOWS)} rows)")
    print(f"     Tab 2: Tools ({len(TOOLS)} rows)")
    print(f"     Tab 3: Platforms ({len(PLATFORMS)} rows)")
    print(f"     Tab 4: Outputs ({len(OUTPUTS)} rows)")
    print(f"     Tab 5: Data Flows ({len(DATA_FLOWS)} rows)")


if __name__ == "__main__":
    build_excel()
