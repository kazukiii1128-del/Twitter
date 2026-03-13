"""
Generate a Consulting Fee Invoice (bank submission).
Fleeters Inc. issues invoice to Orbiters Co.,Ltd.

Usage:
  python tools/generate_consulting_invoice.py
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "Data Storage", "export")

# ---- Configuration ----
INVOICE_NO = "FLT-CONSULTING-2602-001"
INVOICE_DATE = datetime(2026, 2, 23)
AMOUNT = 60000.00
DESCRIPTION = "Consulting Fee"
CURRENCY = "USD"

# Issuer (Fleeters - service provider, receives payment)
ISSUER = {
    "name": "Fleeters Inc.",
    "address": "30 N Gould St. Ste 32663, Sheridan, WY 82801, USA",
    "tel": "+1 10-6209-3352",
    "email": "wj.choi@orbiters.co.kr",
    "signer": "Wonjun Choi",
    "signer_title": "President",
}

# Bill To (Orbiters - payer)
BILL_TO = {
    "name": "ORBITERS Co., Ltd.",
    "address": "Unit 509, 25, Ttukseom-ro 1-gil, Seongdong-gu, Seoul, Republic of Korea, 04778",
    "tel": "+82 10 6209 3352",
    "email": "wj.choi@orbiters.co.kr",
}

# Bank Info (Fleeters' receiving bank account)
BANK_INFO = {
    "remittee": "FLEETERS INC",
    "bank": "JPMorgan Chase Bank, N.A., Santa Monica, CA",
    "swift": "CHASUS33",
    "account": "522203317",
    "routing": "021000021",  # Wire transfer routing number
}

OUTPUT_FILE = os.path.join(OUTPUT_DIR, "2026-02_ConsultingFee_Invoice_FLT_to_ORBI.xlsx")


def create_invoice():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ---- Styles ----
    title_font = Font(name="Calibri", size=20, bold=True)
    section_font = Font(name="Calibri", size=11, bold=True)
    label_font = Font(name="Calibri", size=10, bold=True)
    data_font = Font(name="Calibri", size=10)
    small_font = Font(name="Calibri", size=9)
    amount_font = Font(name="Calibri", size=14, bold=True)
    total_amount_font = Font(name="Calibri", size=12, bold=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='medium'))
    top_bottom_border = Border(
        top=Side(style='medium'), bottom=Side(style='medium')
    )

    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 2     # spacer
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 5     # spacer
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 2     # spacer

    # ============================================================
    # ROW 2: INVOICE title
    # ============================================================
    ws.merge_cells('B2:F2')
    ws['B2'] = "INVOICE"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # Underline below title
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}3'].border = Border(top=Side(style='medium'))

    # ============================================================
    # ROW 4-5: Invoice No & Date (right side)
    # ============================================================
    ws['E4'] = "Invoice No.:"
    ws['E4'].font = label_font
    ws['E4'].alignment = right_align
    ws['F4'] = INVOICE_NO
    ws['F4'].font = data_font

    ws['E5'] = "Invoice Date:"
    ws['E5'].font = label_font
    ws['E5'].alignment = right_align
    ws['F5'] = INVOICE_DATE.strftime("%B %d, %Y")
    ws['F5'].font = data_font

    # ============================================================
    # ROW 7-12: FROM (Issuer) - LFU
    # ============================================================
    ws['B7'] = "FROM:"
    ws['B7'].font = section_font

    ws['B8'] = ISSUER["name"]
    ws['B8'].font = Font(name="Calibri", size=11, bold=True)

    ws['B9'] = ISSUER["address"]
    ws['B9'].font = data_font

    ws['B10'] = f"Tel: {ISSUER['tel']}"
    ws['B10'].font = data_font

    ws['B11'] = f"Email: {ISSUER['email']}"
    ws['B11'].font = data_font

    # ============================================================
    # ROW 7-12: BILL TO (Orbiters) - right side
    # ============================================================
    ws['E7'] = "BILL TO:"
    ws['E7'].font = section_font

    ws['E8'] = BILL_TO["name"]
    ws['E8'].font = Font(name="Calibri", size=11, bold=True)

    ws['E9'] = BILL_TO["address"]
    ws['E9'].font = data_font
    ws.merge_cells('E9:F9')
    ws['E9'].alignment = left_wrap

    ws['E10'] = f"Tel: {BILL_TO['tel']}"
    ws['E10'].font = data_font

    ws['E11'] = f"Email: {BILL_TO['email']}"
    ws['E11'].font = data_font

    # ============================================================
    # ROW 13: Separator
    # ============================================================
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}13'].border = Border(bottom=Side(style='thin', color="CCCCCC"))

    # ============================================================
    # ROW 15-16: Table header
    # ============================================================
    table_headers = [
        ('B', "No."),
        ('C', "Description"),
        ('E', "Currency"),
        ('F', "Amount"),
    ]
    for col, text in table_headers:
        cell = ws[f'{col}15']
        cell.value = text
        cell.font = label_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center
    # Fill D15 too (part of the table)
    ws['D15'].fill = header_fill
    ws['D15'].border = thin_border

    # ============================================================
    # ROW 16: Data row
    # ============================================================
    ws['B16'] = 1
    ws['B16'].font = data_font
    ws['B16'].alignment = center
    ws['B16'].border = thin_border

    ws.merge_cells('C16:D16')
    ws['C16'] = DESCRIPTION
    ws['C16'].font = data_font
    ws['C16'].alignment = left_wrap
    ws['C16'].border = thin_border
    ws['D16'].border = thin_border

    ws['E16'] = CURRENCY
    ws['E16'].font = data_font
    ws['E16'].alignment = center
    ws['E16'].border = thin_border

    ws['F16'] = AMOUNT
    ws['F16'].font = amount_font
    ws['F16'].alignment = right_align
    ws['F16'].number_format = '#,##0.00'
    ws['F16'].border = thin_border

    # ============================================================
    # ROW 17: Empty row (for potential additional items)
    # ============================================================
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}17'].border = thin_border

    # ============================================================
    # ROW 18: Total
    # ============================================================
    ws.merge_cells('B18:D18')
    ws['B18'] = "TOTAL"
    ws['B18'].font = total_amount_font
    ws['B18'].alignment = Alignment(horizontal='right', vertical='center')
    ws['B18'].fill = total_fill
    ws['B18'].border = thin_border
    ws['C18'].border = thin_border
    ws['D18'].border = thin_border

    ws['E18'] = CURRENCY
    ws['E18'].font = total_amount_font
    ws['E18'].alignment = center
    ws['E18'].fill = total_fill
    ws['E18'].border = thin_border

    ws['F18'] = AMOUNT
    ws['F18'].font = Font(name="Calibri", size=14, bold=True)
    ws['F18'].alignment = right_align
    ws['F18'].number_format = '#,##0.00'
    ws['F18'].fill = total_fill
    ws['F18'].border = thin_border

    # ============================================================
    # ROW 20: Amount in words
    # ============================================================
    ws['B20'] = "Amount in Words:"
    ws['B20'].font = label_font
    ws.merge_cells('C20:F20')
    ws['C20'] = "US DOLLARS SIXTY THOUSAND ONLY"
    ws['C20'].font = data_font

    # ============================================================
    # ROW 22-23: Payment Terms
    # ============================================================
    ws['B22'] = "Payment Terms:"
    ws['B22'].font = label_font
    ws['C22'] = "T/T (Telegraphic Transfer)"
    ws['C22'].font = data_font

    # ============================================================
    # ROW 25-30: Bank Information
    # ============================================================
    ws['B25'] = "BANK INFORMATION"
    ws['B25'].font = section_font
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}25'].border = Border(bottom=Side(style='thin'))

    bank_rows = [
        ("Remittee's Name:", BANK_INFO["remittee"]),
        ("Bank Name:", BANK_INFO["bank"]),
        ("SWIFT Code:", BANK_INFO["swift"]),
        ("Account No.:", BANK_INFO["account"]),
        ("Wire Routing No.:", BANK_INFO["routing"]),
    ]
    for i, (label, value) in enumerate(bank_rows):
        r = 27 + i
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = label_font
        ws[f'C{r}'] = value
        ws[f'C{r}'].font = data_font

    # ============================================================
    # ROW 33-36: Signature
    # ============================================================
    ws['E33'] = "Authorized Signature"
    ws['E33'].font = label_font
    ws['E33'].alignment = center

    # Signature line
    ws.merge_cells('E35:F35')
    ws['E35'].border = Border(bottom=Side(style='thin'))
    ws['F35'].border = Border(bottom=Side(style='thin'))

    ws.merge_cells('E36:F36')
    ws['E36'] = f"{ISSUER['signer']} / {ISSUER['name']}"
    ws['E36'].font = data_font
    ws['E36'].alignment = center

    ws.merge_cells('E37:F37')
    ws['E37'] = ISSUER["signer_title"]
    ws['E37'].font = small_font
    ws['E37'].alignment = center

    # ============================================================
    # Print settings
    # ============================================================
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Set print area
    ws.print_area = 'A1:G40'

    # ---- Save ----
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


if __name__ == "__main__":
    print("=" * 60)
    print("Consulting Fee Invoice Generator")
    print("=" * 60)
    print(f"  Issuer:    {ISSUER['name']}")
    print(f"  Bill To:   {BILL_TO['name']}")
    print(f"  Amount:    {CURRENCY} {AMOUNT:,.2f}")
    print(f"  Date:      {INVOICE_DATE.strftime('%Y-%m-%d')}")
    print(f"  Invoice #: {INVOICE_NO}")
    print()

    path = create_invoice()

    print(f"DONE! Invoice saved to:")
    print(f"  {path}")
    print()
