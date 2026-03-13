"""
Generate Export Invoice (bank submission) - product shipment.
LFU (LittlefingerUSA Inc.) issues to Fleeters Inc.

Usage:
  python tools/generate_export_invoice.py
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "Data Storage", "export")

# ---- Configuration ----
INVOICE_NO = "USA-FLEETERS-26-02-1"
INVOICE_DATE = datetime(2026, 2, 23)
CURRENCY = "USD"

# Exporter / Issuer (LFU - receives payment)
EXPORTER = {
    "name": "LittlefingerUSA Inc.",
    "address": "A-320, 3, Godeung-ro, Sujeong-gu, Seongnam-si, Gyeonggi-do, Republic of Korea, 13105",
    "tel": "82-10-4803-4704",
    "email": "grosmimi.usa@gmail.com",
    "signer": "Chung Hae Jung",
    "signer_title": "President / LittlefingerUSA Inc.",
}

# Importer / Buyer (Fleeters - payer)
IMPORTER = {
    "name": "Fleeters Inc.",
    "address": "30 N Gould St. Ste 32663, Sheridan, WY 82801, USA",
    "tel": "82-10-4803-4704",
    "email": "grosmimi.usa@gmail.com",
}

# Trade terms
DESTINATION = "USA"
PAYMENT = "100% by T/T advance before shipment, General Transaction"
PRICE_TERMS = "EXW"
ORIGIN = "Republic of Korea"

# Line items
ITEMS = [
    {"no": 1, "description": "Grosmimi Cups", "amount": 55000.00},
    {"no": 2, "description": "Grosmimi Replacement Parts", "amount": 15000.00},
]

# Bank Info (LFU's receiving bank account)
BANK_INFO = {
    "company": "ORBITERS Co.,Ltd.",
    "remittee": "LITTLEFINGERUSA",
    "bank": "KEB Hana Bank",
    "swift": "KOEXKRSE",
    "account": "630-010399-748",
}

OUTPUT_FILE = os.path.join(OUTPUT_DIR, "2026-02_Grosmimi_EXW_LFU_FLT_USA.xlsx")


def create_invoice():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commercial Invoice"

    # ---- Styles ----
    title_font = Font(name="Calibri", size=14, bold=True)
    section_font = Font(name="Calibri", size=11, bold=True)
    label_font = Font(name="Calibri", size=9, bold=True)
    data_font = Font(name="Calibri", size=9)
    small_font = Font(name="Calibri", size=8)
    amount_font = Font(name="Calibri", size=10, bold=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # ---- Column widths (CIPL style) ----
    col_widths = {
        'A': 2,
        'B': 14,
        'C': 14,
        'D': 14,
        'E': 14,
        'F': 14,
        'G': 14,
        'H': 14,
        'I': 14,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ============================================================
    # ROW 2: Title
    # ============================================================
    ws.merge_cells('B2:I2')
    ws['B2'] = "COMMERCIAL INVOICE"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center')

    # ============================================================
    # ROW 4-8: Exporter info (left) + Invoice info (right)
    # ============================================================
    ws['B4'] = "Shipper / Exporter"
    ws['B4'].font = label_font

    ws['B5'] = EXPORTER["name"]
    ws['B5'].font = Font(name="Calibri", size=10, bold=True)

    ws.merge_cells('B6:E6')
    ws['B6'] = EXPORTER["address"]
    ws['B6'].font = data_font
    ws['B6'].alignment = left_wrap

    ws['B7'] = f"Email: {EXPORTER['email']}"
    ws['B7'].font = data_font

    ws['B8'] = f"Tel: {EXPORTER['tel']}"
    ws['B8'].font = data_font

    # Right side - destination, date, invoice no
    ws['G4'] = "Destination"
    ws['G4'].font = label_font
    ws['H4'] = DESTINATION
    ws['H4'].font = data_font

    ws['G7'] = "Date"
    ws['G7'].font = label_font
    ws['H7'] = INVOICE_DATE
    ws['H7'].font = data_font
    ws['H7'].number_format = 'YYYY-MM-DD'

    ws['G8'] = "Invoice No."
    ws['G8'].font = label_font
    ws['H8'] = INVOICE_NO
    ws['H8'].font = data_font

    # ============================================================
    # ROW 9-13: Importer info
    # ============================================================
    ws['B9'] = "Consignee / Importer"
    ws['B9'].font = label_font

    ws['B10'] = IMPORTER["name"]
    ws['B10'].font = Font(name="Calibri", size=10, bold=True)

    ws.merge_cells('B11:E11')
    ws['B11'] = IMPORTER["address"]
    ws['B11'].font = data_font
    ws['B11'].alignment = left_wrap

    ws['B12'] = f"Email: {IMPORTER['email']}"
    ws['B12'].font = data_font

    ws['B13'] = f"Tel: {IMPORTER['tel']}"
    ws['B13'].font = data_font

    # ============================================================
    # ROW 15-19: Trade terms
    # ============================================================
    terms = [
        ("Destination", DESTINATION),
        ("Payment", PAYMENT),
        ("Price Terms", PRICE_TERMS),
        ("Country of Origin", ORIGIN),
    ]
    for i, (label, value) in enumerate(terms):
        r = 15 + i
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = label_font
        ws[f'D{r}'] = value
        ws[f'D{r}'].font = data_font

    # Separator
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}20'].border = Border(bottom=Side(style='thin', color="CCCCCC"))

    # ============================================================
    # ROW 21: Table header
    # ============================================================
    table_headers = [
        ('B', "No."),
        ('C', "Description"),
        ('G', "Currency"),
        ('H', "Unit Price"),
        ('I', "Amount (USD)"),
    ]
    for col, text in table_headers:
        cell = ws[f'{col}21']
        cell.value = text
        cell.font = label_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # Merge description header across C-F
    ws.merge_cells('C21:F21')
    # Apply border/fill to merged cells
    for c in ['D', 'E', 'F']:
        ws[f'{c}21'].border = thin_border
        ws[f'{c}21'].fill = header_fill

    # ============================================================
    # Data rows
    # ============================================================
    total_amount = 0
    row = 22
    for item in ITEMS:
        ws[f'B{row}'] = item["no"]
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].alignment = center
        ws[f'B{row}'].border = thin_border

        ws.merge_cells(f'C{row}:F{row}')
        ws[f'C{row}'] = item["description"]
        ws[f'C{row}'].font = data_font
        ws[f'C{row}'].alignment = left_wrap
        ws[f'C{row}'].border = thin_border
        for c in ['D', 'E', 'F']:
            ws[f'{c}{row}'].border = thin_border

        ws[f'G{row}'] = CURRENCY
        ws[f'G{row}'].font = data_font
        ws[f'G{row}'].alignment = center
        ws[f'G{row}'].border = thin_border

        ws[f'H{row}'] = ""
        ws[f'H{row}'].border = thin_border

        ws[f'I{row}'] = item["amount"]
        ws[f'I{row}'].font = Font(name="Calibri", size=10)
        ws[f'I{row}'].alignment = right_align
        ws[f'I{row}'].number_format = '#,##0.00'
        ws[f'I{row}'].border = thin_border

        total_amount += item["amount"]
        row += 1

    # Empty spacer row
    for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{c}{row}'].border = thin_border
    row += 1

    # ============================================================
    # Total row
    # ============================================================
    total_row = row
    ws.merge_cells(f'B{total_row}:F{total_row}')
    ws[f'B{total_row}'] = "TOTAL"
    ws[f'B{total_row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'B{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'B{total_row}'].fill = total_fill
    ws[f'B{total_row}'].border = thin_border
    for c in ['C', 'D', 'E', 'F']:
        ws[f'{c}{total_row}'].border = thin_border
        ws[f'{c}{total_row}'].fill = total_fill

    ws[f'G{total_row}'] = CURRENCY
    ws[f'G{total_row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'G{total_row}'].alignment = center
    ws[f'G{total_row}'].fill = total_fill
    ws[f'G{total_row}'].border = thin_border

    ws[f'H{total_row}'] = ""
    ws[f'H{total_row}'].fill = total_fill
    ws[f'H{total_row}'].border = thin_border

    ws[f'I{total_row}'] = total_amount
    ws[f'I{total_row}'].font = Font(name="Calibri", size=12, bold=True)
    ws[f'I{total_row}'].alignment = right_align
    ws[f'I{total_row}'].number_format = '#,##0.00'
    ws[f'I{total_row}'].fill = total_fill
    ws[f'I{total_row}'].border = thin_border

    # ============================================================
    # Amount in words
    # ============================================================
    ws[f'B{total_row + 2}'] = "Amount in Words:"
    ws[f'B{total_row + 2}'].font = label_font
    ws.merge_cells(f'C{total_row + 2}:I{total_row + 2}')
    ws[f'C{total_row + 2}'] = "US DOLLARS SEVENTY THOUSAND ONLY"
    ws[f'C{total_row + 2}'].font = data_font

    # ============================================================
    # Bank Information
    # ============================================================
    bank_start = total_row + 4
    ws[f'B{bank_start}'] = "Bank Information"
    ws[f'B{bank_start}'].font = section_font
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}{bank_start}'].border = Border(bottom=Side(style='thin'))

    bank_rows = [
        ("Remittee's Name:", BANK_INFO["remittee"]),
        ("Bank Name:", BANK_INFO["bank"]),
        ("SWIFT Code:", BANK_INFO["swift"]),
        ("Account No.:", BANK_INFO["account"]),
    ]
    for i, (label, value) in enumerate(bank_rows):
        r = bank_start + 2 + i
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = label_font
        ws[f'C{r}'] = value
        ws[f'C{r}'].font = data_font

    # ============================================================
    # Signature
    # ============================================================
    sig_row = bank_start + 8
    ws[f'G{sig_row}'] = "Authorized Signature"
    ws[f'G{sig_row}'].font = label_font
    ws[f'G{sig_row}'].alignment = center

    ws.merge_cells(f'G{sig_row + 2}:I{sig_row + 2}')
    ws[f'G{sig_row + 2}'].border = Border(bottom=Side(style='thin'))
    ws[f'H{sig_row + 2}'].border = Border(bottom=Side(style='thin'))
    ws[f'I{sig_row + 2}'].border = Border(bottom=Side(style='thin'))

    ws.merge_cells(f'G{sig_row + 3}:I{sig_row + 3}')
    ws[f'G{sig_row + 3}'] = EXPORTER["signer_title"]
    ws[f'G{sig_row + 3}'].font = data_font
    ws[f'G{sig_row + 3}'].alignment = center

    # ============================================================
    # Print settings
    # ============================================================
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # ---- Save ----
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


if __name__ == "__main__":
    total = sum(item["amount"] for item in ITEMS)
    print("=" * 60)
    print("Export Invoice Generator (Grosmimi)")
    print("=" * 60)
    print(f"  Exporter:  {EXPORTER['name']}")
    print(f"  Importer:  {IMPORTER['name']}")
    print(f"  Amount:    {CURRENCY} {total:,.2f}")
    print(f"  Terms:     {PRICE_TERMS}, {PAYMENT}")
    print(f"  Date:      {INVOICE_DATE.strftime('%Y-%m-%d')}")
    print(f"  Invoice #: {INVOICE_NO}")
    print()

    path = create_invoice()

    print(f"DONE! Invoice saved to:")
    print(f"  {path}")
    print()
