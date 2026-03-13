"""
gorgias_cs_template_builder.py

Fetches all closed Gorgias tickets from the past N months, classifies each
conversation with Claude Haiku, then synthesizes canonical CS reply templates
per complaint category using Claude Sonnet.

Output:
  - Google Sheets (CS Template Library)
  - .tmp/cs_templates_YYYY-MM-DD.xlsx (local backup)

Usage:
    python tools/gorgias_cs_template_builder.py [--months 6] [--min-messages 2] [--dry-run]

Prerequisites:
    - .env: GORGIAS_DOMAIN, GORGIAS_EMAIL, GORGIAS_API_KEY, ANTHROPIC_API_KEY
    - .env: GOOGLE_SERVICE_ACCOUNT_PATH=credentials/google_service_account.json
    - .env: GORGIAS_CS_SHEET_ID=1Y1hrdqGZxe3KPlA0rT38B3Tc_q84sP4PumAq25j6zkg
    - Service Account JSON shared as editor on the target spreadsheet
"""

import os
import sys
import json
import argparse
import time
import re
from datetime import datetime, timezone, timedelta
from collections import defaultdict

import requests
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment
import anthropic

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

load_dotenv()

DOMAIN        = os.getenv("GORGIAS_DOMAIN")
EMAIL         = os.getenv("GORGIAS_EMAIL")
API_KEY       = os.getenv("GORGIAS_API_KEY")
ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY")
SA_PATH       = os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH", "credentials/google_service_account.json")
SHEET_ID      = os.getenv("GORGIAS_CS_SHEET_ID", "1Y1hrdqGZxe3KPlA0rT38B3Tc_q84sP4PumAq25j6zkg")

for var, val in [("GORGIAS_DOMAIN", DOMAIN), ("GORGIAS_EMAIL", EMAIL), ("GORGIAS_API_KEY", API_KEY)]:
    if not val:
        raise ValueError(f"{var} not found in .env")

BASE_URL = f"https://{DOMAIN}.gorgias.com/api"
AUTH     = HTTPBasicAuth(EMAIL, API_KEY)
HEADERS  = {"Accept": "application/json"}
AI       = anthropic.Anthropic(api_key=ANTHROPIC_KEY) if ANTHROPIC_KEY else None

# Output columns matching CS_Template_Library_Readable format
SHEET_HEADERS = [
    "pattern_id",
    "problem_category",
    "resolution_category",
    "macro_name",
    "고객 최초문의 요약(영어)",
    "권장 첫 답변(영어)",
    "체크리스트(내부)",
    "금지/주의 표현",
    "업데이트 로그",
]

# ---------------------------------------------------------------------------
# API helper
# ---------------------------------------------------------------------------

def api_get(endpoint: str, params: dict = None) -> dict:
    url = f"{BASE_URL}/{endpoint.lstrip('/')}"
    resp = requests.get(url, auth=AUTH, headers=HEADERS, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# Step 1: Fetch closed tickets from last N months
# ---------------------------------------------------------------------------

def fetch_closed_tickets(months: int) -> list[dict]:
    """
    Fetch all tickets via cursor pagination, then filter client-side:
    - status == "closed"
    - created_datetime >= cutoff (months ago)

    Note: Gorgias API does not support combining status + order_by in one call,
    so we fetch all and filter in Python (same pattern as gorgias_analyze.py).
    """
    cutoff = datetime.now(timezone.utc) - timedelta(days=months * 30)
    cutoff_str = cutoff.strftime("%Y-%m-%dT%H:%M:%SZ")

    all_raw = []
    cursor = None
    page = 1
    stop = False

    print(f"Step 1: Fetching tickets (will filter closed + since {cutoff_str[:10]} client-side)...")

    while not stop:
        params = {"limit": 100, "order_by": "created_datetime:desc"}
        if cursor:
            params["cursor"] = cursor

        data = api_get("tickets", params)
        batch = data.get("data", [])

        for ticket in batch:
            created = ticket.get("created_datetime") or ""
            try:
                dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                if dt < cutoff:
                    # Tickets are ordered newest-first; once we hit cutoff, stop
                    stop = True
                    break
            except Exception:
                pass
            all_raw.append(ticket)

        print(f"  Page {page}: {len(batch)} tickets fetched (pre-filter total: {len(all_raw)})")

        meta = data.get("meta") or {}
        next_cursor = meta.get("next_cursor")
        if not batch or not next_cursor or stop:
            break
        cursor = next_cursor
        page += 1

    # Client-side filter: closed only
    closed = [t for t in all_raw if t.get("status") == "closed"]
    print(f"  Closed tickets in date range: {len(closed)} / {len(all_raw)} total\n")
    return closed


# ---------------------------------------------------------------------------
# Step 2: Fetch messages for a ticket
# ---------------------------------------------------------------------------

def fetch_messages(ticket_id: int) -> list[dict]:
    all_msgs = []
    cursor = None

    while True:
        params = {"limit": 30}
        if cursor:
            params["cursor"] = cursor

        data = api_get(f"tickets/{ticket_id}/messages", params)
        batch = data.get("data", [])
        all_msgs.extend(batch)

        meta = data.get("meta") or {}
        next_cursor = meta.get("next_cursor")
        if not batch or not next_cursor:
            break
        cursor = next_cursor

    return all_msgs


def format_conversation(messages: list[dict]) -> str:
    """Format messages as readable conversation text."""
    lines = []
    for msg in messages:
        body = (msg.get("body_text") or "").strip()
        if not body:
            continue
        # Clean excessive whitespace
        body = re.sub(r'\n{3,}', '\n\n', body)
        body = body[:1200]  # Limit per message

        from_agent = msg.get("from_agent", False)
        sender_obj = msg.get("sender") or {}
        sender_type = sender_obj.get("type") or ""

        if from_agent or sender_type == "agent":
            label = "[Agent]"
        else:
            label = "[Customer]"

        lines.append(f"{label}\n{body}")

    return "\n\n---\n\n".join(lines)


# ---------------------------------------------------------------------------
# Step 3: Classify each ticket with Claude Haiku
# ---------------------------------------------------------------------------

CLASSIFY_SYSTEM = """You are a customer service analyst. Given a support conversation,
extract classification in valid JSON only. No extra text."""

CLASSIFY_PROMPT = """\
Conversation:
{conversation}

Return ONLY this JSON (no markdown, no explanation):
{{
  "problem_category": "<one of: Shipping, Returns & Refunds, Product Issue, Order Change/Cancel, Payment & Billing, Account & Login, Wholesale/B2B, Generic, Other>",
  "resolution_category": "<one of: Provide information, Process refund, Process replacement, Update order, Apologize, Escalate, Close>",
  "customer_complaint": "<1-2 sentences: what the customer complained about>",
  "agent_resolution": "<1-2 sentences: how the agent resolved it>"
}}"""


def classify_ticket(ticket_id: int, subject: str, conversation: str) -> dict | None:
    if not AI:
        return None

    truncated = conversation[:2500]
    if len(conversation) > 2500:
        truncated += "\n...[truncated]"

    prompt = CLASSIFY_PROMPT.format(conversation=truncated)

    try:
        resp = AI.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            system=CLASSIFY_SYSTEM,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        # Strip markdown code fences if present
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)
        result = json.loads(raw)
        return result
    except json.JSONDecodeError as e:
        print(f"    WARN: JSON parse error for ticket #{ticket_id}: {e}")
        return None
    except Exception as e:
        print(f"    WARN: Claude classify failed for ticket #{ticket_id}: {e}")
        return None


# ---------------------------------------------------------------------------
# Step 4: Synthesize template per group with Claude Sonnet
# ---------------------------------------------------------------------------

SYNTHESIZE_SYSTEM = """You are an expert customer service manager writing a
response template library. Create professional, empathetic templates based on
real conversation examples. Use {{ticket.customer.firstname}} for the customer name."""

SYNTHESIZE_PROMPT = """\
Problem Category: {problem_category}
Resolution Category: {resolution_category}

Here are {count} real examples of conversations in this category:

{examples}

Based on these real examples, create a reusable CS template entry. Return ONLY this JSON:
{{
  "macro_name": "<short descriptive name, e.g. 'Shipping: Delayed Order'>",
  "customer_summary": "<2-3 sentences describing when a customer sends this type of inquiry>",
  "response_template": "<full response template in English, professional & empathetic tone, use {{{{ticket.customer.firstname}}}} for name, include relevant placeholder notes in [brackets]>",
  "checklist": "<internal checklist as bullet points with • prefix, e.g.:\\n• Confirm: Order number\\n• Check: Tracking status>",
  "prohibited": "<expressions to avoid, as bullet points with • prefix, or 'None' if none>"
}}"""


def synthesize_template(problem_category: str, resolution_category: str, examples: list[dict]) -> dict | None:
    if not AI:
        return None

    # Pick up to 5 representative examples
    sample = examples[:5]
    example_texts = []
    for i, ex in enumerate(sample, 1):
        complaint = ex.get("customer_complaint", "")
        resolution = ex.get("agent_resolution", "")
        conv_snippet = ex.get("conversation", "")[:600]
        example_texts.append(
            f"--- Example {i} ---\n"
            f"Customer complaint: {complaint}\n"
            f"How agent resolved: {resolution}\n"
            f"Conversation snippet:\n{conv_snippet}"
        )

    prompt = SYNTHESIZE_PROMPT.format(
        problem_category=problem_category,
        resolution_category=resolution_category,
        count=len(sample),
        examples="\n\n".join(example_texts),
    )

    try:
        resp = AI.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1500,
            system=SYNTHESIZE_SYSTEM,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)

        # First attempt: direct parse
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            pass

        # Second attempt: extract fields with regex (handles apostrophes/newlines in values)
        result = {}
        field_patterns = {
            "macro_name":        r'"macro_name"\s*:\s*"((?:[^"\\]|\\.)*)"',
            "customer_summary":  r'"customer_summary"\s*:\s*"((?:[^"\\]|\\.)*)"',
            "response_template": r'"response_template"\s*:\s*"((?:[^"\\]|\\.)*)"',
            "checklist":         r'"checklist"\s*:\s*"((?:[^"\\]|\\.)*)"',
            "prohibited":        r'"prohibited"\s*:\s*"((?:[^"\\]|\\.)*)"',
        }
        for field, pattern in field_patterns.items():
            m = re.search(pattern, raw, re.DOTALL)
            if m:
                result[field] = m.group(1).replace('\\n', '\n').replace('\\"', '"')

        if len(result) >= 3:  # At least 3 fields recovered
            return result

        print(f"    WARN: JSON parse and regex both failed for ({problem_category}/{resolution_category})")
        return None
    except Exception as e:
        print(f"    WARN: Sonnet synthesis failed ({problem_category}/{resolution_category}): {e}")
        return None


# ---------------------------------------------------------------------------
# Step 5: Write to Google Sheets
# ---------------------------------------------------------------------------

def write_to_sheets(records: list[dict], sheet_id: str, sa_path: str) -> bool:
    """Write template records to Google Sheets. Returns True on success."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("ERROR: gspread or google-auth not installed. Run: pip install gspread")
        return False

    if not os.path.exists(sa_path):
        print(f"ERROR: Service account JSON not found at {sa_path}")
        print("  See workflow doc for setup instructions.")
        return False

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    try:
        creds = Credentials.from_service_account_file(sa_path, scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
    except Exception as e:
        print(f"ERROR: Could not open spreadsheet: {e}")
        return False

    # Find or create sheet
    sheet_name = "CS_Template_Library"
    try:
        ws = sh.worksheet(sheet_name)
        ws.clear()
        print(f"  Cleared existing sheet: {sheet_name}")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_name, rows=200, cols=len(SHEET_HEADERS))
        print(f"  Created new sheet: {sheet_name}")

    # Write header
    ws.append_row(SHEET_HEADERS)

    # Write data rows
    for rec in records:
        row = [
            rec.get("pattern_id", ""),
            rec.get("problem_category", ""),
            rec.get("resolution_category", ""),
            rec.get("macro_name", ""),
            rec.get("customer_summary", ""),
            rec.get("response_template", ""),
            rec.get("checklist", ""),
            rec.get("prohibited", ""),
            rec.get("update_log", ""),
        ]
        ws.append_row(row)
        time.sleep(0.3)  # Avoid Sheets API rate limit

    # Bold header row
    try:
        ws.format("A1:I1", {"textFormat": {"bold": True}})
    except Exception:
        pass  # Formatting is optional

    print(f"  Wrote {len(records)} templates to Google Sheets.")
    return True


# ---------------------------------------------------------------------------
# Step 6: Write local Excel backup
# ---------------------------------------------------------------------------

def write_excel_backup(records: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CS Templates"

    ws.append(SHEET_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [10, 20, 22, 30, 40, 55, 35, 30, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    for rec in records:
        row = [
            rec.get("pattern_id", ""),
            rec.get("problem_category", ""),
            rec.get("resolution_category", ""),
            rec.get("macro_name", ""),
            rec.get("customer_summary", ""),
            rec.get("response_template", ""),
            rec.get("checklist", ""),
            rec.get("prohibited", ""),
            rec.get("update_log", ""),
        ]
        ws.append(row)
        for col_idx in range(1, len(SHEET_HEADERS) + 1):
            ws.cell(row=ws.max_row, column=col_idx).alignment = wrap
        ws.row_dimensions[ws.max_row].height = 100

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"  Local backup saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Build CS Template Library from Gorgias conversations")
    parser.add_argument("--months",       type=int, default=6,  help="Months of history to fetch (default: 6)")
    parser.add_argument("--min-messages", type=int, default=2,  dest="min_messages",
                        help="Min message count per ticket (default: 2)")
    parser.add_argument("--dry-run",      action="store_true",  help="Skip Google Sheets write, save Excel only")
    args = parser.parse_args()

    from output_utils import get_output_path
    output_path = get_output_path("cs", "cs_templates")

    # ------------------------------------------------------------------
    # Step 1: Fetch tickets
    # ------------------------------------------------------------------
    all_tickets = fetch_closed_tickets(args.months)

    # Filter by min message count
    candidates = [t for t in all_tickets if (t.get("messages_count") or 0) >= args.min_messages]
    print(f"Tickets with >= {args.min_messages} messages: {len(candidates)} / {len(all_tickets)}\n")

    if not candidates:
        print("No qualifying tickets found. Exiting.")
        return

    # ------------------------------------------------------------------
    # Step 2 & 3: Fetch messages + classify each ticket
    # ------------------------------------------------------------------
    print(f"Step 2-3: Fetching conversations and classifying {len(candidates)} tickets...")
    classified = []

    for i, ticket in enumerate(candidates, start=1):
        tid = ticket.get("id")
        subject = ticket.get("subject") or ""
        msg_count = ticket.get("messages_count") or 0
        print(f"  [{i}/{len(candidates)}] Ticket #{tid} ({msg_count} msgs) — {subject[:50]}")

        # Fetch messages
        try:
            messages = fetch_messages(tid)
        except Exception as e:
            print(f"    SKIP: message fetch error: {e}")
            continue

        if not messages:
            continue

        conversation = format_conversation(messages)
        if not conversation.strip():
            continue

        # Classify
        result = classify_ticket(tid, subject, conversation)
        if result:
            result["ticket_id"] = tid
            result["subject"] = subject
            result["conversation"] = conversation
            classified.append(result)
            print(f"    → {result.get('problem_category')} / {result.get('resolution_category')}")
        else:
            print(f"    → classification skipped")

        # Small delay to avoid Gorgias rate limit
        time.sleep(0.1)

    print(f"\nClassified {len(classified)} tickets.\n")

    if not classified:
        print("No tickets classified. Check ANTHROPIC_API_KEY. Exiting.")
        return

    # ------------------------------------------------------------------
    # Step 4: Group by category and synthesize templates
    # ------------------------------------------------------------------
    print("Step 4: Grouping and synthesizing templates...")

    groups = defaultdict(list)
    for item in classified:
        key = (item.get("problem_category", "Other"), item.get("resolution_category", "Generic"))
        groups[key].append(item)

    print(f"  Found {len(groups)} unique category groups:")
    for key, items in sorted(groups.items(), key=lambda x: -len(x[1])):
        print(f"    {key[0]} / {key[1]}: {len(items)} tickets")

    templates = []
    today_log = datetime.now().strftime("%Y-%m-%d")

    for idx, ((problem_cat, resolution_cat), examples) in enumerate(
        sorted(groups.items(), key=lambda x: -len(x[1])), start=1
    ):
        print(f"\n  Synthesizing template {idx}/{len(groups)}: {problem_cat} / {resolution_cat} ({len(examples)} examples)")

        result = synthesize_template(problem_cat, resolution_cat, examples)
        if not result:
            # Fallback: create minimal template from classification data
            result = {
                "macro_name": f"{problem_cat}: {resolution_cat}",
                "customer_summary": examples[0].get("customer_complaint", ""),
                "response_template": "(Template generation failed — please write manually)",
                "checklist": "• Confirm: Order number\n• Review: Customer history",
                "prohibited": "None",
            }

        pattern_id = f"CS-{idx:03d}"
        templates.append({
            "pattern_id": pattern_id,
            "problem_category": problem_cat,
            "resolution_category": resolution_cat,
            "macro_name": result.get("macro_name", f"{problem_cat}: {resolution_cat}"),
            "customer_summary": result.get("customer_summary", ""),
            "response_template": result.get("response_template", ""),
            "checklist": result.get("checklist", ""),
            "prohibited": result.get("prohibited", "None"),
            "update_log": f"v1 {today_log}: auto-generated from {len(examples)} tickets",
        })

    print(f"\nGenerated {len(templates)} templates.\n")

    # ------------------------------------------------------------------
    # Step 5: Write to Google Sheets (unless dry-run)
    # ------------------------------------------------------------------
    if args.dry_run:
        print("Step 5: [DRY RUN] Skipping Google Sheets write.")
    else:
        print("Step 5: Writing to Google Sheets...")
        success = write_to_sheets(templates, SHEET_ID, SA_PATH)
        if success:
            print(f"  Google Sheets URL: https://docs.google.com/spreadsheets/d/{SHEET_ID}")
        else:
            print("  Google Sheets write failed — check credentials. Local backup will still be saved.")

    # ------------------------------------------------------------------
    # Step 6: Local Excel backup
    # ------------------------------------------------------------------
    print(f"\nStep 6: Saving local Excel backup...")
    write_excel_backup(templates, output_path)

    print(f"\nDone! {len(templates)} CS templates generated.")
    if args.dry_run:
        print(f"Review results in: {output_path}")
        print("When satisfied, re-run without --dry-run to push to Google Sheets.")


if __name__ == "__main__":
    main()
