"""
Polar Ads 카테고리 분류 툴
- Data Storage/Polar Ads 카테고리 분류_Raw.xlsx 의 I열(Brand), J열(Category) 채우기
- 매칭 우선순위 (행 단위):
  1. Facebook Ads 랜딩 URL → Shopify 제품 핸들 → SKU → Product Variant 파일
  2. 컬렉션 URL → 브랜드 추론
  3. 캠페인명 키워드 → 브랜드 추론
  4. Adset명 키워드 → 브랜드 추론  ← NEW
  5. Ad명 키워드 → 브랜드 추론      ← NEW
  6. 매칭 불가 → Non-classified / Non-classified
"""

import os
import re
import json
import time
import urllib.request
import urllib.parse
from collections import defaultdict, Counter
from dotenv import load_dotenv

load_dotenv()

# ── 경로 설정 ─────────────────────────────────────────────────
from output_utils import get_output_path, get_latest_file, get_intermediate_path, DATA_STORAGE
POLAR_ADS_FILE      = get_latest_file("polar", "ads_classification") or os.path.join(DATA_STORAGE, "_archive", "Polar Ads 카테고리 분류_Raw.xlsx")
POLAR_ADS_SHEET     = "ads Copy (1)"
PRODUCT_VARIANT_FILE = os.path.join(DATA_STORAGE, "polar", "Product_Variant_Reference.xlsx")
PRODUCT_VARIANT_SHEET = "Custom Report"
FB_ADS_FILE         = get_intermediate_path("", "facebook_ads.xlsx")

SHOPIFY_SHOP  = os.getenv("SHOPIFY_SHOP", "mytoddie.myshopify.com")
SHOPIFY_TOKEN = os.getenv("SHOPIFY_ACCESS_TOKEN")
API_VERSION   = "2024-01"

# ── 브랜드명 키워드 매핑 ───────────────────────────────────────
# 우선순위: 앞에 있을수록 높음
BRAND_KEYWORDS = [
    # Grosmimi: 이름 또는 약자 또는 제품군(ppsu/stainless는 Grosmimi 전용)
    (["grosmimi", " gm ", "| gm |", "gm |", "| gm", "gm_",
      "grosm", "grossmimi"],                                         "Grosmimi"),
    (["beemymagic", "beemy"],                                        "Beemymagic"),
    (["naeiae"],                                                      "Naeiae"),
    (["babyrabbit", "baby rabbit", "baby-rabbit", "babyrabbit"],     "BabyRabbit"),
    (["alpremio"],                                                    "Alpremio"),
    # CHA&MOM: 약자 cm 포함 (단, commemoi보다 낮은 우선순위)
    (["cha&mom", "cha mom", "cha&mom", "chamom", " cm ", "| cm |",
      "cm |", "| cm", "_cm_", "cm_"],                               "CHA&MOM"),
    (["hattung"],                                                     "Hattung"),
    (["commemoi", "comme moi", "commemo"],                            "Comme Moi"),
    (["bamboobebe", "bamboo bebe"],                                   "BambooBebe"),
    (["ride & go", "ride&go", "ride-go", "ridego", "ride go"],       "RIDE & GO"),
    (["nature love mere"],                                            "Nature Love Mere"),
    (["easy shower", "shower stand"],                                 "Easy Shower"),
    # 제품군으로 브랜드 추론 (겹치지 않는 제품군 활용)
    (["ppsu"],                                                        "Grosmimi"),  # PPSU = Grosmimi 전용
    (["stainless steel", "stainless straw", "sls cup"],              "Grosmimi"),  # Grosmimi 전용
    (["straw cup", "straw"],                                         "Grosmimi"),  # Straw = Grosmimi 전용
    (["rice snack", "pop rice"],                                      "Naeiae"),    # Naeiae 전용
    (["skincare", "lotion", "hair wash", "body wash"],                "CHA&MOM"),  # CHA&MOM 전용
]

# 브랜드별 기본 카테고리 (특정 제품 파악 불가 시 사용)
BRAND_DEFAULT_CATEGORY = {
    "Alpremio":        "Alpremio Seat",
    "BabyRabbit":      "Non-classified",
    "BambooBebe":      "BambooBebe",
    "Beemymagic":      "Beemymagic",
    "CHA&MOM":         "Non-classified",
    "CX":              "CX",
    "Comme Moi":       "Non-classified",
    "Easy Shower":     "Easy Shower",
    "Grosmimi":        "Non-classified",
    "Hattung":         "Hattung",
    "Naeiae":          "Naeiae Pop Rice Snack",
    "Nature Love Mere":"Nature Love Mere",
    "RIDE & GO":       "RIDE & GO",
}

# 카테고리 키워드 (브랜드 파악 후 캠페인/URL 텍스트에서 카테고리 추론)
CATEGORY_KEYWORDS = [
    (["ppsu straw cup", "ppsu"],                      "PPSU Straw Cup"),
    (["knotted flip", "knotted straw"],               "KNOTTED Flip Top Straw Cup"),
    (["flip top straw", "flip top"],                  "Flip Top Straw Cup"),
    (["stainless straw", "stainless steel straw", "sls cup", "sls straw"],
                                                      "Stainless Steel Straw Cup"),
    (["stainless steel food tray", "stainless food tray"],
                                                      "Stainless Steel Food Tray"),
    (["stainless tumbler", "stainless steel tumbler"],"Stainless Steel Tumbler"),
    (["ppsu tumbler"],                                "PPSU Tumbler"),
    (["heart tray", "beemeal heart"],                 "Beemeal Heart Tray"),
    (["beemeal bowl clover"],                         "Beemeal Bowl Clover"),
    (["beemeal bowl double", "beemeal double"],       "Beemeal Bowl Double-handle"),
    (["beemeal bowl heart"],                          "Beemeal Bowl Heart"),
    (["modular dish"],                                "Heart Tray Modular Dish"),
    (["pop rice", "rice snack"],                      "Naeiae Pop Rice Snack"),
    (["book stand"],                                  "Comme Moi Book Stand"),
    (["drawing board"],                               "Comme Moi Double-Sided Drawing Board"),
    (["kids stool", "kids' stool"],                   "Comme Moi Kids' Stool"),
    (["hair & body", "hair body wash"],               "CHA&MOM Hair & Body Wash"),
    (["intense cream", "phyto cream"],                "CHA&MOM Phyto Seline Intense Cream"),
    (["moisture lotion", "lotion"],                   "CHA&MOM Phyto Seline Moisture Lotion"),
    (["ppsu baby bottle", "baby bottle"],             "PPSU Baby Bottle"),
    (["straw brush", "brush"],                        "Straw Brush"),
    (["teether"],                                     "Teether"),
    (["lunch bag"],                                   "Lunch Bag"),
    (["loungewear"],                                  "Loungewear"),
    (["underwear"],                                   "Underwear"),
    (["socks"],                                       "Socks"),
    (["tumbler"],                                     "Stainless Steel Tumbler"),   # fallback tumbler (대부분 SS Tumbler 제휴)
    # straw fallback은 infer_category_from_text 함수에서 compound 로직으로 처리
    (["tray"],                                        "Stainless Steel Food Tray"), # fallback tray
]

# 항상 Non-classified로 분류할 패턴
NON_CLASSIFIED_PATTERNS = ["dsh |", "| dsh", "dsh tof", "dsh mof"]

# General로 분류할 URL 패턴
GENERAL_URL_PATTERNS = ["/discount/", "target.com", "amazon.com", "play.google.com",
                         "walmart.com", "lovecarekorea"]


# ── Shopify API ────────────────────────────────────────────────
_shopify_cache = {}

def get_product_skus_by_handle(handle):
    """Shopify 제품 핸들 → SKU 목록"""
    if handle in _shopify_cache:
        return _shopify_cache[handle]

    if not SHOPIFY_TOKEN:
        return []

    url = (f"https://{SHOPIFY_SHOP}/admin/api/{API_VERSION}/products.json"
           f"?handle={urllib.parse.quote(handle)}&fields=variants&limit=1")
    req = urllib.request.Request(url, headers={"X-Shopify-Access-Token": SHOPIFY_TOKEN})
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            products = data.get("products", [])
            skus = []
            for p in products:
                for v in p.get("variants", []):
                    if v.get("sku"):
                        skus.append(v["sku"])
            _shopify_cache[handle] = skus
            time.sleep(0.5)  # rate limit
            return skus
    except Exception as e:
        print(f"    [API ERROR] handle={handle}: {e}")
        return []


# ── 참조 데이터 로딩 ──────────────────────────────────────────
def load_sku_map():
    """SKU → (Brand, Category) 딕셔너리 반환"""
    import openpyxl
    wb = openpyxl.load_workbook(PRODUCT_VARIANT_FILE, read_only=True, data_only=True)
    ws = wb[PRODUCT_VARIANT_SHEET]
    sku_map = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        sku, cat, brand = row[1], row[4], row[5]
        if sku and brand and str(sku) not in sku_map:
            sku_map[str(sku)] = (brand, cat or "Non-classified")
    print(f"  SKU 매핑 로드: {len(sku_map)}개")
    return sku_map


def load_campaign_urls():
    """Campaign ID → landing URL 딕셔너리 반환"""
    import openpyxl
    wb = openpyxl.load_workbook(FB_ADS_FILE, read_only=True, data_only=True)
    ws = wb.active
    url_map = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cid = str(row[0]) if row[0] else None
        url = row[6] or ""
        if cid:
            url_map[cid] = url
    print(f"  Facebook Ads 캠페인 로드: {len(url_map)}개")
    return url_map


# ── 분류 로직 ─────────────────────────────────────────────────
def infer_brand_from_text(text):
    """텍스트(캠페인명, URL 등)에서 브랜드 추론"""
    t = text.lower()
    for keywords, brand in BRAND_KEYWORDS:
        for kw in keywords:
            if kw in t:
                return brand
    return None


def infer_category_from_text(text):
    """텍스트에서 카테고리 추론"""
    t = text.lower()
    for keywords, cat in CATEGORY_KEYWORDS:
        for kw in keywords:
            if kw in t:
                return cat
    # Straw compound check: stainless/steel + straw → SS Straw Cup, straw만 → Non-classified
    if "straw" in t:
        if "stainless" in t or "steel" in t:
            return "Stainless Steel Straw Cup"
        return "Non-classified"  # PPSU vs Stainless 구분 불가
    return None


def extract_handle_from_url(url):
    """URL에서 Shopify 제품 핸들 추출"""
    m = re.search(r"/products/([^/?#&,\s]+)", url)
    return m.group(1) if m else None


def extract_collection_from_url(url):
    """URL에서 컬렉션명 추출"""
    m = re.search(r"/collections/([^/?#&,\s]+)", url)
    return m.group(1) if m else None


def classify_by_url(url, sku_map):
    """단일 URL로 (brand, category) 분류. 실패 시 None 반환"""
    if not url:
        return None

    url = url.strip()

    # General 패턴 URL → Non-classified
    for pat in GENERAL_URL_PATTERNS:
        if pat in url:
            return ("Non-classified", "Non-classified")

    # 제품 URL → Shopify API → SKU 매칭
    if "/products/" in url:
        handle = extract_handle_from_url(url)
        if handle:
            skus = get_product_skus_by_handle(handle)
            for sku in skus:
                if sku in sku_map:
                    return sku_map[sku]
            # SKU 매칭 실패 → 핸들 텍스트에서 추론
            brand = infer_brand_from_text(handle)
            if brand:
                cat = infer_category_from_text(handle) or BRAND_DEFAULT_CATEGORY.get(brand, "Non-classified")
                return (brand, cat)

    # 컬렉션 URL → 브랜드 추론
    if "/collections/" in url:
        col = extract_collection_from_url(url)
        if col and col != "all":
            brand = infer_brand_from_text(col)
            if brand:
                cat = infer_category_from_text(col) or BRAND_DEFAULT_CATEGORY.get(brand, "Non-classified")
                return (brand, cat)

    return None


def classify_by_text(text):
    """텍스트에서 (brand, category) 추론. 실패 시 None"""
    if not text:
        return None
    brand = infer_brand_from_text(str(text))
    if brand:
        cat = infer_category_from_text(str(text)) or BRAND_DEFAULT_CATEGORY.get(brand, "General")
        return (brand, cat)
    return None


def classify_campaign(campaign_name, campaign_id, url_map, sku_map):
    """
    캠페인 → (Brand, Category) 결정
    우선순위: 랜딩 URL → 캠페인명 키워드 → Non-classified
    """
    name = str(campaign_name or "").strip()
    name_lower = name.lower()

    # ─ Step 1: 랜딩 URL 기반 분류 ─
    landing_urls_raw = url_map.get(str(campaign_id), "")
    if landing_urls_raw:
        for url in landing_urls_raw.split(","):
            result = classify_by_url(url.strip(), sku_map)
            if result and result[0] not in ("Non-classified",):
                return result

    # ─ Step 2: 캠페인명 키워드 기반 분류 ─
    # zezebaebae는 스토어명, 단독으로는 브랜드 아님
    if "zezebaebae" in name_lower:
        # zezebaebae 외에 다른 브랜드 키워드가 있으면 사용
        brand = infer_brand_from_text(name.replace("zezebaebae", ""))
        if brand:
            cat = infer_category_from_text(name) or BRAND_DEFAULT_CATEGORY.get(brand, "Non-classified")
            return (brand, cat)
        return ("Non-classified", "Non-classified")

    brand = infer_brand_from_text(name)
    if brand:
        cat = infer_category_from_text(name) or BRAND_DEFAULT_CATEGORY.get(brand, "Non-classified")
        return (brand, cat)

    # ─ Step 3: 랜딩 URL 있지만 브랜드 못 찾은 경우 → Non-classified ─
    if landing_urls_raw:
        for url in landing_urls_raw.split(","):
            result = classify_by_url(url.strip(), sku_map)
            if result:
                return result
        return ("Non-classified", "Non-classified")

    # ─ Step 4: URL 없고 이름도 불명 ─
    return ("Non-classified", "Non-classified")


def classify_row(camp_name, camp_id, adset_name, ad_name, url_map, sku_map):
    """
    행 단위 분류: Ad명 텍스트 우선 → Adset명 → 캠페인(URL+이름) 순
    - Ad명에 구체적 브랜드 키워드가 있으면 캠페인 URL 분류보다 우선 적용
    - 예: "easy shower support handle" → Easy Shower (캠페인이 Grosmimi PPSU여도 무시)
    - 예: "commemoi adjustable book stand" → Comme Moi / Book Stand
    """
    # 1순위: Ad명 텍스트에서 브랜드 추론
    result = classify_by_text(ad_name)
    if result and result[0] not in ("Non-classified",):
        return result

    # 2순위: Adset명 텍스트에서 브랜드 추론
    result = classify_by_text(adset_name)
    if result and result[0] not in ("Non-classified",):
        return result

    # 3순위: 캠페인 (랜딩 URL → SKU 매핑 → 캠페인명 키워드)
    return classify_campaign(camp_name, camp_id, url_map, sku_map)


# ── 메인 ─────────────────────────────────────────────────────
def main():
    import openpyxl

    print("=== Polar Ads 카테고리 분류 시작 ===\n")

    # 참조 데이터 로드
    sku_map = load_sku_map()
    url_map = load_campaign_urls()

    # Polar Ads Raw 파일 열기 (쓰기 가능)
    print(f"\n  Polar Ads Raw 파일 열기...")
    wb = openpyxl.load_workbook(POLAR_ADS_FILE)
    ws = wb[POLAR_ADS_SHEET]

    # 캠페인별 분류 (중복 API 호출 방지)
    campaign_cache = {}  # campaign_id → (brand, category)

    total = 0
    updated = 0
    summary = defaultdict(int)

    print(f"\n  행 처리 중...\n")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        camp_name  = row[0].value                                    # A: Campaign
        camp_id    = str(row[1].value) if row[1].value else ""       # B: Campaign Id
        adset_name = row[2].value                                    # C: Adset
        ad_name    = row[4].value                                    # E: Ad
        total += 1

        # 행 단위 분류 (Ad명까지 활용)
        brand, cat = classify_row(camp_name, camp_id, adset_name, ad_name, url_map, sku_map)

        prev_brand = row[8].value
        row[8].value = brand   # I열
        row[9].value = cat     # J열
        updated += 1
        summary[f"{brand} / {cat}"] += 1

        # 변경된 행만 출력
        if brand != prev_brand:
            label = f"{str(ad_name or '')[:30]:30s} (camp: {str(camp_name or '')[:25]})"
            print(f"  [{row_idx:4d}] {label} → {brand} / {cat}")

    ads_output = get_output_path("polar", "ads_classification")
    wb.save(ads_output)
    print(f"\n{'='*60}")
    print(f"완료! 총 {total}행 중 {updated}행 업데이트 → {ads_output}")
    print(f"\n=== 분류 결과 요약 ===")
    for label, count in sorted(summary.items(), key=lambda x: -x[1]):
        print(f"  {label}: {count}행")
    print(f"\n저장 완료: {POLAR_ADS_FILE}")


if __name__ == "__main__":
    main()
