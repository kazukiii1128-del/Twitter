"""
WAT Tool: Generate professional consulting-grade Excel report from scraped jobs JSON.
Input:  .tmp/jobs_raw.json
Output: .tmp/remote_support_jobs_YYYYMMDD.xlsx

Usage:
    python tools/generate_excel.py
    python tools/generate_excel.py --input .tmp/jobs_raw.json --output .tmp/report.xlsx
"""

import json
import argparse
import re
import logging
from pathlib import Path
from datetime import datetime
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Color palette (consulting neutral) ───────────────────────────────────────
NAVY       = "1F3864"
SLATE_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
OFF_WHITE  = "F5F7FA"
GOLD       = "C9A42A"
MID_GRAY   = "6C757D"
DARK_GRAY  = "343A40"
LINK_BLUE  = "1155CC"

# ── Borders ───────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_bottom():
    return Border(bottom=Side(style="medium", color=SLATE_BLUE))

# ── Column definitions for Job Listings sheet ─────────────────────────────────
# (header, field_key, col_width, wrap_text)
COLUMNS = [
    ("#",            None,                 5,  False),
    ("Job Title",    "job_title",         34,  False),
    ("Company",      "company",           22,  False),
    ("Location",     "location",          18,  False),
    ("Salary",       "salary",            24,  False),
    ("Experience",   "experience_level",  14,  False),
    ("Type",         "employment_type",   12,  False),
    ("Posted",       "date_posted",       14,  False),
    ("Skills",       "skills",            30,  True),
    ("Description",  "description_snippet", 52, True),
    ("Category",     "category",          14,  False),
    ("Apply",        "job_url",           10,  False),
]


# ── Data loading ──────────────────────────────────────────────────────────────

def load_jobs(input_path: Path) -> tuple[list[dict], str]:
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["jobs"], data.get("scraped_at", datetime.now().isoformat())


# ── Statistics ────────────────────────────────────────────────────────────────

def parse_salary_midpoint(salary_str: str) -> float | None:
    if not salary_str:
        return None
    nums = re.findall(r"\d[\d,]*", salary_str.replace(",", ""))
    try:
        if len(nums) >= 2:
            return (float(nums[0]) + float(nums[1])) / 2
        elif len(nums) == 1:
            return float(nums[0])
    except ValueError:
        pass
    return None


def compute_stats(jobs: list[dict]) -> dict:
    companies = [j.get("company", "").strip() for j in jobs if j.get("company")]
    locations = [j.get("location", "").strip() for j in jobs if j.get("location")]

    all_skills: list[str] = []
    for j in jobs:
        raw = j.get("skills", [])
        if isinstance(raw, list):
            all_skills.extend(s.strip().lower() for s in raw if s)
        elif isinstance(raw, str) and raw:
            all_skills.extend(s.strip().lower() for s in raw.split(",") if s.strip())

    midpoints = [m for j in jobs if (m := parse_salary_midpoint(j.get("salary", ""))) is not None]
    avg_salary = round(sum(midpoints) / len(midpoints)) if midpoints else None

    return {
        "total_jobs":       len(jobs),
        "unique_companies": len(set(companies)),
        "unique_locations": len(set(locations)),
        "jobs_with_salary": len(midpoints),
        "avg_salary":       avg_salary,
        "top_companies":    Counter(companies).most_common(10),
        "top_skills":       Counter(all_skills).most_common(15),
        "top_locations":    Counter(locations).most_common(10),
    }


# ── Style helpers ─────────────────────────────────────────────────────────────

def hdr(cell, bg=NAVY, fg=WHITE, size=11, bold=True, halign="center"):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border    = thin_border()


def body(cell, bg=WHITE, size=10, halign="left", valign="top", wrap=False):
    cell.font      = Font(name="Calibri", size=size, color=DARK_GRAY)
    cell.fill      = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border    = thin_border()


# ── Dashboard sheet ───────────────────────────────────────────────────────────

def _breakdown_table(ws, row: int, col: int, title: str, headers: list[str], rows: list):
    """Draw a titled two-column breakdown table starting at (row, col)."""
    # Title bar
    for dc in range(2):
        c = ws.cell(row=row, column=col + dc)
        c.fill = PatternFill(fill_type="solid", fgColor=SLATE_BLUE)
        c.border = thin_border()
    title_cell = ws.cell(row=row, column=col)
    title_cell.value = title
    title_cell.font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

    # Column headers
    for dc, h in enumerate(headers):
        c = ws.cell(row=row + 1, column=col + dc, value=h)
        hdr(c, bg=LIGHT_BLUE, fg=NAVY, size=9)

    # Data rows
    for ri, (name, count) in enumerate(rows):
        r = row + 2 + ri
        bg = OFF_WHITE if ri % 2 == 0 else WHITE
        name_c  = ws.cell(row=r, column=col,     value=name)
        count_c = ws.cell(row=r, column=col + 1, value=count)
        body(name_c,  bg=bg, size=9)
        body(count_c, bg=bg, size=9, halign="center")


def build_dashboard(wb: openpyxl.Workbook, stats: dict, scraped_at: str) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY

    # ── Background wash ───────────────────────────────────────────────────
    bg_fill = PatternFill(fill_type="solid", fgColor=OFF_WHITE)
    for r in range(1, 50):
        for c in range(1, 20):
            ws.cell(row=r, column=c).fill = bg_fill

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells("B2:K2")
    t = ws["B2"]
    t.value     = "Remote Support Jobs  —  Market Overview"
    t.font      = Font(name="Calibri", bold=True, size=20, color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 40

    ws.merge_cells("B3:K3")
    s = ws["B3"]
    s.value     = (
        f"Source: dailyremote.com    "
        f"Filter: Full-time · Maternity benefits    "
        f"Scraped: {scraped_at[:10]}"
    )
    s.font      = Font(name="Calibri", size=9, color=MID_GRAY, italic=True)
    s.alignment = Alignment(horizontal="left")
    ws.row_dimensions[3].height = 16

    # Divider under subtitle
    for c in range(2, 12):
        ws.cell(row=4, column=c).border = Border(
            bottom=Side(style="medium", color=NAVY)
        )
    ws.row_dimensions[4].height = 4

    # ── KPI blocks (row 6-8) ──────────────────────────────────────────────
    kpis = [
        ("Total Jobs",    stats["total_jobs"]),
        ("Companies",     stats["unique_companies"]),
        ("Locations",     stats["unique_locations"]),
        ("With Salary",   stats["jobs_with_salary"]),
        ("Avg. Salary",   f"${stats['avg_salary']:,}" if stats["avg_salary"] else "N/A"),
    ]
    kpi_cols = [2, 4, 6, 8, 10]  # B D F H J

    for col, (label, value) in zip(kpi_cols, kpis):
        # KPI box background
        for r in range(6, 9):
            ws.cell(row=r, column=col).fill = PatternFill(fill_type="solid", fgColor=WHITE)
            ws.cell(row=r, column=col).border = thin_border()
            ws.cell(row=r, column=col + 1).fill = PatternFill(fill_type="solid", fgColor=WHITE)

        lbl = ws.cell(row=6, column=col, value=label)
        lbl.font      = Font(name="Calibri", bold=True, size=8, color=MID_GRAY)
        lbl.alignment = Alignment(horizontal="center", vertical="bottom")
        ws.merge_cells(
            start_row=6, start_column=col, end_row=6, end_column=col + 1
        )

        val = ws.cell(row=7, column=col, value=value)
        val.font      = Font(name="Calibri", bold=True, size=22, color=GOLD)
        val.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(
            start_row=7, start_column=col, end_row=8, end_column=col + 1
        )

    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 34
    ws.row_dimensions[8].height = 8

    # Spacer row
    ws.row_dimensions[9].height = 8

    # ── Breakdown tables (row 10+) ────────────────────────────────────────
    _breakdown_table(ws, 10, 2,  "Top 10 Companies", ["Company", "Listings"], stats["top_companies"])
    _breakdown_table(ws, 10, 5,  "Top 15 Skills",    ["Skill",   "Count"],    stats["top_skills"])
    _breakdown_table(ws, 10, 8,  "Top 10 Locations", ["Location","Jobs"],     stats["top_locations"])

    # ── Column widths ─────────────────────────────────────────────────────
    widths = {"A": 2, "B": 24, "C": 12, "D": 3, "E": 22, "F": 10,
              "G": 3, "H": 22, "I": 10, "J": 3, "K": 22, "L": 10}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ── Job Listings sheet ────────────────────────────────────────────────────────

def build_listings(wb: openpyxl.Workbook, jobs: list[dict]) -> None:
    ws = wb.create_sheet("Job Listings")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = SLATE_BLUE

    # Header row
    for ci, (header, _, width, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=header)
        hdr(c)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28

    # Freeze first row
    ws.freeze_panes = "A2"

    # AutoFilter
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    # Data rows
    for ri, job in enumerate(jobs):
        row_idx = ri + 2
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE

        for ci, (_, field, _, wrap) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=ci)

            if field is None:
                # Row number
                cell.value = ri + 1
                body(cell, bg=bg, halign="center")

            elif field == "skills":
                raw = job.get("skills", [])
                if isinstance(raw, list):
                    cell.value = ", ".join(raw)
                else:
                    cell.value = str(raw) if raw else ""
                body(cell, bg=bg, wrap=wrap)

            elif field == "job_url":
                url = job.get("job_url", "")
                if url and url.startswith("http"):
                    cell.value     = "View Job"
                    cell.hyperlink = url
                    cell.font      = Font(
                        name="Calibri", size=10,
                        color=LINK_BLUE, underline="single", bold=True
                    )
                    cell.fill      = PatternFill(fill_type="solid", fgColor=bg)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    cell.border    = thin_border()
                else:
                    cell.value = url
                    body(cell, bg=bg, halign="center")

            else:
                cell.value = job.get(field, "") or ""
                body(cell, bg=bg, wrap=wrap)

        # Row height: taller for rows with wrapped content
        ws.row_dimensions[row_idx].height = 55

    # Format as Excel Table (enables sort arrows in addition to AutoFilter)
    last_data_row = len(jobs) + 1
    table_ref = f"A1:{last_col}{last_data_row}"
    tbl = Table(displayName="JobListings", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,   # We handle fill manually
        showColumnStripes=False,
    )
    ws.add_table(tbl)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Excel report from scraped jobs JSON")
    parser.add_argument("--input",  default=".tmp/jobs_raw.json")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(
            f"Input not found: {input_path}\nRun 'python tools/scrape_jobs.py' first."
        )

    ts = datetime.now().strftime("%Y%m%d")
    output_path = Path(args.output) if args.output else Path(f".tmp/remote_support_jobs_{ts}.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Loading jobs from {input_path}...")
    jobs, scraped_at = load_jobs(input_path)
    logger.info(f"Loaded {len(jobs)} jobs")

    stats = compute_stats(jobs)
    logger.info(
        f"Stats: {stats['total_jobs']} jobs | "
        f"{stats['unique_companies']} companies | "
        f"{stats['unique_locations']} locations"
    )

    wb = openpyxl.Workbook()
    # Remove default blank sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    logger.info("Building Dashboard sheet...")
    build_dashboard(wb, stats, scraped_at)

    logger.info("Building Job Listings sheet...")
    build_listings(wb, jobs)

    wb.active = wb["Dashboard"]
    wb.save(output_path)
    logger.info(f"Saved: {output_path}")

    print(f"\n{'='*55}")
    print(f"  Excel report created: {output_path}")
    print(f"  Total jobs:     {stats['total_jobs']}")
    print(f"  Companies:      {stats['unique_companies']}")
    print(f"  Locations:      {stats['unique_locations']}")
    if stats["avg_salary"]:
        print(f"  Avg. salary:    ${stats['avg_salary']:,}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
